"""Turn the design into DataTable CSVs for Unreal.

    python tools/generate_datatables.py            # write the CSVs
    python tools/generate_datatables.py --check    # verify, change nothing

Output goes to game/Data/. Each CSV matches a USTRUCT declared in
game/Source/Cataclysm/Data/CataclysmDataRows.h, and an Unreal automation test
loads every CSV through its struct so a mismatch fails the build rather than
being discovered when something silently reads nothing.

THERE ARE TWO SOURCES, NOT ONE. Most tables come from the design workbook,
docs/All_Things_Cataclysm.xlsx, which is what a person edits. The two enemy
tables come instead from sim/cataclysm_sim/enemy_stats.py, because that is where
the enemy stat block is designed and where its self-checks live. `TABLES` holds
the workbook builders and `MODEL_TABLES` the Python-model ones; everything after
that treats them alike.

GENERATED FROM THE MODEL RATHER THAN COPIED OUT OF IT, deliberately. This project
already keeps one copy of a model by hand -- sim/cataclysm_sim/scoring.py against
the DungeonSimulator repository -- and CLAUDE.md records that it drifted silently
twice, which is why sim/verify_scoring_port.py had to be written. A second
hand-maintained copy would be the same mistake. Continuous integration runs
`--check`, so editing enemy_stats.py without regenerating fails the pull request.

WHY THIS IS NOT A LOOP OVER SHEETS

Not every workbook table comes from a sheet that is already a plain table with
one entity per row. Several sheets need reshaping, so each table has its own
handler:

  Enchantments      two independent tables side by side in one sheet, positives
                    in columns A-D and negatives in F-I
  Enemy Modifiers   a matrix, one column per Cataclysm, each cell holding
                    "Name: Description"
  Buffs, Debuffs    single column, no header row, each cell "Name: Description"
  DoTs              same shape as Buffs

Row names are derived and must be stable, because a DataTable row name is the key
everything else references. Renaming one silently breaks every reference to it.

THE ITEM BASES AND AFFIXES SHEETS HAVE A SECOND READER. `sim/cataclysm_sim/
affixes.py` holds the same pool, because that is where the design rules are
enforced and where tuning happens. The workbook is authoritative and is what a
person edits; `tools/tests/test_affix_sheets_match_the_model.py` compares the two
so they cannot drift apart silently, which a copy in this project has done before.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import pathlib
import re
import sys

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
OUTPUT_DIR = REPO_ROOT / "game" / "Data"

#: The simulation package. It is not installed, so the enemy builders below reach
#: it by path, the same way tools/tests/test_brute_matches_the_model.py does.
SIM_ROOT = REPO_ROOT / "sim"

CATACLYSM_TYPES = ["Demonic", "Death", "War", "Pestilence", "Famine",
                   "Celestial", "Chaos", "Void", "Generic"]


class DataError(Exception):
    """A problem in the source data that must stop generation."""


# --------------------------------------------------------------------------
# helpers

def clean(value) -> str:
    """Normalise a cell to a trimmed string, collapsing internal whitespace."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def number(value, field: str, row: int) -> float:
    text = clean(value)
    if not text:
        raise DataError(f"row {row}: {field} is empty")
    try:
        return float(text)
    except ValueError as error:
        raise DataError(f"row {row}: {field} is {text!r}, not a number") from error


def row_name(*parts: str) -> str:
    """A stable, readable DataTable row name.

    Only letters, digits and underscores survive, because a row name is used as
    an FName and referenced from data.
    """
    joined = "_".join(p for p in parts if p)
    name = re.sub(r"[^A-Za-z0-9]+", "_", joined).strip("_")
    if not name:
        raise DataError(f"could not build a row name from {parts!r}")
    return name


def split_named(text: str) -> tuple[str, str]:
    """Split a "Name: Description" cell. Returns ("", text) if there is no colon."""
    if ":" in text:
        name, _, description = text.partition(":")
        return name.strip(), description.strip()
    return "", text


def unique(rows: list[dict], table: str) -> list[dict]:
    """Make row names unique by suffixing duplicates, and report how many."""
    seen: dict[str, int] = {}
    for row in rows:
        base = row["Name"]
        if base in seen:
            seen[base] += 1
            row["Name"] = f"{base}_{seen[base]}"
        else:
            seen[base] = 0
    return rows


# --------------------------------------------------------------------------
# per-sheet handlers

def dungeon_modifiers(book) -> list[dict]:
    out = []
    for index, raw in enumerate(book["Dungeon Modifiers"].iter_rows(values_only=True), 1):
        if index == 1 or not raw or not clean(raw[0]):
            continue
        cataclysm, name, weight, description = (clean(raw[0]), clean(raw[1]),
                                                raw[2], clean(raw[3]))
        if not name:
            raise DataError(f"Dungeon Modifiers row {index}: modifier name is empty")
        if cataclysm not in CATACLYSM_TYPES:
            raise DataError(f"Dungeon Modifiers row {index}: "
                            f"{cataclysm!r} is not a Cataclysm type")
        out.append({"Name": row_name(cataclysm, name), "CataclysmType": cataclysm,
                    "ModifierName": name,
                    "Weight": number(weight, "Weight", index),
                    "Description": description})
    return unique(out, "Dungeon Modifiers")


#: Riders any shape may carry, and what each one means.
#:
#: Eight of the sixteen designed Demonic skills leave a burning patch of ground
#: behind whatever else they do, so the ground zone is a RIDER on a shape rather
#: than a shape of its own. Issue #37's own table hints at this: it lists
#: "persistent ground zone" against "used by most of the above" instead of
#: naming skills, where every other entry names skills.
SHAPE_RIDERS = {
    "Burn": "1 if the skill sets what it hits alight, 0 or absent if not",

    # WHOM A BUFF BURNS, WHICH `Burn` ABOVE DOES NOT SAY. A self buff
    # stating `Burn` may mean "the pinned enemies in my radius, once a
    # second" (the Spear's Held Fast) or "whatever strikes me in melee"
    # (the Greataxe's Burning Wrath). One key cannot mean both, so the
    # second is its own. Added 2026-09-02.
    "BurnsAttackers": "1 if a running buff sets alight whatever strikes its "
                      "holder in melee, 0 or absent if not",
    "GroundRadius": "metres of burning ground left behind, 0 for none",
    "GroundDuration": "seconds that ground burns",
    "GroundPercent": "percent of the skill's damage that ground deals per "
                     "second. Decided 2026-08-14 on issue #361: standing in it "
                     "for its whole GroundDuration costs one hit of the skill "
                     "that left it, so this is 100 divided by GroundDuration. "
                     "It keeps burning ground area denial rather than a second "
                     "damage source, and stops a longer patch being "
                     "automatically a bigger one",
    "GroundHitsAllies": "1 if that ground burns everything standing in it "
                        "whatever side it is on, including whatever left it. "
                        "Absent or 0 means it burns only the caster's enemies, "
                        "which is what every player skill wants. The Hellhound's "
                        "fire trail is the only thing that sets it",
    "StunSeconds": "how long the target is stunned, if the skill stuns. The "
                   "anti-stun-lock rule in section VI still applies: it is the "
                   "duration, not a promise that a stun lands",
    "Knockback": "metres the skill pushes what it hits away from the caster, 0 "
                 "or absent for none. A RIDER RATHER THAN A STRIKE PARAMETER, "
                 "decided 2026-08-15 on issue #626: displacement is not specific "
                 "to one kind of skill. A strike, a leap, a charge and an enemy "
                 "slam can all shove, and while this was a parameter of Strike "
                 "alone, Shockwave Leap knocked back in its prose and could not "
                 "say so in its data. That is the same argument that made a "
                 "burning patch of ground a rider",
    "FinalHitPercent": "percent of weapon damage a closing hit deals, if any",
    "HealthCostPercent": "percent of current health one use costs, if any",
    "Effect": "the named status effect this applies, from the Buffs, Debuffs "
              "or DoTs sheet. More than one may be named, comma separated",

    # --- Added 2026-09-01 with the Demonic verb rewrite ---------------------
    #
    # THE BUCKET IS IN THE NAME, NOT IN A VALUE. `MoreDamagePer` multiplies
    # separately and `IncreasedDamagePer` joins the additive sum, which is the
    # distinction section VI draws and the single most important thing about a
    # damage number in this project. A shared key with a bucket beside it would
    # have made the two one typo apart.
    "MoreDamagePer": "percentage points of MORE damage -- the multiplicative "
                     "bucket -- per unit of ScalingSource",
    "IncreasedDamagePer": "percentage points of INCREASED damage -- the "
                          "additive bucket -- per unit of ScalingSource",

    # A FLAT MULTIPLIER ON A CONDITION RATHER THAN A RATE PER COUNT, which is
    # what separates this from the two keys above. Those need a ScalingSource
    # to say what they count; the condition here is where the attacker stood.
    # Added 2026-09-02 for the Dagger's Emberpierce, whose "40% more damage
    # from behind" had no parameter at all and so did nothing.
    "MoreDamageFromBehind": "percentage points of MORE damage -- the "
                            "multiplicative bucket -- on a blow landed from "
                            "behind the target",
    "DurationPer": "seconds added to the skill's own duration per unit of "
                   "ScalingSource",
    "ScalingSource": "what the three keys above count, one of SCALING_SOURCES",

    # Displacement that a distance cannot say. `Knockback` above stays exactly
    # what issue #626 made it -- metres pushed away from the caster -- because
    # that decision generalised it on purpose and three enemy abilities write
    # it. This rider carries the five verbs a distance does not express: put on
    # the floor, thrown upward, hauled inward, dragged along, or nailed in
    # place. A skill that simply shoves still writes Knockback.
    "ForcedMovement": "what the skill does to a target beyond pushing it, one "
                      "or more of FORCED_MOVEMENTS",
    "ForcedMovementDistance": "metres a pull or drag carries the target",
    "ForcedMovementDuration": "seconds the target is held where it was put",

    # Geometry that outlives the hit. Distinct from the burning ground rider
    # above: that is a damage patch, this changes where a fight can happen.
    "Terrain": "persistent geometry the skill leaves, one of TERRAIN_KINDS",
    "TerrainSize": "metres -- radius for a pit, fissure or thicket, length for "
                   "a wall",
    "TerrainDuration": "seconds the geometry persists",

    # A skill that does not fire, or does not pay out, without its condition.
    "Requires": "a condition the skill needs, one or more of REQUIREMENTS",
    "Immune": "what the caster cannot be subjected to while the skill runs, "
              "one or more of IMMUNITIES. Not immunity to DAMAGE, which is a "
              "different axis and is issue #1162",

    # NOT `Duration`, DELIBERATELY. Anathema's ten seconds is how long its curse
    # sits on an enemy and Butcher's Bill's ten seconds is how long the skill
    # itself runs. They were the same word, and a reader could not tell which
    # one a row meant.
    "EffectDuration": "seconds an applied Effect lasts, never a channel length",
    "EffectMagnitude": "size of the applied Effect, in whatever unit that "
                       "effect is measured in",

    # THE ONE THING AN AURA GIVES RATHER THAN TAKES. Added 2026-09-02 for
    # Conflagration's "allies within it deal 8% increased fire damage", which
    # had no key at all, so half the row lived only in its prose. Issue #1182.
    #
    # `increased` IS THE ADDITIVE BUCKET AND THE NAME SAYS SO, the same way
    # `MoreDamagePer` and `IncreasedDamagePer` below carry their bucket in their
    # names. It joins the sum every gear affix and passive node joins.
    #
    # NOT `IncreasedDamagePer`, WHICH IS A DIFFERENT NUMBER. That is per unit of
    # a ScalingSource and sizes one blow. This is a flat grant to somebody else
    # for as long as they stand inside.
    "AllyIncreasedDamage": "percentage points of increased damage an aura "
                           "grants allies standing in it, scoped to the "
                           "skill's own element",

    # THE OTHER HALF OF THE SAME EVENT `ScalingSource=HitTaken` READS. Added
    # 2026-09-02 for Living Pyre, "returns health equal to 25% of the damage
    # that hit dealt", which had no key. Of what reached health, not of what
    # was sent. It is NOT leech: leech is what a hit gives back to whoever
    # landed it, paid out over three seconds. Issue #1162.
    "HealthFromHitTaken": "percent of a blow taken that is returned to the "
                          "caster as health while the skill runs",

    # THE OTHER WAY A SKILL MAY USE A BLOW TAKEN. Added 2026-09-02 for Martyr's
    # Ember, which holds it as damage to give back rather than turning it into
    # health at once. Three keys because the row states three numbers -- "40% of
    # all damage you take is stored", "capped at 200% weapon damage", and a
    # spend the row does NOT state, whose value is a judgement recorded in
    # docs/DECISIONS.md. Issue #1162.
    "StoresFromHitTaken": "percent of a blow taken that is put into the "
                          "skill's store instead of being paid back",
    "StoreCapPercent": "ceiling on that store, as percent of weapon damage",
    "StoreSpentPerHit": "percent of weapon damage one landed blow takes out "
                        "of the store and adds to what it hit. Once per enemy "
                        "hit, not once per use",

    "OnDeath": "what happens when an affected enemy dies, one of ON_DEATHS",
    "OnDeathRange": "metres the on-death effect reaches for its next target",

    "ConsumeBurn": "1 if the skill spends the target's burn rather than only "
                   "applying it. The Sword's verb",
    "ConsumeRadius": "metres of whatever consumption produces",

    # SPREADING FIRE WITHOUT CONSUMING IT, which is the opposite of the two
    # keys directly above. Consuming takes the burn OUT of a target and
    # spends it; this leaves it burning and lights its neighbours as well.
    # The Wand's Hex of Cinders is the only row that states it, and issue
    # #1146 records why it is not written with GroundRadius instead.
    "SpreadWhen": "what a TARGET must already carry for the skill to spread "
                  "fire from it, one of SPREAD_CONDITIONS. Not a condition "
                  "on the cast, which is Requires",
    "SpreadRadius": "metres the fire spreads from a target that met "
                    "SpreadWhen",

    "MaxDamagePercent": "ceiling on a skill whose damage scales, as percent of "
                        "weapon damage",
    "MinDamagePercent": "floor for a charged skill released immediately, as "
                        "percent of weapon damage",

    "ChargeTime": "seconds of hold before a full release",
    "ChargeBreaksOn": "what cancels a hold and loses the skill, one or more of "
                      "CHARGE_BREAKS. None means nothing can",

    "RefundsCooldown": "which cooldown the skill returns, one of "
                       "REFUND_TARGETS. What triggers it is the row's Trigger "
                       "tag",
    "Untargetable": "1 if the caster cannot be hit while the skill runs",
}

#: Closed value sets for the text riders above.
#:
#: CLOSED FOR THE SAME REASON THE PARAMETER LIST IS. A misspelled value that
#: parsed would be a skill that runs and quietly does nothing, which is the
#: failure `parse_shape_params` exists to prevent one level up.
SCALING_SOURCES = frozenset({
    "Kill", "Burning", "Second", "Meter", "HitTaken", "Consume", "Consumed",
    "Bounce", "Pierced", "Pinned", "HealthMissing",
})
FORCED_MOVEMENTS = frozenset({"Knockdown", "Launch", "Pull", "Drag", "Pin"})
TERRAIN_KINDS = frozenset({"Pit", "Wall", "Fissure", "Thicket"})
REQUIREMENTS = frozenset({"Burning", "Target", "Stationary", "RearHit"})
ON_DEATHS = frozenset({"Leap", "SpreadDebuff", "Release"})

#: What a skill may make its caster immune to while it runs.
#:
#: THE DESIGN'S OWN WORDS, from the table in section VI of
#: `docs/Cataclysm_GDD_v2.md` listing which effects the anti-stun-lock rule
#: covers: Stun, Knockdown, Slow, Displacement and Madness. Pin is added
#: because it is an effect a row could reasonably refuse and the engine already
#: applies it, and CrowdControl names all six at once -- which is what "immunity
#: to all crowd control" says and what saves a row from listing six things to
#: mean one.
#:
#: SEVEN ROWS ACROSS FOUR WEAPONS STATE ONE. Section VI names five of them
#: outright -- Living Pyre, Unstoppable Force, Forge Stance, Bull Rush and
#: Cinder Rush -- and the Greatsword's Unbroken and Inexorable are the other
#: two. None had a parameter until 2026-09-02.
IMMUNITIES = frozenset({
    "Stun", "Knockdown", "Slow", "Displacement", "Pin", "Madness",
    "CrowdControl",
})
CHARGE_BREAKS = frozenset({"Stagger", "Death", "Movement", "None"})
REFUND_TARGETS = frozenset({"Self", "Movement"})
TARGET_MODES = frozenset({"All", "Nearest", "Furthest"})

#: What a target may already carry for a skill to spread fire from it.
#:
#: ONE VALUE, because one row states the key. It is a closed set rather than
#: free text for the reason every other rider here is: a misspelling would
#: otherwise reach the engine, fail to match any branch, and spread nothing
#: while the row read as though it worked.
SPREAD_CONDITIONS = frozenset({"Burning"})

#: Which closed set each text rider draws from. A rider absent from this map is
#: a number.
TEXT_PARAM_VALUES = {
    "ScalingSource": SCALING_SOURCES,
    "ForcedMovement": FORCED_MOVEMENTS,
    "Terrain": TERRAIN_KINDS,
    "SpreadWhen": SPREAD_CONDITIONS,
    "Requires": REQUIREMENTS,
    "Immune": IMMUNITIES,
    "OnDeath": ON_DEATHS,
    "ChargeBreaksOn": CHARGE_BREAKS,
    "RefundsCooldown": REFUND_TARGETS,
    "TargetMode": TARGET_MODES,
}

#: The parameters whose value is a name rather than a number.
#:
#: `Minions` is a comma-separated list of `Type:Count` pairs naming what a skill
#: produces, added with issue #338. It is a list rather than one name because
#: Iron Fortress deploys two ballistae AND three spike traps, which no single
#: key and value can say. Every type it names must exist in the Minion Types
#: sheet, and `validate_minion_references` refuses one that does not.
TEXT_PARAMS = frozenset({"Mode", "Effect", "Minions"}
                        | set(TEXT_PARAM_VALUES))

#: How a `Minions` entry is written: `Imp:1`, or `Ballista:2, SpikeTrap:3`.
MINION_ENTRY = re.compile(r"^\s*([A-Za-z][A-Za-z0-9]*)\s*:\s*(\d+)\s*$")


def parse_minions(text: str, where: str) -> dict[str, int]:
    """Read a `Minions` value into {type name: how many}."""
    out: dict[str, int] = {}
    for piece in text.split(","):
        piece = piece.strip()
        if not piece:
            continue
        match = MINION_ENTRY.match(piece)
        if not match:
            raise DataError(
                f"{where}: minion entry {piece!r} is not Type:Count")
        name, count = match.group(1), int(match.group(2))
        if name in out:
            raise DataError(f"{where}: {name} is listed twice")
        if count < 1:
            raise DataError(
                f"{where}: {name} is produced {count} times, so naming it says "
                "nothing")
        out[name] = count
    if not out:
        raise DataError(f"{where}: Minions is set and names nothing")
    return out

#: The closed list of shapes, and which parameters each one reads.
#:
#: A SHAPE NAMES WHICH TEMPLATE RUNS. It is deliberately NOT read off the Tags
#: column, even though that column already carries Type.Projectile,
#: Type.AOE.PointBlank and the rest. Two reasons, and the second is the one that
#: would have caused a real bug:
#:
#:   The tags do not decide it. Molten Cleave carries Type.AOE.PointBlank,
#:   Type.Strike AND Type.AOE.Persistent, and nothing says which is the primary
#:   behaviour. Infernal Plunge is a leap and carries no tag saying so.
#:
#:   The tags already have a job. UCataclysmStatPipeline::ModifierApplies scopes
#:   every gear increase by the tags of the skill in hand. Dispatching on them
#:   too would mean adding a tag to make a skill's shape work silently changed
#:   which gear modifiers apply to it.
#:
#: Path of Exile draws the same line: its gems.json carries `types` (the internal
#: list, whose stated purpose is deciding which support gems may support a skill)
#: separately from the behaviour the skill's own ActiveSkills.dat id names.
#: A PROJECTILE STATES `Speed` OR `Flight`, NOT BOTH. `Speed` is centimetres per
#: second and describes something travelling flat, which is what all 398 player
#: projectile rows use. `Flight` is seconds in the air and describes a LOB
#: following real projectile motion onto the point it was aimed at, whose speed
#: and arc height both fall out of the time. `SHAPE_PARAMS` in
#: `sim/cataclysm_sim/enemy_abilities.py` is the same table and says the same.
#: Issue #465.
SHAPE_PARAMS = {
    "Strike": {"Radius", "Angle", "MaxTargets", "Duration", "Interval",
               # Buried Fire leaves the greatsword in the ground and the player
               # fights bare-handed until it is pulled free. Only it states this.
               "DisarmsUntilRecalled"},
    # A PROJECTILE MAY NOW REPEAT. Butcher's Bill throws thirty axes over ten
    # seconds, which needs the same Count, Duration and Interval a Summon has
    # always had; before this it could only be written as prose. `TargetMode`
    # says who a repeating projectile picks, and is the same word and the same
    # values as the column of that name in the Minion Types sheet.
    "Projectile": {"Range", "Radius", "Pierce", "Returns", "Speed", "Arc",
                   "Count", "Duration", "Interval", "TargetMode",
                   "ScalesWithAttackSpeed", "Bounces", "SpreadCurses",
                   "CommandStrike",
                   "TetherTargets", "TetherLength", "TetherDuration"},
    # `Interval` JOINED ON 2026-09-02 WITH THE SPEAR'S HELD FAST, whose
    # sentence has a per-second half: "any pinned enemy within 12 meters is
    # set alight again each second it is held". Until then a self buff was
    # the one lasting shape that could not repeat, so that half of the row
    # could not be written down at all. Strike, Projectile, Movement, Summon,
    # Deployable and Aura all already had it.
    "SelfBuff": {"Duration", "Interval", "Radius", "RangeIncrease"},
    # Inexorable walks for three seconds and Everywhere at Once flickers for
    # four, so a movement can now last and repeat. `RearHits` says every blow it
    # lands counts as struck from behind.
    #
    # `MaxTargets` WAS MISSING AND THE ENGINE HAS ALWAYS READ IT. All three
    # movement modes that hit anything hand it straight to the targeting search:
    # a charge caps what it catches along the line, and a leap and a blink cap
    # what they catch at the ends. A row stating it would therefore have worked,
    # and this table refused to let one say so.
    #
    # WHAT MADE IT VISIBLE is the Spear's Nail Down, which impales "the first
    # enemy you reach" where the Whip's Reel takes "every enemy the line
    # crosses". The two are both charges and differ in nothing else, so without
    # this there is no way to write the difference down at all.
    "Movement": {"Mode", "Range", "Radius", "MaxTargets", "Duration", "Interval",
                 "RearHits"},
    # Subjugate takes ONE enemy, which is what MaxTargets says everywhere else.
    "Summon": {"Range", "Radius", "Count", "MaxActive", "Duration", "Interval",
               "Minions", "MaxTargets", "Possess", "FervourReserve",
               "HealthThresholdPercent"},
    # AN EIGHTH SHAPE, added with issue #338. A summon spawns things that walk to
    # the enemy; a deployable places things that stay where they are put. The
    # split is behavioural and the data already carried it as a tag -- Bolt
    # Turret, Ballista and Iron Fortress all have Type.Deployable and the three
    # summons have Type.Summon -- but the Shape column was empty for all three,
    # so every number they state lived only in their prose description where no
    # code could read it.
    #
    # The name matches the existing tag rather than inventing a new word.
    "Deployable": {"Range", "Radius", "Count", "MaxActive", "Duration",
                   "Interval", "Minions", "HealthPercent"},
    "Aura": {"Radius", "Duration", "Interval"},
    "Debuff": {"Range", "Radius", "MaxTargets", "Duration",
               "SpreadWhen", "SpreadRadius"},
}

#: The movement modes, and what each one is.
#:
#: THREE WERE ADDED 2026-09-01 with the Demonic verb rewrite, because three
#: weapons' verbs are movements nothing existing could describe. `Swap`
#: exchanges the caster's position with a minion's, which is the Staff's
#: Vesselstep and the only movement in the game that costs something the player
#: owns. `Recall` returns to a mark left earlier rather than departing, which is
#: what separates the Dagger's Echo from its Ashwalk. `Flicker` repeats,
#: arriving at one enemy after another, and is the only movement that is also an
#: ultimate.
MOVEMENT_MODES = {"Leap", "Charge", "Blink", "Swap", "Recall", "Flicker"}

#: The two shapes a basic attack may take. `docs/Cataclysm_GDD_v2.md` says the
#: basic attack is the weapon's own swing -- a Strike for a melee weapon, a
#: Projectile for a ranged one -- so any other shape in that column is a mistyped
#: cell rather than a design. Issue #524.
BASIC_ATTACK_SHAPES = frozenset({"Strike", "Projectile"})


def parse_shape_params(text: str, shape: str, where: str) -> dict[str, str]:
    """Read a `Key=Value; Key=Value` cell, refusing anything the shape cannot use.

    REFUSING IS THE POINT. A parameter this returned silently as zero looks
    exactly like a parameter nobody wrote, which is how a cooldown of zero went
    unnoticed across 77 skills in issue #155. A radius of zero hits nothing, so
    a misspelled `Radiuss` would produce a skill that runs and does nothing.
    """
    params: dict[str, str] = {}
    if not text:
        return params

    allowed = SHAPE_PARAMS[shape] | set(SHAPE_RIDERS)
    for piece in text.split(";"):
        piece = piece.strip()
        if not piece:
            continue
        if "=" not in piece:
            raise DataError(f"{where}: shape parameter {piece!r} is not Key=Value")
        key, value = (part.strip() for part in piece.split("=", 1))
        if key not in allowed:
            raise DataError(
                f"{where}: {shape} has no parameter {key!r}. It reads "
                f"{sorted(allowed)}.")
        if key in params:
            raise DataError(f"{where}: parameter {key!r} is given twice")
        if not value:
            raise DataError(f"{where}: parameter {key!r} has no value")
        if key == "Mode":
            if value not in MOVEMENT_MODES:
                raise DataError(
                    f"{where}: Mode is {value!r}, not one of "
                    f"{sorted(MOVEMENT_MODES)}")
        elif key in TEXT_PARAM_VALUES:
            # A CLOSED SET, CHECKED HERE. These read as text like Mode does, so
            # nothing below would have caught a misspelling: "ForcedMovement=Nockdown"
            # would have been stored and then matched no branch in the engine's
            # parser, which is a skill that runs and does not shove.
            for one in (v.strip() for v in value.split(",")):
                if not one:
                    continue
                if one not in TEXT_PARAM_VALUES[key]:
                    raise DataError(
                        f"{where}: {key}={one!r} is not one of "
                        f"{sorted(TEXT_PARAM_VALUES[key])}")
            params[key] = value
        elif key in TEXT_PARAMS:
            pass  # checked against the effect list by validate_skill_effects
        else:
            try:
                float(value)
            except ValueError:
                raise DataError(
                    f"{where}: parameter {key!r} is {value!r}, not a number"
                ) from None
        params[key] = value

    return params


def check_basic_attack(shape: str, params: str, arms_the_holder: bool,
                       is_weapon: bool, where: str) -> None:
    """Refuse a basic attack that is missing, misplaced, or not pure weapon damage.

    THE BASIC ATTACK IS THE ANCHOR. `game/Data/SkillSlots.csv` puts it at 100%
    and every other slot is a percentage of it, so a basic attack carrying a
    burn, a stun or a patch of ground would make the anchor deal more than the
    weapon does and every other slot's percentage would then describe something
    else. That is why the riders are refused here rather than merely discouraged.

    A weapon that grants no attack damage grants no basic attack either, which is
    the Shield and was decided on issue #619. It is asked about through
    `arms_the_holder` rather than by name so a second such weapon behaves the
    same way without an edit here.

    BOTH HALVES OF `is_weapon and arms_the_holder` ARE LOAD-BEARING. Flat attack
    damage is not by itself a weapon: the Vambraces glove base grants 12 of it,
    and a glove has no swing to describe.
    """
    if params and not shape:
        raise DataError(f"{where}: has basic attack parameters but no shape")

    if not shape:
        if is_weapon and arms_the_holder:
            raise DataError(
                f"{where}: is a weapon that grants flat attack damage and has no "
                "basic attack shape. Every armed weapon needs one, or a character "
                "holding it has nothing between its cooldowns. Issue #524.")
        return

    if not is_weapon:
        raise DataError(f"{where}: is not a weapon but states a basic attack. "
                        "Only a weapon supplies one.")
    if not arms_the_holder:
        raise DataError(
            f"{where}: grants no flat attack damage but states a basic attack, "
            "which would deal 100% of nothing. Issue #619.")
    if shape not in BASIC_ATTACK_SHAPES:
        raise DataError(
            f"{where}: basic attack shape is {shape!r}. A basic attack is the "
            f"weapon's own swing, so it is one of {sorted(BASIC_ATTACK_SHAPES)}.")

    parsed = parse_shape_params(params, shape, where)
    riders = sorted(set(parsed) - SHAPE_PARAMS[shape])
    if riders:
        raise DataError(
            f"{where}: basic attack carries {riders}. A basic attack is 100% "
            "weapon damage and nothing else, which is what makes it the anchor "
            "every other slot is measured against.")


def weapon_skills(book) -> list[dict]:
    out = []
    for index, raw in enumerate(book["Weapon Skills"].iter_rows(values_only=True), 1):
        if index == 1 or not raw or not clean(raw[0]):
            continue
        weapon, damage, slot = clean(raw[0]), clean(raw[1]), clean(raw[2])
        name = clean(raw[3])
        shape = clean(raw[6]) if len(raw) > 6 else ""
        params = clean(raw[7]) if len(raw) > 7 else ""
        crit = clean(raw[8]) if len(raw) > 8 else ""
        damage_percent = clean(raw[9]) if len(raw) > 9 else ""
        cooldown = clean(raw[10]) if len(raw) > 10 else ""
        mana_cost = clean(raw[11]) if len(raw) > 11 else ""
        where = f"Weapon Skills row {index} ({damage} {weapon} {slot})"

        if shape and shape not in SHAPE_PARAMS:
            raise DataError(f"{where}: shape {shape!r} is not one of "
                            f"{sorted(SHAPE_PARAMS)}")
        if params and not shape:
            raise DataError(f"{where}: has shape parameters but no shape")
        if shape and not name:
            raise DataError(f"{where}: has a shape but no skill name")

        # Parsed and thrown away. The game parses it again from the CSV; this
        # call is here so a bad cell fails generation rather than producing a
        # skill that runs and does nothing.
        if shape:
            parse_shape_params(params, shape, where)

        out.append({"Name": row_name(damage, weapon, slot),
                    "WeaponType": weapon, "DamageType": damage, "Slot": slot,
                    "SkillName": name,
                    "SkillDescription": clean(raw[4]),
                    "Tags": tags_with_slot(clean(raw[5]), slot, where),
                    "Shape": shape,
                    "ShapeParams": params,
                    "CritChancePercent": skill_crit_chance(crit, where),
                    "DamagePercent": skill_number(
                        damage_percent, "damage percentage", where),
                    "Cooldown": skill_number(cooldown, "cooldown", where),
                    "ManaCost": skill_number(mana_cost, "mana cost", where)})
    return unique(out, "Weapon Skills")


#: What the DamagePercent, Cooldown and ManaCost columns hold for a skill that
#: states none of its own, and takes its slot's figure instead.
#:
#: A SLOT IS A KEY AND A SKILL IS WORTH WHAT IT IS WORTH, decided by the project
#: owner on 2026-08-22 and recorded in docs/DECISIONS.md. Any skill may go in any
#: slot, so a skill taking its damage from whichever slot it happened to be put
#: in would be worth 250% of weapon damage on one key and 400% on another.
#:
#: EVERY CELL IS BLANK TODAY AND THAT IS DELIBERATE. The mechanism lands before
#: the numbers so that nothing changes until a number is written: a blank falls
#: back to the slot's figure, which is what the game did before. Writing the 112
#: designed skills' numbers is the rest of issue #836.
#:
#: NOT ZERO, for the same reason the critical strike sentinel is not. A Support
#: skill deals 0% of weapon damage, a skill may have no cooldown, and a skill may
#: be free, so zero has to mean zero in all three.
UNSTATED_SKILL_NUMBER = -1.0


def skill_number(cell: str, what: str, where: str) -> float:
    """A per-skill damage, cooldown or mana cost, or -1 when the cell is blank.

    Blank is the ordinary case and is not a fault: every one of the 398 rows is
    blank today.
    """
    if not cell:
        return UNSTATED_SKILL_NUMBER

    try:
        value = float(cell)
    except ValueError:
        raise DataError(
            f"{where}: states a {what} of {cell!r}, which is not a number. "
            "Leave the cell blank to take the slot's figure.") from None

    # NEGATIVE IS REFUSED RATHER THAN PASSED THROUGH, because -1 is the sentinel
    # meaning the row says nothing. A row meaning to state -1 of something would
    # be read as stating nothing at all, and nothing would report it.
    if value < 0.0:
        raise DataError(
            f"{where}: states a {what} of {value}. A negative figure is not "
            "meaningful, and -1 already means the row says nothing. Leave the "
            "cell blank instead.")

    return value


#: What the CritChancePercent column holds for a skill that states no critical
#: strike chance of its own, and takes the 5% default instead.
#:
#: NOT ZERO, AND THAT IS THE WHOLE REASON A SENTINEL IS NEEDED. The decision
#: recorded in docs/DECISIONS.md on 2026-08-04 says the 5% is "a default and not
#: a floor: a skill that states 1% gets 1%, which is what lets a skill be
#: designed to crit less than average". A skill designed never to critically
#: strike states 0, so zero has to mean zero. The same convention -1 carries on
#: the Cataclysm.CritRoll console variable, where it means "roll normally".
UNSTATED_CRIT_CHANCE = -1.0


def skill_crit_chance(cell: str, where: str) -> float:
    """The Crit Chance cell as a number, or -1 when the cell is blank.

    Blank is the ordinary case and is not a fault: all 398 rows are blank today,
    and a skill that says nothing about critical strikes should not have to.
    """
    if not cell:
        return UNSTATED_CRIT_CHANCE

    try:
        chance = float(cell)
    except ValueError:
        raise DataError(
            f"{where}: states a critical strike chance of {cell!r}, which is not "
            "a number. Leave the cell blank for the 5% default.") from None

    # THE CAP IS 100 AND IT IS HARD. docs/Cataclysm_GDD_v2.md's caps table says
    # "Above 100% it means nothing", and both the model and the engine clamp
    # there -- HARD_CAPS in sim/cataclysm_sim/character.py and CritChanceCap in
    # CataclysmCombatAttributeSet.h. A row stating 150 would be silently clamped
    # to 100, so it is refused here instead. Whether a keystone may lift that cap
    # is an open design question, issue #658; if it is ever lifted, this bound
    # moves with it.
    if not 0.0 <= chance <= 100.0:
        raise DataError(
            f"{where}: states a critical strike chance of {chance}, which is "
            "outside 0 to 100. The cap is hard; see the caps table in "
            "docs/Cataclysm_GDD_v2.md.")

    return chance


def tags_with_slot(written: str, slot: str, where: str) -> str:
    """Add the row's slot tag to its Tags cell, derived from the Slot column.

    WHY IT IS DERIVED AND NOT WRITTEN. The Tags sheet declares a tag for every
    ability slot, and the Weapon Skills sheet used to carry two of them by hand:
    `Slot.Movement` on every Movement row and `Slot.Ultimate` on every Ultimate
    row, with `Slot.Heavy`, `Slot.Special`, `Slot.Support` and `Slot.Aura` on
    nothing at all. The design says increases are scoped by tag and lists slot
    tags among the scopes, so an affix reading "increased Heavy Attack damage"
    would be a modifier scoped to `Slot.Heavy` -- which would have applied to no
    skill in the game, and nothing would have reported it. Issue #156.

    Deriving it rather than writing it into 398 rows means the Slot column is the
    only place a row's slot is stated, so the tag cannot disagree with it. That
    is the same reason `Cataclysm.Input.EveryAbilitySlotHasAGeneratedTag` exists.

    A SLOT TAG WRITTEN BY HAND IS REFUSED, rather than merged. Allowing both
    would put the slot in two places again, which is the thing this removes.
    """
    tags = [t.strip() for t in written.split(",") if t.strip()]

    hand_written = [t for t in tags if t.startswith("Slot.")]
    if hand_written:
        raise DataError(
            f"{where}: carries the slot tag {hand_written[0]!r} in its Tags "
            f"cell. Slot tags come from the Slot column and are added by the "
            f"generator; remove it from the sheet.")

    if not slot:
        return ", ".join(tags)

    tags.append(f"Slot.{slot}")
    return ", ".join(tags)


def enchantments(book, negative: bool) -> list[dict]:
    """The sheet holds two independent tables side by side.

    Positives occupy columns A-D, negatives F-I. They are NOT paired: the design
    states positives and negatives roll independently, so a row in one table has
    no relationship to the row beside it.
    """
    first = 5 if negative else 0
    out = []
    for index, raw in enumerate(book["Enchantments"].iter_rows(values_only=True), 1):
        if index == 1 or not raw or len(raw) <= first + 3:
            continue
        text = clean(raw[first])
        if not text:
            continue
        kind = "Negative" if negative else "Positive"
        out.append({"Name": row_name(kind, text[:48]),
                    "Effect": text,
                    "EnchantmentType": clean(raw[first + 1]),
                    "Weight": number(raw[first + 2], "Weight", index),
                    "Tags": clean(raw[first + 3]),
                    "IsNegative": "True" if negative else "False"})
    return unique(out, f"Enchantments ({'negative' if negative else 'positive'})")


def enemy_modifiers(book) -> list[dict]:
    """A matrix: one column per Cataclysm, each cell "Name: Description"."""
    rows = list(book["Enemy Modifiers"].iter_rows(values_only=True))
    if not rows:
        raise DataError("Enemy Modifiers is empty")

    headers = [clean(h) for h in rows[0]]
    out = []
    for column, header in enumerate(headers):
        if not header:
            continue
        cataclysm = header.replace(" Modifiers", "").strip()
        if cataclysm not in CATACLYSM_TYPES:
            raise DataError(f"Enemy Modifiers column {column + 1}: "
                            f"{header!r} does not name a Cataclysm type")
        for raw in rows[1:]:
            if column >= len(raw):
                continue
            text = clean(raw[column])
            if not text:
                continue
            name, description = split_named(text)
            if not name:
                name = text[:40]
            out.append({"Name": row_name(cataclysm, name),
                        "CataclysmType": cataclysm,
                        "ModifierName": name,
                        "Description": description})
    return unique(out, "Enemy Modifiers")


#: The numeric columns of the Buffs, Debuffs and DoTs sheets, in sheet order,
#: paired with the field each becomes. Read positionally because those three
#: sheets have NO HEADING ROW, which is a documented property of them rather
#: than an oversight: `docs/README.md` says so, `test_docs_readme_sheet_table_
#: is_true.py` names all three in its HEADERLESS set, and giving them one would
#: change every row count in that table.
#:
#: SO THE ORDER OF THIS TUPLE IS THE SCHEMA. Inserting an entry anywhere but the
#: end silently re-reads every column after it, and nothing would report an
#: error -- a duration would arrive as a strength and both would be wrong.
#: Append only, and say what a new column means in `docs/README.md` at the same
#: time, because the sheet itself cannot say.
STATUS_EFFECT_NUMBERS: tuple[tuple[int, str], ...] = (
    (1, "DurationSeconds"),         # column B
    (2, "PercentOfHit"),            # column C
    (3, "Strength"),                # column D
    (4, "StrengthCap"),             # column E
    (5, "DurationCap"),             # column F
    (6, "PercentOfCurrentHealth"),  # column G
    (7, "FlatDamagePerTick"),       # column H
)


def status_effects(book) -> list[dict]:
    """Buffs, Debuffs and DoTs: "Name: Description", then seven optional numbers.

    The first row is data, not a header. Reading it as a header would silently
    drop one effect from each of the three sheets.

    EVERY NUMBER IS OPTIONAL because most rows state none. An empty cell reads
    as zero, and an effect worth zero applies nothing -- which is the honest
    answer for an effect nobody has designed yet, and is what
    `UCataclysmSkillEffects::BurnNumbers` checks for before applying anything.

    COLUMN B IS HOW LONG AN EFFECT LASTS. It was added for Burn, which every one
    of the sixteen designed Demonic skills applies and which stated no duration
    and no damage anywhere in the design -- so a skill reading "sets each one
    alight" applied an effect with no numbers in it.

    THREE COLUMNS ARE ALTERNATIVE WAYS OF SAYING WHAT ONE TICK DEALS, and an
    effect states exactly one of them. All three are per tick and not a total;
    `docs/DECISIONS.md` carries that decision, made on 2026-08-24.

    * `FlatDamagePerTick`, column H, is a plain amount. Bleed, Poison, Disease,
      Burn and Necrosis use it. The project owner chose it over a percent of the
      hit on 2026-08-24, because a percent of the hit multiplies twice -- the hit
      already grows with the difficulty tier and the three damage over time stats
      multiply on top of it -- while a flat amount grows only with those stats
      and so stays level across the eight tiers.
    * `PercentOfHit`, column C, is a percent of the hit that applied the effect.
      NOTHING USES IT AS OF 2026-08-24. It was Burn's base until the decision
      above. Kept rather than removed because a skill stating its own effect is
      the obvious future caller and removing it would churn the struct and two
      C++ tests for no gain.
    * `PercentOfCurrentHealth`, column G, is for an effect measured against the
      target. Only Void Splinter uses it, at 1% a second. It cannot go through
      the ordinary damage over time path, because current health falls between
      ticks and so the per-tick amount is not fixed. Issue #915.

    SEPARATE COLUMNS RATHER THAN ONE COLUMN AND A STRING NAMING ITS BASIS,
    because a misspelled basis would silently read as one of the others and a
    number cannot be misspelled.

    COLUMNS D TO F ARE AN EFFECT'S STRENGTH AND ITS CAPS. They were added for
    issue #904, because four of the eleven ailments could not be written down at
    all without them:

    * `Strength` is the effect's own magnitude in whatever unit it names --
      Cripple's 30% slow, Weaken's 20% damage reduction, Shred's 10 resistance,
      Necrosis's total denial of healing.
    * `StrengthCap` is where that magnitude stops and rolls into duration
      instead, which is the design's rule for every effect that has one. Empty
      means no NUMERIC cap: Shred's cap is the target's own resistance reaching
      zero, which is a property of the target and not a number of this effect's.
    * `DurationCap` is where the duration stops. Only Stun has one, and Stun is
      the one effect whose scaling stops dead rather than rolling over, because
      its magnitude IS its duration and there is nothing to roll into.
    """
    out = []
    for sheet, kind in (("Buffs", "Buff"), ("Debuffs", "Debuff"), ("DoTs", "DoT")):
        for index, raw in enumerate(book[sheet].iter_rows(values_only=True), 1):
            if not raw:
                continue
            text = clean(raw[0])
            if not text:
                continue
            name, description = split_named(text)
            if not name:
                name = text[:40]
            row = {"Name": row_name(kind, name), "EffectKind": kind,
                   "EffectName": name, "Description": description}
            for column, field in STATUS_EFFECT_NUMBERS:
                cell = raw[column] if len(raw) > column else None
                row[field] = (number(cell, field, index)
                              if cell is not None else 0.0)
            out.append(row)
    return unique(out, "Status Effects")


def gems(book) -> list[dict]:
    """Gem effects and their value at each of the eight rarity tiers.

    The sheet's "Everyday Gemstone" column holds the effect description, and that
    description STATES the Everyday value: "10% chance to apply void splinter"
    means Everyday is 10%. The seven numeric columns that follow are Quality
    through Cataclysmic, and the series continues from the value in the text --
    10%, then 30, 50, 75, 100, 125, 150, 200.

    So all eight tiers have a value; only the first is written in prose. It is
    extracted here so consumers get eight numbers rather than seven and a
    sentence. Every one of the 25 gems states a percentage, which is checked.
    """
    rows = list(book["Gems"].iter_rows(values_only=True))
    tiers = ["Quality", "Superb", "Masterful", "Legendary",
             "Mythical", "Ascendant", "Cataclysmic"]
    out = []
    seen_names: dict[str, int] = {}
    for index, raw in enumerate(rows[1:], start=2):
        if not raw or not clean(raw[0]):
            continue
        name, effect = clean(raw[0]), clean(raw[1])

        # `\d*\.?\d+` and NOT `\d+(?:\.\d+)?`, which is what this was until issue
        # #246. That pattern requires a digit before the decimal point, so on
        # "Increases Magic Find by .5%" it matched "5%" and made the Everyday
        # value 0.05 -- ten times the stated figure, and larger than the six
        # rarity tiers above it. Nothing reported an error; the gem simply
        # shipped with a value its own description contradicted.
        match = re.search(r"(\d*\.?\d+)\s*%", effect)
        if not match:
            raise DataError(f"Gems row {index}: {name!r} states no percentage in "
                            f"its effect text, so the Everyday value cannot be "
                            f"read: {effect!r}")
        everyday = float(match.group(1)) / 100.0

        # A GEM NAME IS WHAT THE PLAYER READS, so two gems cannot share one.
        # `unique` below would silently key the second Gem_Of_Recovery_1 and both
        # rows would import, leaving two different gems on the ground under one
        # name with no way to tell them apart. That is what happened: one raised
        # health regeneration and one reduced cooldowns, for a month (issue #211).
        # Checked here rather than in a test because the generator is the only
        # place that sees both the sheet and the row key.
        if name in seen_names:
            raise DataError(
                f"Gems row {index}: {name!r} is already the name of the gem on "
                f"row {seen_names[name]}. Two gems cannot share a name: a player "
                f"reading it cannot tell which one they picked up, and the row "
                f"keys differ only by a numeric suffix nothing explains.")
        seen_names[name] = index

        entry = {"Name": row_name("Gem", name), "GemName": name,
                 "Effect": effect,
                 "GemType": clean(raw[9]) if len(raw) > 9 else "",
                 "Everyday": everyday}
        for offset, tier in enumerate(tiers, start=2):
            entry[tier] = number(raw[offset], tier, index) if offset < len(raw) else 0.0

        # A RARER GEM HAS TO BE WORTH MORE. Gear and gem rarity equal the
        # difficulty tier, so the ladder is the whole of a gem's progression: a
        # tier that pays less than the one below it means a player who finds a
        # rarer gem should keep the commoner one, and nothing in the interface
        # would tell them so.
        #
        # Checked here rather than in a test because the generator is the only
        # place that sees the prose value and the seven numeric ones together.
        # Issue #246: reading ".5%" as 5% put Of The Goblin's Everyday value
        # above its next six tiers, and the table shipped that way because no
        # check compared them.
        ladder = [entry[t] for t in ("Everyday", *tiers)]
        for lower, higher in zip(ladder, ladder[1:], strict=False):
            if higher <= lower:
                raise DataError(
                    f"Gems row {index}: {name!r} does not get better as it gets "
                    f"rarer. Its eight values are {ladder}, and {higher} is not "
                    f"above {lower}. The first of those is read from the effect "
                    f"text {effect!r} and the rest are the sheet's numeric "
                    f"columns.")

        out.append(entry)
    return unique(out, "Gems")


def parse_tier(value, effect: str, field: str, row: int) -> tuple[str, float, float]:
    """Read one city-upgrade tier cell as (kind, value, interval).

    The sheet has no column saying which kind a cell is, so it is inferred from
    the cell's own notation and the effect text. The four kinds, as stated by the
    project owner:

      "10/10%"  IntervalPercent  two values at once. The effect reads "every X
                                 days ... Y%", and the tier improves BOTH: the
                                 interval drops and the magnitude rises. Here
                                 interval 10 days, magnitude 10%.
      "3x"      Multiplier       multiplies the effect.
      "0.3"     Percent          a percentage increase, stored as a fraction.
      "10"      Flat             a flat improvement, in whatever unit the effect
                                 names -- days, floors, a count of dungeons.

    Percent and Flat are told apart by whether the effect text mentions a
    percentage, because both are bare numbers. That is a heuristic on prose, and
    it is why an explicit kind column in the sheet would be better.
    """
    text = clean(value)
    if not text:
        return "", 0.0, 0.0

    # Must be tested before the others: this cell contains a percent sign AND a
    # bare number, so either of the later branches would match it and lose half
    # the value.
    if "/" in text:
        interval_text, _, magnitude_text = text.partition("/")
        magnitude_text = magnitude_text.strip().rstrip("%")
        try:
            interval = float(interval_text.strip())
            magnitude = float(magnitude_text) / 100.0
        except ValueError as error:
            raise DataError(f"City Upgrades row {row}: {field} is {text!r}, which "
                            f"is not 'days/percent'") from error
        return "IntervalPercent", magnitude, interval

    if text.lower().endswith("x"):
        try:
            return "Multiplier", float(text[:-1]), 0.0
        except ValueError as error:
            raise DataError(f"City Upgrades row {row}: {field} is {text!r}, which "
                            f"is not a multiplier") from error

    try:
        amount = float(text)
    except ValueError as error:
        raise DataError(f"City Upgrades row {row}: {field} is {text!r}, which is "
                        f"none of a percentage, a flat number, a multiplier or "
                        f"a days/percent pair") from error

    return ("Percent" if "%" in effect else "Flat"), amount, 0.0


def city_upgrades(book) -> list[dict]:
    """City upgrades and their tier 2 and tier 3 scaling.

    A trailing asterisk on the branch name marks a ONE-TIME USE upgrade: it fires
    once and is spent, rather than being a standing improvement. The asterisk is
    turned into a flag here and stripped from the branch name, so nothing has to
    parse punctuation to know how an upgrade behaves.

    One row has no branch. It is also one-time use, has no tiers, and is a last
    resort rather than a city improvement. Which branch it belongs to has not
    been decided, so Branch is left empty and BranchUndecided marks it.
    """
    out = []
    for index, raw in enumerate(book["City Upgrades"].iter_rows(values_only=True), 1):
        if index == 1 or not raw:
            continue
        branch, effect = clean(raw[0]), clean(raw[1])
        if not effect:
            continue

        one_time = branch.endswith("*") or not branch
        branch = branch.rstrip("*").strip()

        t2_kind, t2_value, t2_interval = parse_tier(
            raw[2] if len(raw) > 2 else "", effect, "Tier 2", index)
        t3_kind, t3_value, t3_interval = parse_tier(
            raw[3] if len(raw) > 3 else "", effect, "Tier 3", index)

        out.append({"Name": row_name(branch or "Unbranched", effect[:40]),
                    "Branch": branch,
                    "BranchUndecided": "True" if not branch else "False",
                    "IsOneTimeUse": "True" if one_time else "False",
                    "Effect": effect,
                    # The raw cell is kept alongside the parsed values so nothing
                    # is lost if the inference above is ever wrong.
                    "Tier2Raw": clean(raw[2]) if len(raw) > 2 else "",
                    "Tier2Kind": t2_kind,
                    "Tier2Value": t2_value,
                    "Tier2IntervalDays": t2_interval,
                    "Tier3Raw": clean(raw[3]) if len(raw) > 3 else "",
                    "Tier3Kind": t3_kind,
                    "Tier3Value": t3_value,
                    "Tier3IntervalDays": t3_interval})
    return unique(out, "City Upgrades")


def crafting_materials(book) -> list[dict]:
    """The Crafting sheet, which holds two kinds of row.

    EIGHTEEN OF THEM ARE MATERIALS AND NINETEEN ARE CRAFTING ACTIONS. A material
    says which tier it belongs to in its Tier and Source cell -- "Tier 3 (Rare).
    Drop from Dungeon Bosses/Elites." -- and an action such as "Reroll Affix
    Value" says nothing of the kind, because it is a thing you do rather than a
    thing you hold.

    THE TIER IS PUBLISHED AS A NUMBER so the engine does not have to read prose.
    A drop picks a tier and then picks equally among the materials in it, which
    is what `roll_material_tier` in `sim/cataclysm_sim/loot.py` says belongs to
    "whoever holds that table". Zero means the row is an action rather than a
    material, so nothing can drop it.

    AND SO IS AN UPGRADE STONE'S LEVEL, for the same reason and by the same
    route. The ten stones are named "Upgrade Stone +1" through "+10", so the
    level is already stated; publishing it as a number means the engine does not
    parse a name at run time, which it would otherwise do for every material
    that drops. Zero for everything that is not a stone.

    A STONE'S LEVEL IS NOT ITS TIER. The tier is the rarity band it drops in and
    the level is how far it upgrades a piece. The ten were deliberately spread
    two to a band, so a band holds two different levels and neither can be
    derived from the other. Issue #863 is what needs the level: the best stone
    that can drop is capped by the difficulty tier.
    """
    out = []
    for index, raw in enumerate(book["Crafting"].iter_rows(values_only=True), 1):
        if index == 1 or not raw or not clean(raw[0]):
            continue
        name = clean(raw[0])
        source = clean(raw[1])
        found = MATERIAL_TIER_AND_SOURCE.match(source or "")
        stone = UPGRADE_STONE_NAME.match(name or "")
        upgrade_level = int(stone.group(1)) if stone else 0
        if upgrade_level > MAX_UPGRADE_STONE_LEVEL:
            raise DataError(
                f"Crafting row {index}: {name!r} names upgrade level "
                f"{upgrade_level}, and gear runs from +1 to "
                f"+{MAX_UPGRADE_STONE_LEVEL}")
        out.append({"Name": row_name("Material", name), "MaterialName": name,
                    "Tier": int(found.group(1)) if found else 0,
                    "UpgradeLevel": upgrade_level,
                    "TierAndSource": source,
                    "PrimaryUse": clean(raw[2]),
                    "Functions": clean(raw[3]),
                    "CRMetric": clean(raw[6]) if len(raw) > 6 else "",
                    "Formula": clean(raw[7]) if len(raw) > 7 else "",
                    "Outcome": clean(raw[8]) if len(raw) > 8 else ""})
    return unique(out, "Crafting")


IMPLICIT_KINDS = ("flat", "increased")
AFFIX_KINDS = ("Stat", "Resistance", "Ailment", "Hybrid")
AFFIX_POSITIONS = ("prefix", "suffix")


def _header_index(rows: list, sheet: str) -> dict[str, int]:
    """Map header text to column index, so a reordered sheet still reads."""
    if not rows:
        raise DataError(f"{sheet} is empty")
    headers = {clean(h): i for i, h in enumerate(rows[0]) if clean(h)}
    if not headers:
        raise DataError(f"{sheet} has no header row")
    return headers


def _cell(raw, headers: dict[str, int], name: str) -> str:
    index = headers.get(name)
    if index is None or index >= len(raw):
        return ""
    return clean(raw[index])


def item_bases(book) -> list[dict]:
    """One item base per row, with up to two implicits in fixed columns.

    An implicit belongs to the BASE rather than to the slot: a chest built for
    armour and one built for evasion are different bases in the same slot, and
    choosing between them is a decision made before any affix is involved.

    Values are the STATED ones, meaning the fully upgraded figures the design
    document quotes. For a two-handed weapon that is the figure BEFORE the
    two-handed multiplier doubles it, so a Greatsword is 78 here and supplies
    156 in play.
    """
    rows = list(book["Item Bases"].iter_rows(values_only=True))
    headers = _header_index(rows, "Item Bases")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        name = _cell(raw, headers, "Base Name")
        if not name:
            continue
        slot = _cell(raw, headers, "Slot")
        if not slot:
            raise DataError(f"Item Bases row {index}: {name} has no slot")

        hands = _cell(raw, headers, "Hands")
        if hands and hands not in ("1", "2", "1.0", "2.0"):
            raise DataError(f"Item Bases row {index}: {name} is "
                            f"{hands}-handed; a weapon has 1 or 2")

        # Attacks per second, and only a weapon has one. It is NOT an implicit:
        # a two-handed weapon doubles every implicit it carries, which is right
        # for damage and would make a Greatsword swing twice as fast as a Sword.
        # See issue #120 and WeaponBase.attack_speed in
        # sim/cataclysm_sim/affixes.py.
        attack_speed = _cell(raw, headers, "Attack Speed")
        if hands and not attack_speed:
            raise DataError(
                f"Item Bases row {index}: the {name} is a weapon with no attack "
                "speed. Every weapon needs one, because the weapon is where "
                "that base comes from and an increase to zero is worth nothing.")
        if attack_speed and not hands:
            raise DataError(
                f"Item Bases row {index}: {name} is not a weapon but has an "
                "attack speed. Only a weapon supplies that base.")

        # THE BASIC ATTACK'S SHAPE AND REACH. It lives on the weapon base rather
        # than in the Weapon Skills sheet because it does not vary by damage
        # type: the design document calls it weapon damage itself, so there is
        # one per weapon rather than one per weapon-and-damage-type pair. Decided
        # 2026-08-15 on issue #524, and the design document says the matrix holds
        # one skill per NON-BASIC slot, which putting it here keeps true.
        basic_shape = _cell(raw, headers, "Basic Shape")
        basic_params = _cell(raw, headers, "Basic Shape Params")

        # HOW MANY CELLS THE PIECE OCCUPIES IN THE CARRIED BAG, from the
        # footprint table in docs/Inventory_Screen_Design.md. Issue #855.
        #
        # REQUIRED OF EVERY BASE. A base with no footprint could not be put
        # in the bag at all, and a zero would read as an item that takes no
        # room rather than as a base somebody forgot.
        wide = _cell(raw, headers, "Cells Wide")
        high = _cell(raw, headers, "Cells High")
        for label, figure in (("Cells Wide", wide), ("Cells High", high)):
            if not figure or int(float(figure)) < 1:
                raise DataError(
                    f"Item Bases row {index}: {name} has no {label}. Every\n"
                    "base needs a footprint of at least one cell; see the\n"
                    "footprint table in docs/Inventory_Screen_Design.md.")

        entry = {
            "Name": row_name(slot, name),
            "BaseName": name,
            "Slot": slot,
            "Hands": int(float(hands)) if hands else 0,
            "SubType": _cell(raw, headers, "Sub-Type"),
            "WeaponType": _cell(raw, headers, "Weapon Type"),
            "MaxDamageTypes": int(float(
                _cell(raw, headers, "Max Damage Types") or 0)),
            "AttackSpeed": float(attack_speed) if attack_speed else 0.0,
            "BasicShape": basic_shape,
            "BasicShapeParams": basic_params,
            "CellsWide": int(float(wide)),
            "CellsHigh": int(float(high)),
        }

        implicits = 0
        arms_the_holder = False
        for slot_index in (1, 2):
            stat = _cell(raw, headers, f"Implicit {slot_index} Stat")
            kind = _cell(raw, headers, f"Implicit {slot_index} Kind")
            value = _cell(raw, headers, f"Implicit {slot_index} Value")
            if stat:
                if kind not in IMPLICIT_KINDS:
                    raise DataError(
                        f"Item Bases row {index}: {name} implicit {slot_index} "
                        f"kind is {kind!r}, expected one of {list(IMPLICIT_KINDS)}")
                implicits += 1
            # Read off the implicits rather than off the name, which is what
            # player_damage.armed_weapons_in also does. The Shield is the only
            # weapon today that arms nobody, and naming it here instead would
            # make a second such weapon silently wrong.
            #
            # THE TWO ARE NOT IDENTICAL AND THE DIFFERENCE IS DELIBERATE. That
            # function sums any attack_damage implicit; this one requires the
            # kind to be flat. They agree on today's data, because no weapon
            # grants increased attack damage as an implicit. This is the stricter
            # of the two, because a weapon granting only an increase supplies no
            # damage to increase.
            if stat == "attack_damage" and kind == "flat":
                arms_the_holder = True
            entry[f"Implicit{slot_index}Stat"] = stat
            entry[f"Implicit{slot_index}Kind"] = kind
            entry[f"Implicit{slot_index}Value"] = float(value) if value else 0.0

        if implicits == 0:
            raise DataError(f"Item Bases row {index}: {name} grants nothing, so "
                            "it is not a distinct base")

        check_basic_attack(basic_shape, basic_params, arms_the_holder,
                           bool(hands), f"Item Bases row {index} ({name})")
        out.append(entry)

    return unique(out, "Item Bases")


#: What a Weapon Meshes row puts in the Mesh column to mean "draw nothing".
#:
#: A WORD RATHER THAN AN EMPTY CELL, because an empty cell cannot be told apart
#: from a row somebody started and did not finish. Issue #1125 asked for exactly
#: this distinction: "Drawing nothing is a reasonable answer; drawing nothing
#: silently is not."
DRAWS_NOTHING = "None"

#: Where a weapon mesh path has to start. A mesh outside /Game/ is not in the
#: project's content at all and could never be loaded.
CONTENT_PREFIX = "/Game/"


def weapon_meshes(book) -> list[dict]:
    """Which mesh is drawn in the hand for each weapon base. Source: Weapon Meshes.

    WHY THIS IS A SHEET OF ITS OWN RATHER THAN COLUMNS ON ITEM BASES. Issue
    #1125. A mesh path is an art binding, not a design decision: it changes when
    the art changes and says nothing about how the weapon plays. Item Bases is a
    design sheet and keeping third-party asset paths out of it means a new
    weapons pack moves one sheet rather than editing the design. This follows
    Element Visuals, which is the same shape for the same reason.

    THE ROW KEY MATCHES ITEM BASES, so `FCataclysmItem::Base` looks a row up
    directly with no translation. Both are built by `row_name("Weapon", name)`.

    EVERY WEAPON BASE MUST APPEAR AND NOTHING ELSE MAY. A base with no row would
    silently draw nothing, which is the thing this table exists to make
    impossible, and a row for a base that does not exist is a typo that would
    otherwise never be noticed.
    """
    rows = list(book["Weapon Meshes"].iter_rows(values_only=True))
    headers = _header_index(rows, "Weapon Meshes")

    # Which bases are weapons, read off Item Bases rather than listed here, so
    # adding a weapon base to the design cannot leave this table behind.
    base_rows = list(book["Item Bases"].iter_rows(values_only=True))
    base_headers = _header_index(base_rows, "Item Bases")
    weapon_bases = {
        _cell(raw, base_headers, "Base Name")
        for raw in base_rows[1:]
        if _cell(raw, base_headers, "Slot") == "Weapon"
        and _cell(raw, base_headers, "Base Name")
    }

    out = []
    covered = set()
    for index, raw in enumerate(rows[1:], start=2):
        name = _cell(raw, headers, "Base Name")
        if not name:
            continue

        where = f"Weapon Meshes row {index} ({name})"

        if name not in weapon_bases:
            raise DataError(
                f"{where}: there is no weapon base called {name!r} in the Item\n"
                f"Bases sheet. This table is keyed by base name and every key\n"
                f"has to name a real weapon.")
        if name in covered:
            raise DataError(f"{where}: {name!r} already has a row.")
        covered.add(name)

        mesh = _cell(raw, headers, "Mesh")
        if not mesh:
            raise DataError(
                f"{where}: the Mesh cell is empty. A base that should draw\n"
                f"nothing says {DRAWS_NOTHING!r}, so that a row nobody filled\n"
                f"in can be told apart from a weapon that is meant to be\n"
                f"invisible. The Fist is the designed example: unarmed draws\n"
                f"nothing on purpose.")

        if mesh != DRAWS_NOTHING and not mesh.startswith(CONTENT_PREFIX):
            raise DataError(
                f"{where}: the mesh is {mesh!r}, which does not start with\n"
                f"{CONTENT_PREFIX!r}. A mesh outside the project's content\n"
                f"cannot be loaded. Write {DRAWS_NOTHING!r} to draw nothing.")

        scale = _cell(raw, headers, "Scale")
        try:
            scale_value = float(scale)
        except (TypeError, ValueError) as exc:
            raise DataError(
                f"{where}: the scale is {scale!r}, which is not a number.\n"
                f"Write 1 for a mesh drawn at the size the pack authored."
            ) from exc

        if scale_value <= 0.0:
            raise DataError(
                f"{where}: the scale is {scale_value}. A scale of zero or less\n"
                f"is a weapon that cannot be seen, which is what writing\n"
                f"{DRAWS_NOTHING!r} in the Mesh column is for.")

        out.append({
            "Name": row_name("Weapon", name),
            "BaseName": name,
            # Emptied here rather than carrying the word through, so the game
            # tests one thing -- "is the path empty" -- instead of comparing
            # against a magic word it would have to keep in step with this file.
            "Mesh": "" if mesh == DRAWS_NOTHING else mesh,
            "Scale": scale_value,
        })

    missing = sorted(weapon_bases - covered)
    if missing:
        raise DataError(
            "Weapon Meshes has no row for " + ", ".join(missing) + ".\n"
            "Every weapon base needs one, because a base with no row would\n"
            f"draw nothing without saying so. Write {DRAWS_NOTHING!r} in the\n"
            "Mesh column for a weapon that is meant to be invisible.")

    return unique(out, "Weapon Meshes")


def affixes(book) -> list[dict]:
    """One rollable affix per row, across four kinds.

    Stat        one stat, flat or increased, with a top value
    Resistance  a family covering `Breadth` damage types at `Top Value` each
    Ailment     a chance to apply an effect a gem also applies
    Hybrid      two stat affixes at a reduced share each

    A stat that appears as a prefix never appears as a suffix, which is what
    stops one item carrying four of whatever is strongest.
    """
    rows = list(book["Affixes"].iter_rows(values_only=True))
    headers = _header_index(rows, "Affixes")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        name = _cell(raw, headers, "Affix Name")
        if not name:
            continue

        kind = _cell(raw, headers, "Affix Kind")
        if kind not in AFFIX_KINDS:
            raise DataError(f"Affixes row {index}: {name} has kind {kind!r}, "
                            f"expected one of {list(AFFIX_KINDS)}")

        position = _cell(raw, headers, "Position")
        if position not in AFFIX_POSITIONS:
            raise DataError(f"Affixes row {index}: {name} is a {position!r}, "
                            f"expected one of {list(AFFIX_POSITIONS)}")

        allowed = _cell(raw, headers, "Allowed Slots")
        if not allowed:
            raise DataError(f"Affixes row {index}: {name} can roll on no slot")

        value = _cell(raw, headers, "Top Value")
        if kind in ("Stat", "Resistance", "Ailment") and not value:
            raise DataError(f"Affixes row {index}: {name} has no top value")

        # THE WORD THIS AFFIX GIVES AN ITEM'S NAME, and only a suffix has one.
        # An item is called `<rarity> <base> of <word>`, so the first word is
        # the rarity: a prefix has nowhere in the name to appear, and a word on
        # one could never be read. A suffix without one leaves an item that
        # rolled it with nothing to be named after. See sim/cataclysm_sim/
        # naming.py, which mirrors this column.
        name_word = _cell(raw, headers, "Name Word")
        if position == "suffix" and not name_word:
            raise DataError(
                f"Affixes row {index}: the suffix {name} has no name word, so "
                "an item rolling it would have nothing to be called after")
        if position == "prefix" and name_word:
            raise DataError(
                f"Affixes row {index}: the prefix {name} carries the name word "
                f"{name_word!r}, and an item's first word is its rarity, so "
                "that word could never appear")

        out.append({
            "Name": row_name(kind, name),
            "AffixName": name,
            "AffixKind": kind,
            "Position": position,
            "Stat": _cell(raw, headers, "Stat"),
            "ValueKind": _cell(raw, headers, "Value Kind"),
            "TopValue": float(value) if value else 0.0,
            "Breadth": int(float(_cell(raw, headers, "Breadth") or 0)),
            "Ailment": _cell(raw, headers, "Ailment"),
            "Gem": _cell(raw, headers, "Gem"),
            "HybridPart1": _cell(raw, headers, "Hybrid Part 1"),
            "HybridPart2": _cell(raw, headers, "Hybrid Part 2"),
            "AllowedSlots": allowed,
            "NameWord": name_word,
        })

    # TWO AFFIXES SHARING A WORD would make an item's name say less than it
    # looks like it says. A player reading "of Warding" should be able to look
    # it up and find one thing.
    named: dict[str, str] = {}
    for row in out:
        word = row["NameWord"]
        if not word:
            continue
        if word in named:
            raise DataError(f"Affixes: the name word {word!r} is carried by "
                            f"both {named[word]!r} and {row['AffixName']!r}")
        named[word] = row["AffixName"]

    return unique(out, "Affixes")


#: The eight item rarities, weakest first. The Gear Rarity sheet is read in this
#: order, so a rarity missing from the sheet, misspelled there, or listed out of
#: order fails generation rather than producing a ladder the drop cascade walks
#: wrongly.
#:
#: WRITTEN OUT HERE the way CATACLYSM_TYPES is, rather than read from
#: sim/cataclysm_sim/affixes.py. This file keeps the workbook tables and the two
#: model-generated tables apart on purpose, and the rarity ladder belongs to the
#: workbook side. tools/tests/test_generated_loot_tables_match_the_model.py
#: compares this list against the model's RARITIES and against the
#: ECataclysmRarity enum the engine looks these rows up by.
RARITY_LADDER = ("Everyday", "Quality", "Superb", "Masterful",
                 "Legendary", "Mythical", "Ascendant", "Cataclysmic")

#: The highest upgrade level a piece can reach, so a rarity gated above it is a
#: typo rather than a design. Mirrors UCataclysmItemValues::MaxGearLevel and
#: affixes.GEAR_LEVELS.
MAX_GEAR_LEVEL = 10

#: How many hands a weapon takes. Any other value in the Item Sockets sheet is a
#: mistyped cell, since the sheet uses 0 for everything that is not a weapon.
WEAPON_HAND_COUNTS = (1, 2)


def gear_rarity(book) -> list[dict]:
    """One rarity per row: how often it drops, and what a drop of it arrives at.

    THE ROW KEY IS THE RARITY'S OWN NAME, which is also the name of its
    ECataclysmRarity entry. That is the join between the table and the engine:
    the engine walks the ladder by the enum and looks each row up by the enum's
    name, so nothing depends on the order Unreal happens to iterate a DataTable's
    rows in.

    FOUR NUMBERS PER RARITY, and each one is checked against the rarity below it
    because all four move in a known direction as rarity rises:

      Drop Weight        never rises. A rarer thing cannot be more common.
      Gear Level Gate    never falls. Legendary needs +4 and Cataclysmic +10.
      Residue band       neither end falls. A better drop costs more to improve.

    A cell that moves the wrong way is a typo rather than a design, and none of
    them would fail anywhere else: the table would load, the cascade would run,
    and the numbers would simply be wrong.
    """
    rows = list(book["Gear Rarity"].iter_rows(values_only=True))
    headers = _header_index(rows, "Gear Rarity")

    found: dict[str, dict] = {}
    for index, raw in enumerate(rows[1:], start=2):
        rarity = _cell(raw, headers, "Rarity")
        if not rarity:
            continue
        if rarity not in RARITY_LADDER:
            raise DataError(f"Gear Rarity row {index}: {rarity!r} is not a "
                            f"rarity; expected one of {list(RARITY_LADDER)}")
        if rarity in found:
            raise DataError(f"Gear Rarity row {index}: {rarity} appears twice")

        weight = number(_cell(raw, headers, "Drop Weight"), "Drop Weight", index)
        if weight <= 0.0:
            raise DataError(f"Gear Rarity row {index}: {rarity} has a drop "
                            f"weight of {weight:g}, so it can never drop")

        gate = number(_cell(raw, headers, "Gear Level Gate"),
                      "Gear Level Gate", index)
        if not 0 <= gate <= MAX_GEAR_LEVEL:
            raise DataError(f"Gear Rarity row {index}: {rarity} is gated at "
                            f"upgrade level {gate:g}, outside 0 to "
                            f"{MAX_GEAR_LEVEL}")

        lowest = number(_cell(raw, headers, "Residue On Drop Lowest"),
                        "Residue On Drop Lowest", index)
        highest = number(_cell(raw, headers, "Residue On Drop Highest"),
                         "Residue On Drop Highest", index)
        if lowest <= 0.0:
            raise DataError(f"Gear Rarity row {index}: {rarity} drops carrying "
                            f"{lowest:g} residue. Every drop carries some.")
        if highest < lowest:
            raise DataError(f"Gear Rarity row {index}: {rarity}'s residue band "
                            f"runs from {lowest:g} down to {highest:g}")

        # THE COLOUR THE ITEM'S NAME IS DRAWN IN, and the only channel a player
        # has for reading a rarity off the floor at a glance. The design's
        # Interface Colour section names the eight; this column states them.
        # Converted from the sRGB a colour picker shows to the linear an
        # FLinearColor holds, the same way the element visuals are.
        colour = linear_colour(_cell(raw, headers, "Colour"), "Colour",
                               f"Gear Rarity row {index} ({rarity})")

        found[rarity] = {
            "Name": row_name(rarity),
            "Rarity": rarity,
            "DropWeight": weight,
            "GearLevelGate": int(gate),
            "ResidueOnDropLowest": lowest,
            "ResidueOnDropHighest": highest,
            "Colour": colour,
        }

    missing = [rarity for rarity in RARITY_LADDER if rarity not in found]
    if missing:
        raise DataError(f"Gear Rarity has no row for {missing}. Every rarity "
                        "needs one, or a drop rolling it has no weight, no "
                        "gate and no residue.")

    out = [found[rarity] for rarity in RARITY_LADDER]
    for below, above in zip(out[:-1], out[1:], strict=True):
        rising = f"{above['Rarity']} is rarer than {below['Rarity']}, and"
        if above["DropWeight"] > below["DropWeight"]:
            raise DataError(f"Gear Rarity: {rising} drops more often "
                            f"({above['DropWeight']:g} against "
                            f"{below['DropWeight']:g})")
        if above["GearLevelGate"] < below["GearLevelGate"]:
            raise DataError(f"Gear Rarity: {rising} is gated lower "
                            f"(+{above['GearLevelGate']} against "
                            f"+{below['GearLevelGate']})")
        for end in ("ResidueOnDropLowest", "ResidueOnDropHighest"):
            if above[end] < below[end]:
                raise DataError(f"Gear Rarity: {rising} its {end} is smaller "
                                f"({above[end]:g} against {below[end]:g})")

    seen_colours: dict[str, str] = {}
    for row in out:
        if row["Colour"] in seen_colours:
            raise DataError(
                f"Gear Rarity: {row['Rarity']} and {seen_colours[row['Colour']]} "
                "are drawn in the same colour, so a player could not tell them "
                "apart on the floor")
        seen_colours[row["Colour"]] = row["Rarity"]

    return unique(out, "Gear Rarity")


def item_sockets(book) -> list[dict]:
    """The most sockets a piece in each gear slot can have.

    TWO ROWS FOR A WEAPON, because its maximum depends on how many hands it
    takes: three for a one-hander and six for a two-hander, so two one-handed
    weapons match one two-hander. Every other slot has a single row with a hand
    count of 0, which is also how the Item Bases sheet writes a non-weapon.

    POTION SLOTS ARE NOT HERE. They carry one socket each and there are four of
    them, but they are consumables rather than gear and nothing rolls one as a
    drop. They are the difference between the 41 sockets this table describes
    and the 45 the design states across all equipment.
    """
    rows = list(book["Item Sockets"].iter_rows(values_only=True))
    headers = _header_index(rows, "Item Sockets")

    out = []
    seen: set[tuple[str, int]] = set()
    for index, raw in enumerate(rows[1:], start=2):
        slot = _cell(raw, headers, "Slot")
        if not slot:
            continue

        hands = int(number(_cell(raw, headers, "Hands") or 0, "Hands", index))
        if hands and hands not in WEAPON_HAND_COUNTS:
            raise DataError(f"Item Sockets row {index}: {slot} takes {hands} "
                            f"hands; a weapon takes one of "
                            f"{list(WEAPON_HAND_COUNTS)} and everything else 0")

        if (slot, hands) in seen:
            raise DataError(f"Item Sockets row {index}: {slot} at {hands} "
                            "hand(s) appears twice")
        seen.add((slot, hands))

        maximum = number(_cell(raw, headers, "Max Sockets"),
                         "Max Sockets", index)
        if maximum < 1:
            raise DataError(f"Item Sockets row {index}: {slot} holds "
                            f"{maximum:g} sockets, so no gem could ever go in "
                            "one. Remove the slot instead.")

        out.append({
            "Name": row_name(slot, f"{hands}H" if hands else ""),
            "Slot": slot,
            "Hands": hands,
            "MaxSockets": int(maximum),
        })

    return unique(out, "Item Sockets")


def affix_tiers(book) -> list[dict]:
    """How heavily each affix tier is weighted when a drop rolls an affix.

    THE TIERS MUST RUN FROM 1 WITH NO GAP. A drop draws from every tier at or
    below the cap its difficulty tier allows, so a missing tier is not a tier
    that never rolls -- it is a tier the draw has no weight for.
    """
    rows = list(book["Affix Tiers"].iter_rows(values_only=True))
    headers = _header_index(rows, "Affix Tiers")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        cell = _cell(raw, headers, "Tier")
        if not cell:
            continue
        tier = int(number(cell, "Tier", index))

        weight = number(_cell(raw, headers, "Drop Weight"), "Drop Weight", index)
        if weight <= 0.0:
            raise DataError(f"Affix Tiers row {index}: T{tier} has a drop "
                            f"weight of {weight:g}, so it can never roll")

        out.append({
            "Name": row_name(f"T{tier}"),
            "Tier": tier,
            "DropWeight": weight,
        })

    if [row["Tier"] for row in out] != list(range(1, len(out) + 1)):
        raise DataError("Affix Tiers must list every tier from 1 upward in "
                        f"order, and it lists {[r['Tier'] for r in out]}")

    for below, above in zip(out[:-1], out[1:], strict=True):
        if above["DropWeight"] > below["DropWeight"]:
            raise DataError(f"Affix Tiers: T{above['Tier']} is above "
                            f"T{below['Tier']} and rolls more often "
                            f"({above['DropWeight']:g} against "
                            f"{below['DropWeight']:g})")

    return unique(out, "Affix Tiers")



def enemy_drops(book) -> list[dict]:
    """What a kill of each enemy rarity drops, and how good it is.

    THE ROW KEY IS THE ENEMY RARITY, spelt exactly as `EnemyRarities.csv` keys
    its own rows, so the two tables join on it. That table comes from the
    simulation's enemy model and this one from the workbook, and
    `validate_enemy_drop_rarities` cross-checks them rather than either trusting
    the other.

    THREE NUMBERS PER RARITY, and each is checked against the rarity below it:

      Gear Drops       never falls. A rarer enemy cannot give less.
      Material Drops   never falls, for the same reason.
      Magic Find       never falls, and a Common enemy adds none at all,
                       because it is the baseline the ladder is read against.

    AN EXPECTED COUNT RATHER THAN A CHANCE for both drop columns. A chance
    cannot exceed one and a Cataclysm Boss drops twelve items; the fractional
    part is rolled as a probability where the drop happens.
    """
    rows = list(book["Enemy Drops"].iter_rows(values_only=True))
    headers = _header_index(rows, "Enemy Drops")

    out = []
    seen: set[str] = set()
    for index, raw in enumerate(rows[1:], start=2):
        rarity = _cell(raw, headers, "Enemy Rarity")
        if not rarity:
            continue
        if rarity in seen:
            raise DataError(f"Enemy Drops row {index}: {rarity} appears twice")
        seen.add(rarity)

        gear = number(_cell(raw, headers, "Gear Drops"), "Gear Drops", index)
        materials = number(_cell(raw, headers, "Material Drops"),
                           "Material Drops", index)
        for what, value in (("gear", gear), ("crafting material", materials)):
            if value <= 0.0:
                raise DataError(
                    f"Enemy Drops row {index}: a {rarity} drops {value:g} "
                    f"{what} items. Loot quantity multiplies this, so no amount "
                    "of it would make a rate of zero produce anything.")

        magic_find = number(_cell(raw, headers, "Magic Find") or 0,
                            "Magic Find", index)
        if magic_find < 0.0:
            raise DataError(
                f"Enemy Drops row {index}: a {rarity} adds {magic_find:g} magic "
                "find. It is an added percentage and cannot be negative.")

        out.append({
            "Name": row_name(rarity),
            "EnemyRarity": rarity,
            "Step": int(number(_cell(raw, headers, "Step"), "Step", index)),
            "GearDrops": gear,
            "MaterialDrops": materials,
            "MagicFind": magic_find,
        })

    if not out:
        raise DataError("Enemy Drops has no rows")

    if out[0]["MagicFind"] != 0.0:
        raise DataError(
            f"Enemy Drops: the weakest rarity, {out[0]['EnemyRarity']}, adds "
            f"{out[0]['MagicFind']:g} magic find rather than none. It is the "
            "baseline the whole ladder is read against.")

    if [row["Step"] for row in out] != list(range(len(out))):
        raise DataError(
            "Enemy Drops: the Step column has to run from 0 upward in row "
            f"order, and it reads {[row['Step'] for row in out]}")

    for below, above in zip(out[:-1], out[1:], strict=True):
        rising = (f"a {above['EnemyRarity']} is rarer than a "
                  f"{below['EnemyRarity']} and")
        for column, what in (("GearDrops", "gear"),
                             ("MaterialDrops", "crafting materials"),
                             ("MagicFind", "magic find")):
            if above[column] < below[column]:
                raise DataError(
                    f"Enemy Drops: {rising} gives less {what} "
                    f"({above[column]:g} against {below[column]:g})")

    return unique(out, "Enemy Drops")


def material_tiers(book) -> list[dict]:
    """How heavily each crafting material tier is weighted on a drop.

    THE TIERS MUST RUN FROM 1 WITH NO GAP, for the same reason the affix tiers
    must: the roll walks every rung, so a missing one is a rung the cascade has
    no weight for rather than one that never comes up.

    THE MATERIALS COLUMN IS A COUNT, not a list. It says how many crafting
    materials share the tier, which is what turns "one drop in 341 is Extremely
    Rare" into "one drop in 1,023 is Purified Essence" -- the figure the weights
    were actually chosen against. `validate_material_tier_counts` checks it
    against the Crafting sheet, where the materials really are.
    """
    rows = list(book["Material Tiers"].iter_rows(values_only=True))
    headers = _header_index(rows, "Material Tiers")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        cell = _cell(raw, headers, "Tier")
        if not cell:
            continue
        tier = int(number(cell, "Tier", index))

        name = _cell(raw, headers, "Tier Name")
        if not name:
            raise DataError(f"Material Tiers row {index}: tier {tier} has no "
                            "name, and the Crafting sheet names its tiers")

        weight = number(_cell(raw, headers, "Drop Weight"), "Drop Weight", index)
        if weight <= 0.0:
            raise DataError(f"Material Tiers row {index}: {name} has a drop "
                            f"weight of {weight:g}, so it can never roll")

        materials = number(_cell(raw, headers, "Materials"), "Materials", index)
        if materials < 1:
            raise DataError(f"Material Tiers row {index}: {name} holds "
                            f"{materials:g} materials, so rolling it would give "
                            "nothing")

        # THE COLOUR THE MATERIAL'S NAME IS DRAWN IN ON THE FLOOR. The five
        # tiers are a different ladder from the eight gear rarities and they
        # share a surface with them, so they get their own hue family rather
        # than borrowing five of the gear colours. Decided on 2026-08-19; the
        # reasoning and the measurements are in docs/DECISIONS.md.
        colour = linear_colour(_cell(raw, headers, "Colour"), "Colour",
                               f"Material Tiers row {index} ({name})")

        out.append({
            "Name": row_name(f"T{tier}"),
            "Tier": tier,
            "TierName": name,
            "DropWeight": weight,
            "Materials": int(materials),
            "Colour": colour,
        })

    if [row["Tier"] for row in out] != list(range(1, len(out) + 1)):
        raise DataError("Material Tiers must list every tier from 1 upward in "
                        f"order, and it lists {[r['Tier'] for r in out]}")

    for below, above in zip(out[:-1], out[1:], strict=True):
        if above["DropWeight"] > below["DropWeight"]:
            raise DataError(
                f"Material Tiers: {above['TierName']} is above "
                f"{below['TierName']} and rolls more often "
                f"({above['DropWeight']:g} against {below['DropWeight']:g})")

    # TWO TIERS SHARING A COLOUR WOULD BE TWO THINGS A PLAYER COULD NOT TELL
    # APART on the floor, the same fault the Gear Rarity sheet refuses.
    seen_colours: dict[str, str] = {}
    for row in out:
        if row["Colour"] in seen_colours:
            raise DataError(
                f"Material Tiers: {row['TierName']} and "
                f"{seen_colours[row['Colour']]} are drawn in the same colour, "
                "so a player could not tell them apart on the floor")
        seen_colours[row["Colour"]] = row["TierName"]

    return unique(out, "Material Tiers")


#: The two slots that are allowed to have no cooldown, and the reason each has
#: none. The Basic Attack is automatic, so the weapon's attack speed sets its
#: rate; the Aura is a toggle, so there is nothing to wait for. Any other slot
#: reading zero is a forgotten number rather than a decision, which is exactly
#: how issue #155 went unnoticed while 41 enchantments scaled it.
SLOTS_WITHOUT_A_COOLDOWN = frozenset({"Basic", "Aura"})


def skill_slots(book) -> list[dict]:
    """What a skill in each of the seven slots is worth, waits, and costs.

    WHY THIS SHEET EXISTS. These numbers used to live only in
    sim/cataclysm_sim/character.py, where Unreal cannot reach them, so no
    ability could honour a cooldown or a mana cost. They are per slot rather
    than per skill: no designed skill states its own, and a column on the
    Weapon Skills sheet would be 77 copies of seven values.

    A skill states its own figure only when it differs, which is what Skull
    Splitter does at 500% weapon damage.
    """
    rows = list(book["Skill Slots"].iter_rows(values_only=True))
    headers = _header_index(rows, "Skill Slots")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        slot = _cell(raw, headers, "Slot")
        if not slot:
            continue

        cooldown = number(_cell(raw, headers, "Cooldown") or 0, "Cooldown", index)
        lowest = number(_cell(raw, headers, "Cooldown Lowest") or 0,
                        "Cooldown Lowest", index)
        highest = number(_cell(raw, headers, "Cooldown Highest") or 0,
                         "Cooldown Highest", index)
        if not lowest <= cooldown <= highest:
            raise DataError(
                f"Skill Slots row {index}: {slot} has a cooldown of {cooldown}s "
                f"outside its own band of {lowest} to {highest}")
        if cooldown == 0.0 and slot not in SLOTS_WITHOUT_A_COOLDOWN:
            raise DataError(
                f"Skill Slots row {index}: {slot} has no cooldown. Only "
                f"{sorted(SLOTS_WITHOUT_A_COOLDOWN)} may have none; the Basic "
                "Attack is automatic and the Aura is a toggle.")

        damage = number(_cell(raw, headers, "Damage Percent") or 0,
                        "Damage Percent", index)
        damage_low = number(_cell(raw, headers, "Damage Lowest") or 0,
                            "Damage Lowest", index)
        damage_high = number(_cell(raw, headers, "Damage Highest") or 0,
                             "Damage Highest", index)
        if not damage_low <= damage <= damage_high:
            raise DataError(
                f"Skill Slots row {index}: {slot} deals {damage}% of weapon "
                f"damage, outside its own band of {damage_low} to {damage_high}")

        out.append({
            "Name": slot,
            "Slot": slot,
            "DamagePercent": damage,
            "DamageLowest": damage_low,
            "DamageHighest": damage_high,
            "Cooldown": cooldown,
            "CooldownLowest": lowest,
            "CooldownHighest": highest,
            "ManaCost": number(_cell(raw, headers, "Mana Cost") or 0,
                               "Mana Cost", index),
            "ManaOnHit": number(_cell(raw, headers, "Mana On Hit") or 0,
                                "Mana On Hit", index),
            "Note": _cell(raw, headers, "Note"),
        })

    return unique(out, "Skill Slots")


#: The row name carrying the stat line every class inherits.
DEFAULT_CLASS_ROW = "Default"


def class_stats(book) -> list[dict]:
    """One (class, stat) pair per row: its level 1 base and its per-level gain.

    THE DEFAULT LINE IS A ROW SET, NOT A SPECIAL CASE. There are 33 stats and 24
    classes planned, so writing every class out in full would be 792 rows of
    which almost all would repeat the same values. A class named "Default"
    carries the shared line and each real class overrides only the stats that
    express its identity, which is also how the design describes a class: as
    much by what it refuses as by what it takes.

    Anything reading this resolves a class's value for a stat by looking for
    that class's row, then the Default row, then zero.
    """
    rows = list(book["Class Stats"].iter_rows(values_only=True))
    headers = _header_index(rows, "Class Stats")

    out = []
    seen: set[tuple[str, str]] = set()
    for index, raw in enumerate(rows[1:], start=2):
        class_name = _cell(raw, headers, "Class")
        if not class_name:
            continue
        stat = _cell(raw, headers, "Stat")
        if not stat:
            raise DataError(f"Class Stats row {index}: {class_name} names no stat")

        key = (class_name, stat)
        if key in seen:
            raise DataError(
                f"Class Stats row {index}: {class_name} sets {stat} twice, so "
                "one of the two is silently ignored")
        seen.add(key)

        out.append({
            "Name": row_name(class_name, stat),
            "ClassName": class_name,
            "Stat": stat,
            "Base": number(_cell(raw, headers, "Base") or 0, "Base", index),
            "PerLevel": number(_cell(raw, headers, "Per Level") or 0,
                               "Per Level", index),
        })

    if not any(r["ClassName"] == DEFAULT_CLASS_ROW for r in out):
        raise DataError(
            f"Class Stats has no {DEFAULT_CLASS_ROW!r} rows, so every class "
            "would start from nothing")
    return unique(out, "Class Stats")


#: The tag each minion family must carry, so a row's family and its tags cannot
#: say two different things.
#:
#: WHICH ATTRIBUTE IS NO LONGER WRITTEN ON THE MINION. Issue #335 settled Spirit
#: for creatures and Agility for machines, and the first version of this table
#: stored that as a `Scaling Attribute` column. That could only ever express one
#: attribute granting one thing. A minion now carries TAGS, and the separate
#: `Minion Scaling` sheet says what each attribute grants to a minion carrying a
#: given tag -- which is the shape Last Epoch uses, where only stats explicitly
#: tagged for minions reach them and the tags are layered rather than one flag.
MINION_FAMILY_TAG = {"Creature": "Minion.Creature", "Machine": "Minion.Machine"}

#: Which enemy a minion picks. The Ballista deliberately targets the furthest,
#: which is the whole reason this is a column rather than one global rule.
MINION_TARGET_MODES = ("Nearest", "Furthest")


def minion_types(book) -> list[dict]:
    """One row per minion type: its own stat block. Issue #336.

    WHY A TABLE OF ITS OWN, rather than numbers in each skill's Shape Params.
    Two skills can produce the same creature -- Summon Imp and Open the Rift both
    make a lesser imp -- and one skill can produce two, as Iron Fortress makes
    ballistae and spike traps. Numbers in the skill row would exist twice for the
    first case and could not be expressed at all for the second. A skill decides
    how many, how often and how long; the type decides what the thing IS.

    HEALTH AND DAMAGE ARE ABSOLUTE, NOT A SHARE OF THE SUMMONER.
    `docs/Cataclysm_GDD_v2.md` line 1588: "Every minion type has its own stats. A
    minion is not a percentage of its summoner." Each is the type's own base
    raised by the summoner's level, in the same `Base` and `Per Level` shape
    `Class Stats` already uses for character stats.

    THE SUMMONER'S LEVEL IS THE ONLY THING THAT RAISES THEM. Gear does not cross
    unless a modifier names minions, so a minion's damage rises more slowly than
    its summoner's, whose gear rises too. That gap is what minion affixes and
    points in the scaling attribute exist to close, and it is intended rather
    than a fitting error.
    """
    rows = list(book["Minion Types"].iter_rows(values_only=True))
    headers = _header_index(rows, "Minion Types")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        name = _cell(raw, headers, "Minion Type")
        if not name:
            continue

        family = _cell(raw, headers, "Family")
        if family not in MINION_FAMILY_TAG:
            raise DataError(
                f"Minion Types row {index}: {name} is family {family!r}; "
                f"expected one of {sorted(MINION_FAMILY_TAG)}")

        written = _cell(raw, headers, "Tags")
        tags = [t.strip() for t in written.split(",") if t.strip()]
        if not tags:
            raise DataError(
                f"Minion Types row {index}: {name} carries no tags, so nothing "
                "can ever scale it")
        required = MINION_FAMILY_TAG[family]
        if required not in tags:
            raise DataError(
                f"Minion Types row {index}: {name} is a {family} and does not "
                f"carry {required}, so no {family} scaling would reach it")
        if "Type.Minion" not in tags:
            raise DataError(
                f"Minion Types row {index}: {name} does not carry Type.Minion")
        for other, tag in MINION_FAMILY_TAG.items():
            if other != family and tag in tags:
                raise DataError(
                    f"Minion Types row {index}: {name} is a {family} and also "
                    f"carries {tag}, so both families' scaling would reach it")

        mode = _cell(raw, headers, "Target Mode")
        if mode not in MINION_TARGET_MODES:
            raise DataError(
                f"Minion Types row {index}: {name} targets {mode!r}; expected "
                f"one of {list(MINION_TARGET_MODES)}")

        interval = number(_cell(raw, headers, "Attack Interval Seconds"),
                          "Attack Interval Seconds", index)
        if interval <= 0:
            raise DataError(
                f"Minion Types row {index}: {name} attacks every {interval}s, "
                "which is never or continuously rather than at a rate")

        health = number(_cell(raw, headers, "Base Health"), "Base Health", index)
        damage = number(_cell(raw, headers, "Base Damage"), "Base Damage", index)
        if health <= 0:
            raise DataError(
                f"Minion Types row {index}: {name} has {health} base health, so "
                "it dies to anything at level 1")
        if damage < 0:
            raise DataError(
                f"Minion Types row {index}: {name} has {damage} base damage")

        threat = number(_cell(raw, headers, "Threat Percent"),
                        "Threat Percent", index)
        if threat < 0:
            raise DataError(
                f"Minion Types row {index}: {name} draws {threat}% attention")

        out.append({
            "Name": name,
            "Family": family,
            "BaseHealth": health,
            "HealthPerLevel": number(_cell(raw, headers, "Health Per Level"),
                                     "Health Per Level", index),
            "BaseDamage": damage,
            "DamagePerLevel": number(_cell(raw, headers, "Damage Per Level"),
                                     "Damage Per Level", index),
            "AttackIntervalSeconds": interval,
            "MoveSpeed": number(_cell(raw, headers, "Move Speed"),
                                "Move Speed", index),
            "ThreatPercent": threat,
            "ReachCm": number(_cell(raw, headers, "Reach Cm"), "Reach Cm", index),
            "NoticeRadiusCm": number(_cell(raw, headers, "Notice Radius Cm"),
                                     "Notice Radius Cm", index),
            "TargetMode": mode,
            "Tags": ", ".join(tags),
        })

    if not out:
        raise DataError("Minion Types has no rows, so no skill can summon "
                        "anything with stats")
    return unique(out, "Minion Types")


#: Which minion stats an attribute may grant. NOT the character sheet: these are
#: the minion's own stats, held in `Minion Types`, and nothing here reaches the
#: summoner.
MINION_SCALABLE_STATS = ("damage", "health")


def minion_scaling(book) -> list[dict]:
    """What one point of an attribute grants a minion carrying a tag.

    WHY THIS IS NOT ROWS IN `Attributes`. That table is (attribute, stat, percent
    per point) with no tag, and everything reading it sums every attribute that
    names a stat. One shared "minion damage" entry would therefore let a
    summoner's Agility raise an imp, which issue #335 settled it must not: the
    attribute is declared per minion type. A tag is what makes the scoping real.

    IT IS THE SHAPE `character.Modifier` ALREADY USES -- a stat, an amount and
    the tags it requires -- which the project owner stated on 2026-08-02: "The
    player holds all of its own increases, and those increases apply to things
    with matching tags." This is that rule pointed at minions, so the four minion
    affixes in issue #337 need no new machinery.

    ONLY DAMAGE IS FILLED IN, because only damage is decided. The design document
    states "Each grants 1.0% increased minion damage per point". Health is
    expressible here and nobody has chosen a figure for it.
    """
    rows = list(book["Minion Scaling"].iter_rows(values_only=True))
    headers = _header_index(rows, "Minion Scaling")

    out = []
    seen: set[tuple[str, str, str]] = set()
    for index, raw in enumerate(rows[1:], start=2):
        attribute = _cell(raw, headers, "Attribute")
        if not attribute:
            continue

        tag = _cell(raw, headers, "Requires Tag")
        if not tag:
            raise DataError(
                f"Minion Scaling row {index}: {attribute} requires no tag, so "
                "it would reach every minion of every family")

        stat = _cell(raw, headers, "Stat")
        if stat not in MINION_SCALABLE_STATS:
            raise DataError(
                f"Minion Scaling row {index}: {attribute} grants {stat!r}; "
                f"expected one of {list(MINION_SCALABLE_STATS)}")

        key = (attribute, tag, stat)
        if key in seen:
            raise DataError(
                f"Minion Scaling row {index}: {attribute} sets {stat} for {tag} "
                "twice, so one of the two is silently ignored")
        seen.add(key)

        value = number(_cell(raw, headers, "Percent Per Point"),
                       "Percent Per Point", index)
        if value <= 0:
            raise DataError(
                f"Minion Scaling row {index}: {attribute} gives {value} percent "
                f"of minion {stat} per point, so the point is wasted")

        out.append({
            "Name": row_name(attribute, tag, stat),
            "Attribute": attribute,
            "RequiresTag": tag,
            "Stat": stat,
            "PercentPerPoint": value,
        })

    if not out:
        raise DataError("Minion Scaling has no rows, so no attribute raises a "
                        "minion at all")
    return unique(out, "Minion Scaling")


def attributes(book) -> list[dict]:
    """What one point of an attribute is worth, one stat per row.

    ATTRIBUTES ONLY EVER SCALE. A point adds to a stat's sum of increases and
    the sum multiplies the base, so an attribute grants nothing on a stat with
    no base. That is the design working rather than failing.

    Values are PERCENT PER POINT: Vitality reads 2, meaning 2% maximum health
    per point. The simulation stores the same figure as a fraction.
    """
    rows = list(book["Attributes"].iter_rows(values_only=True))
    headers = _header_index(rows, "Attributes")

    out = []
    seen: set[tuple[str, str]] = set()
    for index, raw in enumerate(rows[1:], start=2):
        attribute = _cell(raw, headers, "Attribute")
        if not attribute:
            continue
        stat = _cell(raw, headers, "Stat")
        if not stat:
            raise DataError(f"Attributes row {index}: {attribute} names no stat")

        key = (attribute, stat)
        if key in seen:
            raise DataError(
                f"Attributes row {index}: {attribute} sets {stat} twice")
        seen.add(key)

        value = number(_cell(raw, headers, "Percent Per Point"),
                       "Percent Per Point", index)
        if value <= 0:
            raise DataError(
                f"Attributes row {index}: {attribute} gives {value} percent of "
                f"{stat} per point, so the point is wasted")

        out.append({
            "Name": row_name(attribute, stat),
            "Attribute": attribute,
            "Stat": stat,
            "PercentPerPoint": value,
        })
    return unique(out, "Attributes")


#: The prefix every damage type's tag carries. The eight are declared on the
#: Tags sheet and generated into game/Config/Tags/CataclysmTags.ini.
ELEMENT_TAG_PREFIX = "Element."

#: The three material scales, and what one of them being 1.0 means: leave the
#: Niagara system's own authored value alone. None of them may be zero or
#: negative -- see the check in `element_visuals`.
ELEMENT_SCALES = ("EmissiveMultiplier", "SpawnRateScale", "VelocityScale")


def srgb_to_linear(channel: int) -> float:
    """One 0-255 sRGB channel as linear 0-1, the same curve the engine uses.

    The same function as `srgb_to_linear` in tools/generate_telegraph_material.py
    and as `FLinearColor::FromSRGBColor`, which is what
    ACataclysmTelegraphMarker::ResolveColour calls on the design document's hex.
    """
    c = channel / 255.0
    if c <= 0.04045:
        return c / 12.92
    return ((c + 0.055) / 1.055) ** 2.4


def linear_colour(value, field: str, where: str) -> str:
    """A design document sRGB hex as the text Unreal imports into FLinearColor.

    THE DESIGN STATES sRGB AND THE ENGINE WANTS LINEAR. Section XIII writes
    `#FF7A2E` because that is what a colour picker shows; an FLinearColor is
    linear, so a material fed the raw 0-255 figures divided by 255 renders a
    visibly different colour from the one that was designed. The conversion
    happens here so the workbook keeps the notation a person can copy straight
    out of the design document.

    THE LENGTH IS NOT ENOUGH ON ITS OWN. That was a real bug in this project:
    `FColor::FromHex` does not report bad input, and the word "nonsense" is
    eight characters, so a length-only check accepted it and produced a colour
    nobody asked for. Every character is checked here for the same reason.
    """
    text = clean(value).lstrip("#")
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", text):
        raise DataError(
            f"{where}: {field} is {clean(value)!r}, which is not six hex digits "
            "with an optional leading hash. Section XIII of "
            "docs/Cataclysm_GDD_v2.md states every one of these as sRGB hex.")

    channels = (int(text[i:i + 2], 16) for i in (0, 2, 4))
    red, green, blue = (srgb_to_linear(c) for c in channels)
    return (f"(R={red:.6f},G={green:.6f},B={blue:.6f},A=1.000000)")


def element_visuals(book) -> list[dict]:
    """What each damage type's effects look like. Source: Element Visuals.

    WHY THIS SHEET EXISTS. Issue #549. The eight damage types have an effect
    palette, settled by the project owner and recorded in section XIII of
    docs/Cataclysm_GDD_v2.md, and nothing in the game could read it. Section 5
    of docs/Niagara_Conventions.md specifies this table: eight rows against the
    existing tags, so one Niagara system serves all eight damage types instead
    of eight copies of it. Eight shapes times eight damage types is 64 assets
    built the wrong way and 8 assets plus 8 rows built this one.

    THE ROW KEY IS THE TAG'S LEAF, so `Element.Demonic` keys the row `Demonic`
    and anything holding a tag can reach its row without a lookup table.

    BRIGHTNESS IS NOT CHECKED HERE, deliberately, and this is the one thing
    about this table that is easy to get wrong twice. The design's readability
    rule -- a world surface may not exceed 30% brightness, an effect's primary
    may not fall below 60% -- is already enforced against the design document by
    tools/tests/test_effect_palette_stays_readable.py, and this table is pinned
    to that same document by
    tools/tests/test_element_visuals_match_the_design.py. So the rule reaches
    this table through those two, and a second copy of the CIE lightness formula
    here would be a second thing to keep in step for no extra coverage.
    """
    rows = list(book["Element Visuals"].iter_rows(values_only=True))
    headers = _header_index(rows, "Element Visuals")

    out = []
    for index, raw in enumerate(rows[1:], start=2):
        tag = _cell(raw, headers, "Element Tag")
        if not tag:
            continue

        where = f"Element Visuals row {index} ({tag})"
        if not tag.startswith(ELEMENT_TAG_PREFIX):
            raise DataError(
                f"{where}: the key is {tag!r}, and this table is keyed by a "
                f"damage type's tag, so it must start with "
                f"{ELEMENT_TAG_PREFIX!r}.")

        entry = {
            "Name": row_name(tag[len(ELEMENT_TAG_PREFIX):]),
            "ElementTag": tag,
            "PrimaryColour": linear_colour(
                _cell(raw, headers, "Primary"), "Primary", where),
            "SecondaryColour": linear_colour(
                _cell(raw, headers, "Secondary"), "Secondary", where),
        }

        # A SCALE OF ZERO IS A FORGOTTEN NUMBER, NOT A DECISION, and each of the
        # three fails in a way that looks like a broken effect rather than like
        # bad data: no particles at all, particles that never move, or an effect
        # rendered black. That is the shape of issue #155, where a cooldown of
        # zero went unnoticed across 77 skills because zero read as a value.
        for column, field in zip(("Emissive Multiplier", "Spawn Rate Scale",
                                  "Velocity Scale"), ELEMENT_SCALES,
                                 strict=True):
            scale = number(_cell(raw, headers, column), column, index)
            if scale <= 0:
                raise DataError(
                    f"{where}: {column} is {scale}. A scale of zero or less "
                    "makes the effect invisible rather than neutral; 1.0 is "
                    "what leaves the Niagara system's own value alone.")
            entry[field] = scale

        out.append(entry)

    return unique(out, "Element Visuals")


# --------------------------------------------------------------------------
# the enemy tables, which come from the Python model rather than the workbook

def _enemy_stats():
    """Import `cataclysm_sim.enemy_stats`, adding sim/ to the path if needed.

    Imported here rather than at the top of the file so that generating the
    workbook tables does not depend on the simulation package being importable.

    IMPORTING IT RUNS ITS OWN CHECKS. enemy_stats.py calls five self-checks at
    module scope: every archetype deals a real damage type, none declares more
    resistance than the 70% cap is worth, every body has a width, every creature
    can turn, and no creature's defensive layers combine past what a geared
    player stops. So a model that contradicts itself raises here rather than
    being written out to a CSV.
    """
    if str(SIM_ROOT) not in sys.path:
        sys.path.insert(0, str(SIM_ROOT))
    try:
        from cataclysm_sim import enemy_stats
    except ImportError as error:
        raise DataError(
            f"could not import cataclysm_sim.enemy_stats from {SIM_ROOT}, which "
            f"is where the enemy stat block is designed: {error}") from error
    return enemy_stats


def _six_places(value: float) -> float:
    """Round to six decimal places, so the written table stays readable.

    Python writes the shortest string that reproduces a float exactly, and for a
    rarity multiplier like 0.5 * 1.85 ** 3 that string is 3.1655562500000005.
    Six places is a ten-millionth of a Common enemy's health multiplier, far
    below anything a balance judgement can see, and it keeps the table legible
    to the person reading the diff.

    The drift guard does not depend on this: `--check` rebuilds the whole table
    from the model and compares the text, so changing a model constant makes the
    file stale whatever precision it is written at.
    """
    return round(value, 6)


def enemy_archetypes(_book=None) -> list[dict]:
    """One row per enemy archetype: what KIND of creature it is.

    Source: `ARCHETYPES` in sim/cataclysm_sim/enemy_stats.py. One column per
    field of the `Archetype` dataclass, and no derived values, because deriving
    one here would put a second reading of the design in this file.

    DISTANCES AND SPEEDS ARE IN METRES, as the model states them and as
    game/Data/ClassStats.csv already states the player's movement speed. The
    engine multiplies by `CentimetresPerMetre` when it reads one, the way
    ACataclysmPlayerCharacter::ApplyMovementSpeed does.

    A CHASE SPEED OF 0 IS A SENTINEL, not a creature that cannot move: the model
    defines it as "use MoveSpeed in both states", which is what every enemy but
    the Brute does. It is written out unchanged rather than expanded here,
    because expanding it would mean this file deciding what the model meant.

    THE BASELINE ROW IS NOT A CREATURE. It is in `ARCHETYPES` so the rarity
    ladder can be read without an archetype's multipliers on top, and it is
    written out because this table is the model and an exception here could hide
    a real archetype going missing. Its Role column says so in words.
    """
    model = _enemy_stats()

    out = []
    for kind in model.ARCHETYPES.values():
        out.append({
            "Name": row_name(kind.name),
            "ArchetypeName": kind.name,
            "Role": kind.role,
            "Cataclysm": kind.cataclysm,
            "HealthShare": _six_places(kind.health_share),
            "DamageShare": _six_places(kind.damage_share),
            "ArmorShare": _six_places(kind.armor_share),
            "AttackIntervalSeconds": _six_places(kind.attack_interval),
            "CritChancePercent": _six_places(kind.crit_chance),
            "CritMultiplierPercent": _six_places(kind.crit_multiplier),
            "MoveSpeedMetresPerSecond": _six_places(kind.move_speed),
            "ChaseSpeedMetresPerSecond": _six_places(kind.chase_speed),
            "EvasionPercent": _six_places(kind.evasion),
            "EnergyShieldFraction": _six_places(kind.energy_shield_fraction),
            "BodyRadiusMetres": _six_places(kind.body_radius),
            "TurnRateDegreesPerSecond": _six_places(kind.turn_rate_degrees),
            "ResistancePercent": _six_places(kind.resistance),
        })

    if not out:
        raise DataError("enemy_stats.ARCHETYPES is empty")
    return unique(out, "Enemy Archetypes")


def enemy_rarities(_book=None) -> list[dict]:
    """One row per rarity: how much of an encounter's score each stat is worth.

    Source: `RARITY_ORDER` and the six AT_COMMON and PER_STEP constants in
    sim/cataclysm_sim/enemy_stats.py.

    RARITY SCALES MAGNITUDE AND NOTHING ELSE. Attack interval, criticals,
    movement and resistance belong to the archetype, so they are not here: a
    Legendary Imp is a bigger Imp, not a slower or tougher kind of creature.
    The model's own header explains why it used to be the other way round.

    WHAT A COLUMN MEANS. `stats_for` computes each of the three as

        score * <stat>PerScore * archetype.<stat>Share

    so the multiplier here is already raised to the power of the rarity's step.
    Written expanded rather than as a base and a growth rate so that reading an
    enemy's stats is a table lookup and a multiply, with no exponent at runtime,
    and so the ladder can be read straight off the file: a Cataclysm Boss has
    about 21 times a Common enemy's health.

    Health is additionally floored at 1 by `stats_for`, and energy shield is a
    fraction of health set by the archetype. Neither is a rarity figure.
    """
    model = _enemy_stats()

    # Read directly rather than through getattr with a default: a missing name
    # must raise here. A check that quietly passes when the thing it reads has
    # been renamed is worse than no check, because it reads as coverage.
    weights = model.scoring.RARITY_WEIGHTS
    if list(weights) != list(model.RARITY_ORDER):
        raise DataError(
            "enemy_stats.RARITY_ORDER is "
            f"{list(model.RARITY_ORDER)} but scoring.RARITY_WEIGHTS, which its "
            f"comment calls the authoritative list, is {list(weights)}. "
            "scoring.py is a port that has drifted from its source twice, so "
            "the two lists must be compared rather than assumed equal.")

    out = []
    for rarity in model.RARITY_ORDER:
        step = model.rarity_step(rarity)
        out.append({
            "Name": row_name(rarity),
            "RarityName": rarity,
            "Step": step,
            "HealthPerScore": _six_places(
                model.HEALTH_AT_COMMON * model.HEALTH_PER_STEP ** step),
            "DamagePerScore": _six_places(
                model.DAMAGE_AT_COMMON * model.DAMAGE_PER_STEP ** step),
            "ArmorPerScore": _six_places(
                model.ARMOR_AT_COMMON * model.ARMOR_PER_STEP ** step),
            "SpawnWeight": _six_places(model.spawn_weight(rarity)),
            "BodyScale": _six_places(
                model.BODY_SCALE_AT_COMMON
                * model.BODY_SCALE_PER_STEP ** step),
        })

    if not out:
        raise DataError("enemy_stats.RARITY_ORDER is empty")

    # THE SPAWN WEIGHTS HAVE TO ADD UP TO ONE FLOOR'S WORTH, because they are
    # shares of a floor's population rather than independent chances. The model
    # reads them out of scoring.DUNGEON_SCORE_MIX, and that file is a port which
    # has drifted from its source twice, so this is checked rather than assumed.
    total = sum(row["SpawnWeight"] for row in out)
    if abs(total - 1.0) > 1e-6:
        raise DataError(
            f"the enemy rarity spawn weights sum to {total}, not 1. They are "
            "how common each rarity is on a floor, which the Dungeon Score "
            "Formula section of docs/Cataclysm_GDD_v2.md states outright, and "
            "shares that do not add up leave part of every floor unfilled.")

    return unique(out, "Enemy Rarities")



# --------------------------------------------------------------------------
# The class passive trees, from the node graphs in docs/
#
# A THIRD SOURCE, BESIDE THE WORKBOOK AND THE PYTHON MODEL. The four class trees
# are authored in C:\Projects\PassiveTreeCreator, a separate tool, and exported
# as JSON into docs/. Those files are the design; this turns them into something
# the game can read, in the same way the workbook sheets are turned into the
# other tables.
#
# WHY THE GAME CANNOT READ THE JSON DIRECTLY. `docs/` is not packaged and is not
# a content directory, so nothing in a built game can open it. Every other piece
# of design data reaches the engine as a DataTable and these do the same.
#
# WHAT IS DELIBERATELY NOT CARRIED ACROSS: the colour of a node, the viewport the
# editor was last looking at, and the free-floating text labels the tool lets an
# author place. All three are about the authoring tool's own canvas rather than
# about the tree, and a screen that drew them would be drawing the editor.
# --------------------------------------------------------------------------

#: The four class trees that exist. The other twenty are issue #24.
CLASS_TREES = ("Berserker", "Bulwark", "Saboteur", "Masochist")

#: How many options a capstone offers, from the design document: "Player chooses
#: one of three options per tier." Fixed at three, which is why the options are
#: three pairs of columns rather than a second table.
CAPSTONE_OPTIONS = 3


def _tree_file(name: str) -> pathlib.Path:
    return REPO_ROOT / "docs" / f"{name}_Class_Tree_Final.json"


def _load_tree(name: str) -> dict:
    path = _tree_file(name)
    if not path.is_file():
        raise DataError(f"{path.name} is not present")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as error:
        raise DataError(f"{path.name} is not valid JSON: {error}") from error


def _capstone_threshold(node: dict, tree: str) -> int:
    """How many points must be spent in the tree before this capstone opens.

    THREE PLACES TO LOOK, AND ALL THREE ARE IN USE. The Masochist tree writes it
    as `threshold`, the Saboteur tree as `pointThreshold`, and the Berserker and
    Bulwark trees carry both. Every one of the four also states it in the
    description text. Issue #935 asks for them to be made to agree; until then
    anything reading these files has to try all three, and reading only one of
    them would silently give a capstone a threshold of zero.
    """
    data = node["data"]
    for field in ("threshold", "pointThreshold"):
        value = data.get(field)
        if isinstance(value, (int, float)):
            return int(value)

    matched = re.search(r"(\d+)\s+points", data.get("description", ""))
    if not matched:
        raise DataError(
            f"{tree}: capstone {data.get('name')!r} states no point threshold, "
            f"in a threshold field or in its description")
    return int(matched.group(1))



def _no_duplicates(rows: list[dict], table: str) -> list[dict]:
    """Every row name is distinct, or say which one is not.

    NOT `unique`, WHICH SUFFIXES A DUPLICATE AND SAYS NOTHING. That is the right
    answer for a workbook sheet where two rows can legitimately share a display
    name. These keys are built here out of a tree name and a node identifier, so
    a duplicate means this file built the same key twice and the game would hold
    one node where the design has two.
    """
    seen: set[str] = set()
    for row in rows:
        if row["Name"] in seen:
            raise DataError(f"{table}: two rows are both called {row['Name']!r}")
        seen.add(row["Name"])
    return rows


def passive_nodes(_book=None) -> list[dict]:
    """One row per node, across all four class trees.

    THE ROW KEY IS THE TREE AND THE NODE IDENTIFIER TOGETHER, because node
    identifiers are only unique within a tree. Fourteen are shared by more than
    one of the four -- `capstone_25` is in all of them -- so keying on the node
    identifier alone would silently merge them and the game would hold one tree's
    worth of capstones for four trees.
    """
    out = []
    for tree in CLASS_TREES:
        data = _load_tree(tree)
        for node in data.get("nodes", []):
            body = node.get("data", {})
            kind = node.get("type", "")
            if kind not in ("basic", "keystone", "capstone"):
                raise DataError(f"{tree}: node {node.get('id')} has kind "
                                f"{kind!r}, which is not one of the three")

            position = node.get("position", {})
            options = body.get("options", []) if kind == "capstone" else []

            row = {
                "Name": f"{tree}_{node['id']}",
                "Tree": tree,
                "NodeId": node["id"],
                "Kind": kind,
                "NodeName": clean(body.get("name", "")),
                "Description": clean(body.get("description", "")),
                "MaxPoints": int(body.get("maxPoints", 0)),
                # ZERO FOR EVERYTHING THAT IS NOT A CAPSTONE, which is not a
                # missing value: a basic node and a keystone open when their
                # edges allow, not at a number of points spent in the tree.
                "Threshold": (_capstone_threshold(node, tree)
                              if kind == "capstone" else 0),
                "PositionX": float(position.get("x", 0.0)),
                "PositionY": float(position.get("y", 0.0)),
            }

            # THE SABOTEUR'S FOUR CAPSTONES HAVE NONE, and its own descriptions
            # say to choose one of three. That is issue #935. Empty columns
            # carry the gap into the game rather than hiding it, and the screen
            # says a capstone with no options cannot be taken.
            for index in range(CAPSTONE_OPTIONS):
                option = options[index] if index < len(options) else {}
                row[f"Option{index + 1}Name"] = clean(option.get("name", ""))
                row[f"Option{index + 1}Description"] = clean(
                    option.get("description", ""))

            out.append(row)

    if not out:
        raise DataError("no class tree nodes were read")
    return _no_duplicates(out, "Passive Nodes")


def passive_edges(_book=None) -> list[dict]:
    """One row per dependency edge, across all four class trees.

    AN EDGE IS A REQUIREMENT AND NOT A LINE. `requiredPoints` on an edge from A
    to B means B cannot be taken until A holds that many points. The positions
    of the two nodes are in the node table; nothing here is about drawing.
    """
    out = []
    for tree in CLASS_TREES:
        data = _load_tree(tree)
        node_ids = {node["id"] for node in data.get("nodes", [])}

        for edge in data.get("edges", []):
            source = edge.get("source")
            target = edge.get("target")
            if source not in node_ids or target not in node_ids:
                raise DataError(
                    f"{tree}: edge {edge.get('id')} joins {source} to {target} "
                    f"and one of them is not a node in that tree")

            out.append({
                "Name": f"{tree}_{edge['id']}",
                "Tree": tree,
                "Source": f"{tree}_{source}",
                "Target": f"{tree}_{target}",
                "RequiredPoints": int(edge.get("data", {})
                                      .get("requiredPoints", 0)),
            })

    if not out:
        raise DataError("no class tree edges were read")
    return _no_duplicates(out, "Passive Edges")


#: The states of the character a passive bonus may be made to depend on.
#:
#: HELD HERE AS WELL AS IN C++ ON PURPOSE, and the two have to agree.
#: `UCataclysmPassiveTree::AccumulateInto` turns each of these names into an
#: `ECataclysmStatCondition`, and a name it does not recognise is applied with no
#: condition -- a bonus that holds all the time instead of some of the time,
#: silently and in the player's favour. Refusing an unknown name when the file is
#: written is the only place that can be caught. Issue #959.
#:
#: EACH CARRIES ITS OWN UNITS AND ITS OWN RANGE, since issue #962 added a window
#: measured in seconds beside a threshold measured in percent. One shared range of
#: 0 to 100 would have refused a window longer than a minute and, worse, accepted
#: a percentage written where seconds belong: `seconds_after_health_cost` with 35
#: is a thirty-five second window, which is not what anybody meant and which
#: nothing at run time would report.
#:
#: The value is (lowest, highest, what the number means), or `None` for a
#: condition that compares nothing and therefore takes no value at all.
#:
#: A CONDITION WITH NO NUMBER IS A REAL SHAPE, since issue #962. Four of these
#: compare a reading against a threshold; `while_bleeding` names a kind of effect
#: and the kind is in the name. There is no column that could hold it -- the
#: value column is a float and a tag name is not a number -- so the name carries
#: it, and a value typed beside it is refused rather than quietly ignored. A
#: range of (0, 0) would have accepted a stray zero and read as a threshold of
#: nothing, which is why this is `None` and not a degenerate range.
#: How many options a capstone node offers. Issue #1029.
#:
#: HELD HERE AS WELL AS IN C++ AND THE TWO HAVE TO AGREE.
#: `UCataclysmPassiveTree::CapstoneOptions` is the other copy, and the passive
#: node sheet's own shape is a third statement of it: it carries `Option1Name`
#: through `Option3Name` and no fourth.
#: `test_the_capstone_option_count_matches_the_engine` holds them together.
CAPSTONE_OPTIONS = 3

CONDITIONS = {
    # "While at or below 20% health" is `health_at_or_below` with 20.
    "health_at_or_below": (0.0, 100.0, "a percentage of maximum health"),

    # "While below 20% health" is `health_below` with 20. Issue #1051.
    #
    # A SECOND HEALTH THRESHOLD BECAUSE THE TWO REALLY DIFFER, at exactly the
    # threshold. Seven nodes across the four trees say "at or below" and take
    # the predicate above; The Final Vow's first option, The Last Drop, is the
    # only one in the game that states a health threshold as a STATE and words
    # it "below". `skill_health_cost_above` is the precedent: it exists because
    # "above" genuinely differs from "at or below", and its comment records a
    # character sitting exactly on that threshold correctly getting nothing.
    #
    # NOT SOLVED BY REWORDING THE NODE. The project owner rewrote all twelve
    # Masochist capstone options on 2026-08-27 under issue #1031 and this is
    # their text. `CONDITION_WORDS` in the test beside this file is what refuses
    # a row whose predicate and whose sentence disagree, and widening that entry
    # to accept both wordings would have let a node saying "below" be delivered
    # as "at or below", which is the drift the check exists to catch.
    #
    # "DROPPING BELOW" IS NEITHER OF THESE. That is an EVENT, which The Breaking
    # Point and Rock Bottom both state, and an event is not a condition on a
    # modifier at all: `UCataclysmDamageConversion` holds its threshold as a
    # constant and says why.
    "health_below": (0.0, 100.0, "a percentage of maximum health"),

    # "while you are above 50% health" is `health_above` with 50. Issue #1070.
    #
    # THE THIRD HEALTH PREDICATE AND THE FIRST THAT POINTS UPWARDS. The two
    # above both ask whether health has fallen far enough; Ceaseless Penance is
    # the only node in the game that asks whether it is still high, and it is a
    # STATE rather than an event: a character that heals back above the
    # threshold is holding its debuffs again, and one that falls below it is not.
    #
    # NOT `health_at_or_below` READ BACKWARDS. Strictly above 50 and at or below
    # 50 are complements, so the two cover every character between them, but a
    # row carries one predicate and the stat pipeline has no "not". A node
    # wanting "above" needs a predicate that says so.
    #
    # STRICTLY ABOVE, BECAUSE THE OPTION WRITES "above". The same boundary
    # `skill_health_cost_above` draws and for the same reason: a character
    # sitting exactly on half health is not above it, so its debuffs expire.
    "health_above": (0.0, 100.0, "a percentage of maximum health"),

    # "for 2 seconds after you pay a health cost" is
    # `seconds_after_health_cost` with 2. The upper bound is a sanity limit
    # rather than a design rule: the design's longest window is 5 seconds, and
    # anything past a minute is far likelier to be a percentage in the wrong
    # column than a deliberate window.
    "seconds_after_health_cost": (0.0, 60.0, "a number of seconds"),

    # "for 5 seconds after you take damage of a Cataclysm type other than
    # Demonic" is `seconds_after_foreign_damage` with 5. Same units and
    # same bound as the window above; a different event opens it.
    "seconds_after_foreign_damage": (0.0, 60.0, "a number of seconds"),

    # "a skill whose health cost is above 10% of your maximum health" is
    # `skill_health_cost_above` with 10. Issue #983.
    #
    # THE ONLY ONE OF THESE THAT ASKS ABOUT THE SKILL RATHER THAN THE CHARACTER,
    # so one character using two skills an instant apart answers it differently
    # for each. Nothing the other three can express covers that.
    #
    # STRICTLY ABOVE, NOT AT OR BELOW, unlike `health_at_or_below`, and the
    # difference is reachable rather than pedantic: the Deeper Cuts node at its
    # full ten points adds exactly 10% of maximum health to every skill, so a
    # character with that node and no skill of its own cost sits precisely on
    # this threshold and correctly gets nothing.
    "skill_health_cost_above": (0.0, 100.0, "a percentage of maximum health"),

    # "While you are Bleeding" is `while_bleeding`, and it takes no value.
    # Issue #962. Thirst for Pain is the node.
    #
    # THE FIRST CONDITION THAT ASKS WHAT THE CHARACTER IS CARRYING rather than
    # where one of its own numbers stands, and the first that compares nothing.
    # `UCataclysmDebuffs::IsBleeding` is what answers it.
    #
    # ONE NAME PER KIND OF EFFECT, rather than one name and a column saying
    # which. The same argument the three stack scales below make: a further
    # column here is a row struct change, so a build has to happen before the
    # DataTable asset can be regenerated. This one could not use a column
    # anyway, because the value column holds a number.
    "while_bleeding": None,

    # "While your Fervour is at maximum" is `class_resource_at_maximum`, and it
    # takes no value either. Issue #1026. Communion of Pain is the node.
    #
    # NOT A THRESHOLD WITH THE VALUE SET TO A HUNDRED. A threshold has to be
    # either points or a percentage of the maximum, and the two disagree: the
    # Ritualist's `class_resource` is 150 where every other class's is 100. "At
    # maximum" is the top of whatever bar the class has and is neither.
    #
    # A FUTURE "while above 75 Fervour" WANTS ITS OWN NAME, for the same reason.
    # That one is a points threshold and this one is not, so reusing this with a
    # value would give a Ritualist the wrong answer.
    "class_resource_at_maximum": None,
}

#: The states a passive bonus's SIZE may grow with. Issue #968.
#:
#: A DIFFERENT QUESTION FROM `CONDITIONS` ABOVE. A condition decides whether the
#: bonus applies; a scale decides how large it is when it does. A row may carry
#: both, and neither implies anything about the other.
#:
#: HELD HERE AS WELL AS IN C++ ON PURPOSE, and the two have to agree.
#: `UCataclysmPassiveTree::AccumulateInto` turns each of these names into an
#: `ECataclysmStatScale`. A name it does not recognise is made worth NOTHING
#: rather than worth its full value, because the full value of a scaling bonus is
#: what a character in the most extreme state would get -- for Vicious Onslaught,
#: the bonus for being at death's door handed to one at full health. Refusing an
#: unknown name when the file is written is where it should be caught instead.
#:
#: The value is (lowest step, highest step, what the step means).
SCALES = {
    # "for every 5% of your maximum health that is missing" is `health_missing`
    # with a step of 5. Steps are counted whole and rounded down, so a character
    # 12% below full health has two of them.
    "health_missing": (0.0, 100.0, "a percentage of maximum health"),

    # "for each point of Fervour you currently hold" is `class_resource_held`
    # with a step of 1. Issue #980. Steps are counted whole and rounded down the
    # same way, which cannot show at a step of one.
    #
    # NAMED FOR THE SHARED POOL RATHER THAN FOR FERVOUR. There is one class
    # resource and every class has it; Fervour is only the Masochist's name for
    # it, and a name from one tree would have to change the first time another
    # tree used the same shape. The upper bound is the largest maximum any class
    # line gives the pool.
    "class_resource_held": (0.0, 100.0, "a number of points of the pool"),

    # "for every 5% of your maximum health you currently owe" is `health_owed`
    # with a step of 5. Issue #994. The Reckoning reads the same state with a
    # step of 2. Steps are counted whole and rounded down, the same rule.
    #
    # OWED IS NOT MISSING, and the two are independent readings of health. A
    # character that deferred a cost owes health it is still standing on, so it
    # can be at full health and owe a fifth of it; one that paid the same cost
    # outright is a fifth down and owes nothing. Each has its own node.
    #
    # THE UPPER BOUND IS A STEP OF THE WHOLE POOL. What is OWED may pass a
    # character's maximum health -- that is what The Reckoning kills them for --
    # but a step larger than the pool would be a bonus nothing could ever reach.
    "health_owed": (0.0, 100.0, "a percentage of maximum health"),

    # "for every 1% of life leech you have" is `life_leech` with a step of 1.
    # Issue #1045. Glutton is the node and the only one.
    #
    # A READING OF A STAT AND NOT OF A STATE, which is what makes it unlike
    # every scale above. Those read where a character's health is or how full
    # its pool is, and change from moment to moment; life leech changes when the
    # character's gear or passive points change and not otherwise. So this is a
    # bonus that grows with an investment rather than with a situation.
    #
    # THE UPPER BOUND IS A STEP OF A HUNDRED PERCENT, which is a judgement and
    # not a figure read off anything: nothing in the design states a ceiling on
    # life leech. A hundred is the largest step that could still pay out for a
    # character that leeched every point of damage it dealt, and a step above
    # that would be a bonus nobody could ever reach.
    "life_leech": (0.0, 100.0, "a percentage of life leech"),

    # THREE COUNTS OF STACKS, one per kind. Issues #1002, #1003 and #1004.
    # "Each stack gives +1% increased attack speed per point" is
    # `momentum_stacks` with a step of 1.
    #
    # THREE NAMES RATHER THAN ONE NAME AND A COLUMN SAYING WHICH KIND. A fourth
    # column on this sheet is a row struct change, which means a build before
    # the DataTable asset can be regenerated and a column list to move; three
    # names cost three lines each and read the way the scales above do.
    #
    # THEY ARE NOT INTERCHANGEABLE. Each kind is granted by a different event
    # and lasts a different length of time -- 3, 5 and 8 seconds -- so a row
    # naming the wrong one would count somebody else's stacks and no arithmetic
    # would report it.
    #
    # THE UPPER BOUND IS THE LARGEST CAP ANY KIND HAS. A step larger than the
    # cap would be a bonus no character could ever reach.
    "momentum_stacks": (0.0, 10.0, "a number of stacks"),
    "bloodlust_stacks": (0.0, 10.0, "a number of stacks"),
    "carnage_stacks": (0.0, 10.0, "a number of stacks"),

    # "for each unique debuff on you" is `debuffs_carried` with a step of 1.
    # Issue #962. Four nodes read it and all four count single debuffs.
    #
    # NOT A FOURTH STACK COUNT, THOUGH IT IS COUNTED THE SAME WAY. A stack is
    # granted by an event this project chose to remember and expires on a timer
    # this project chose. A debuff is a gameplay effect somebody applied, and
    # the ability system is already holding the list for its own reasons.
    # `UCataclysmDebuffs` says which of those effects are harmful.
    #
    # THE UPPER BOUND IS A JUDGEMENT AND NOT A DESIGN RULE. Nothing caps how
    # many debuffs a character may carry -- issue #962's own notes say the
    # Vessel of Plagues keystone would need such a cap and there is none -- so
    # this is a sanity limit in the shape the stack counts use. Ten distinct
    # harmful effects at once is already far past anything the game can apply.
    "debuffs_carried": (0.0, 10.0, "a number of debuffs"),
}


def passive_effects(book) -> list[dict]:
    """What a passive node grants, one stat effect per row.

    A NODE SAYS WHAT IT DOES IN ENGLISH AND THIS SAYS IT IN NUMBERS. The four
    class tree files carry a sentence written for a player -- "+2% increased
    Life Leech per point" -- and no stat name, no bucket and no machine-readable
    value. This sheet is where those are authored, which the project owner chose
    on 2026-08-25 over changing the separate tree authoring tool's schema.

    IT COVERS A SMALL PART OF THE TREE AND THAT IS THE HONEST POSITION. Most
    nodes are not stat modifiers at all: they change a rule, generate a class
    resource, or apply only in a condition the three-bucket pipeline cannot
    express. Issue #939 measures the gap and lists what the rest would need.

    A ROW HERE IS OPTIONAL. A node with no row grants nothing, which is what
    every node did before this sheet existed, so an unauthored node is the
    ordinary case rather than an error.

    A NODE MAY HAVE SEVERAL ROWS AND ALL OF THEM APPLY. Issue #953. It could have
    exactly one until then, because the DataTable row name WAS the node name, and
    that shape cannot express the nodes that grant two things: "+1% increased
    Maximum Health and +0.5% increased Armor" is two stats in one sentence, and
    the Masochist's starting node grants three Fervour rates at once. The node is
    now a column of its own and the row name is the node with `#1`, `#2` and so
    on after it, in the order the sheet lists them. Nothing reads the row name;
    it exists because a DataTable needs a unique key.

    A NUMBER SIGN IS THE SEPARATOR BECAUSE NO NODE NAME CAN CONTAIN ONE. Node
    names are a tree name and an identifier from the tree authoring tool joined
    by an underscore, so an underscore would let `A_1` mean either the first row
    of node `A` or the only row of a node really called `A_1`.
    """
    rows = list(book["Passive Effects"].iter_rows(values_only=True))
    headers = _header_index(rows, "Passive Effects")

    out = []
    counts: dict[str, int] = {}
    for index, raw in enumerate(rows[1:], start=2):
        node = clean(_cell(raw, headers, "Node"))
        if not node:
            continue

        if "#" in node:
            raise DataError(
                f"Passive Effects row {index}: the node name {node!r} contains "
                f"a number sign, which this file uses to separate a node from "
                f"the index of its effect. Rename the node.")

        stat = clean(_cell(raw, headers, "Stat"))
        if not stat:
            raise DataError(f"Passive Effects row {index}: {node} names no stat")

        kind = clean(_cell(raw, headers, "Value Kind")).lower()
        if kind not in ("flat", "increased", "more"):
            raise DataError(
                f"Passive Effects row {index}: {node} has value kind {kind!r}, "
                f"which is not flat, increased or more")

        # A STATE OF THE CHARACTER THE BONUS ONLY APPLIES IN. Issue #959. Empty
        # is "always", which is every row before that issue.
        #
        # REFUSED IF THE GAME CANNOT JUDGE IT. A condition the engine does not
        # recognise is applied with no condition at all, which is a bonus that
        # holds all the time instead of some of the time -- silently, and in the
        # player's favour. Refusing here is the only place that can be caught.
        condition = clean(_cell(raw, headers, "Condition")).lower()
        if condition and condition not in CONDITIONS:
            raise DataError(
                f"Passive Effects row {index}: {node} names the condition "
                f"{condition!r}, which the game cannot judge. Known: "
                f"{', '.join(sorted(CONDITIONS))}.")

        condition_value = 0.0
        if condition and CONDITIONS[condition] is None:
            # A CONDITION THAT COMPARES NOTHING TAKES NO VALUE, AND A VALUE
            # BESIDE ONE IS REFUSED RATHER THAN IGNORED. Issue #962. Somebody
            # writing a number next to `while_bleeding` believes it does
            # something; the game never reads it, so the row would be worth
            # something other than what its author thought and nothing at run
            # time would say so. This is the only place that can be caught.
            written = clean(_cell(raw, headers, "Condition Value"))
            if written:
                raise DataError(
                    f"Passive Effects row {index}: {node} carries the condition "
                    f"{condition!r} and a condition value of {written!r}. That "
                    f"condition compares nothing, so the value would be "
                    f"ignored. Leave the column empty.")
        elif condition:
            condition_value = number(_cell(raw, headers, "Condition Value"),
                                     "Condition Value", index)
            low, high, units = CONDITIONS[condition]
            if not low <= condition_value <= high:
                raise DataError(
                    f"Passive Effects row {index}: {node} has a condition value "
                    f"of {condition_value}, and {condition!r} takes {units} "
                    f"between {low:g} and {high:g}.")

        # A STATE THE BONUS'S SIZE GROWS WITH. Issue #968. Empty is a fixed
        # value, which is every row before that issue.
        #
        # REFUSED IF THE GAME CANNOT JUDGE IT, and the direction of the failure
        # is the reason. An unknown scale reaching the game is made worth nothing
        # rather than worth its full value, so the node would silently grant
        # nothing at all.
        scale = clean(_cell(raw, headers, "Scale")).lower()
        if scale and scale not in SCALES:
            raise DataError(
                f"Passive Effects row {index}: {node} names the scale "
                f"{scale!r}, which the game cannot judge. Known: "
                f"{', '.join(sorted(SCALES))}.")

        scale_step = 0.0
        if scale:
            scale_step = number(_cell(raw, headers, "Scale Step"),
                                "Scale Step", index)
            low, high, units = SCALES[scale]

            # A STEP OF NOTHING MAKES THE BONUS WORTH NOTHING AT EVERY STATE,
            # which is why zero is refused here rather than only bounded.
            if scale_step <= 0.0 or not low <= scale_step <= high:
                raise DataError(
                    f"Passive Effects row {index}: {node} has a scaling step of "
                    f"{scale_step}, and {scale!r} takes {units} above 0 and up "
                    f"to {high:g}. A step of nothing is worth nothing at every "
                    f"state.")

        # WHICH OF A CAPSTONE'S THREE OPTIONS THIS ROW BELONGS TO. Issue #1029.
        # Empty is "not an option", which is every row in the four trees except
        # the capstones'.
        #
        # A COLUMN AND NOT THE `Condition` COLUMN ABOVE, and the reason is
        # concrete rather than stylistic. A row may need an option AND a real
        # condition at once: the Second Vow's second option reads "Dropping
        # below 50% health grants immunity to all damage for 5 seconds", which
        # is option 2 and a health trigger together, and one column cannot hold
        # both.
        #
        # WHICH NODES MAY CARRY ONE IS CHECKED IN `validate_passive_effects`,
        # because that is where the node table is in hand and its `Kind` column
        # says which nodes are capstones.
        option = 0
        if clean(_cell(raw, headers, "Option")):
            option = int(number(_cell(raw, headers, "Option"), "Option", index))
            if not 1 <= option <= CAPSTONE_OPTIONS:
                raise DataError(
                    f"Passive Effects row {index}: {node} names option "
                    f"{option}, and a capstone offers {CAPSTONE_OPTIONS}. Leave "
                    f"the column empty for a row that is not a capstone option.")

        counts[node] = counts.get(node, 0) + 1

        out.append({
            "Name": f"{node}#{counts[node]}",
            "Node": node,
            "Stat": stat,
            "ValueKind": kind,
            "ValuePerPoint": number(_cell(raw, headers, "Value Per Point"),
                                    "Value Per Point", index),
            "RequiredTags": clean(_cell(raw, headers, "Required Tags")),
            "Condition": condition,
            "ConditionValue": condition_value,
            "Scale": scale,
            "ScaleStep": scale_step,
            "Option": option,
        })

    # THE SAME NODE AND THE SAME STAT TWICE IS A MISTAKE RATHER THAN A DOUBLE
    # HELPING, and it is one a person editing a spreadsheet makes by copying a
    # row and forgetting to change the stat. Two rows granting the same node two
    # DIFFERENT stats is the whole point of the shape and is allowed.
    #
    # UNLESS THE TWO ARE CONDITIONED DIFFERENTLY, which is a real shape: a node
    # could give one amount always and more of it below a threshold. So the
    # condition is part of what makes a row distinct.
    #
    # AND SO IS THE SCALE, since issue #968, for the same reason: a node could
    # give a fixed amount and a further amount that grows with a state.
    #
    # AND SO IS THE CAPSTONE OPTION, since issue #1029, and that one is not a
    # refinement but a necessity. A capstone's three options are three separate
    # things a player may pick, and two of them may reasonably move the same
    # stat by different amounts. Without the option in this key the second would
    # read as a duplicated row and the sheet would refuse to be written at all.
    pairs: dict[tuple[str, str, str, float, str, float, int], int] = {}
    for row in out:
        key = (row["Node"], row["Stat"], row["Condition"],
               row["ConditionValue"], row["Scale"], row["ScaleStep"],
               row["Option"])
        pairs[key] = pairs.get(key, 0) + 1
    twice = sorted(key for key, count in pairs.items() if count > 1)
    if twice:
        listed = ", ".join(f"{key[0]} granting {key[1]}" for key in twice)
        raise DataError(
            f"the Passive Effects sheet grants the same stat twice on one "
            f"node, under the same condition, scale and capstone option: "
            f"{listed}. Two rows for one node are for two different stats, or "
            f"for the same stat under different conditions, scales or options; "
            f"anything else is a duplicated row.")

    if not out:
        raise DataError("the Passive Effects sheet is empty")
    return out


def item_base_flat_stats(item_bases: list[dict] | None) -> set[str]:
    """The stats an item base supplies as a flat implicit.

    A BASE UNDER AN INCREASE, WHICH IS THE ONLY THING THIS IS FOR. See
    `validate_passive_effects` below for the argument. `attack_damage` is the
    stat that needs it: no class stat line names it, deliberately, because a
    character's damage comes from the weapon in its hands.

    HELD HERE RATHER THAN IN THE TEST THAT ALSO NEEDS IT, so that the rule is
    stated once. `tools/tests/test_passive_effects_match_the_node_text.py`
    imports this function.

    EMPTY IS A LEGITIMATE ANSWER and not a failure: a caller that has no
    ItemBases table gets a smaller set of known stats and so a stricter check,
    which is the safe direction.
    """
    if not item_bases:
        return set()

    supplied: set[str] = set()
    for row in item_bases:
        for index in (1, 2):
            stat = str(row.get(f"Implicit{index}Stat", "") or "").strip()
            kind = str(row.get(f"Implicit{index}Kind", "") or "").strip().lower()
            if stat and kind == "flat":
                supplied.add(stat)
    return supplied


#: An item base column that supplies a stat's base, and the stat it supplies.
#:
#: A SECOND ROUTE FROM A WEAPON TO A BASE, and it is not a flat implicit. Issue
#: #1002. `attack_damage` reaches a character as an implicit on the weapon's
#: base, which `item_base_flat_stats` above sees; a swing rate is a COLUMN
#: instead, because two weapons average theirs and an implicit cannot be
#: averaged. `UCataclysmPlayerClassStats::StatBasesFromWeapons` reads that column
#: and writes the base, and the C++ test's own map of where a stat comes from
#: says exactly this: attack speed comes from "the worn weapons, as a base
#: override from StatBasesFromWeapons".
#:
#: WHY IT MATTERS TO THE CHECK BELOW. Without this, an `increased` row on
#: `attack_speed` is refused as an increase with no base under it -- and that is
#: wrong, because every weapon in the game carries one. The Masochist's Sanguine
#: Momentum node is the first to need it.
ITEM_BASE_COLUMN_STATS = {"AttackSpeed": "attack_speed"}


def item_base_column_stats(item_bases: list[dict] | None) -> set[str]:
    """The stats an item base supplies as a column rather than as an implicit.

    A COLUMN THAT IS ZERO EVERYWHERE SUPPLIES NOTHING, which is why this looks
    at the values rather than only at the header. Every armour base in the file
    carries an `AttackSpeed` of 0.0, and only the fourteen weapon bases carry a
    real one; a header alone would say a stat was supplied by a file that only
    ever writes nothing into it.
    """
    if not item_bases:
        return set()

    supplied: set[str] = set()
    for row in item_bases:
        for column, stat in ITEM_BASE_COLUMN_STATS.items():
            raw = str(row.get(column, "") or "").strip()
            try:
                if raw and float(raw) > 0.0:
                    supplied.add(stat)
            except ValueError:
                continue
    return supplied


#: Stats whose base the ENGINE supplies at the moment it asks for them, mapped to
#: where that base comes from.
#:
#: THE FOURTH KIND OF SUPPLIER, and it is named one stat at a time on purpose.
#: The other three -- a class stat line, a flat row in this sheet, a flat implicit
#: or column on an item base -- are all readable from the generated data, so the
#: check can see them. This one cannot be seen from data at all: the base is an
#: argument C++ passes to `StatForSkill`, and nothing in any CSV mentions it.
#:
#: WHY IT IS STILL A REAL SUPPLIER. The complaint this whole check exists to make
#: is that "an increase with no base under it is worth zero". A base passed by the
#: caller is a base. `attack_speed` is the precedent already accepted beside it:
#: no class line names it, and `StatBasesFromWeapons` writes it from the worn
#: weapon's column.
#:
#: WHY IT IS A DICTIONARY AND NOT A SET. Naming the source is what stops this
#: becoming a place to silence the check. An entry has to say which code supplies
#: the base, so a reader can go and look, and a stat added here without one is
#: obvious.
#:
#: AND NAMING ONE IS A PROMISE SOME CODE HAS TO KEEP. Issue #1025. The first entry
#: below named a constant nothing ever passed, so the exemption silenced this
#: check and the base really was zero: The Breaking Point opened a conversion
#: window of zero seconds and converted nothing, for as long as the node existed.
#: `UCataclysmPlayerClassStats::EngineSuppliedBases` is the map that now keeps the
#: promise, and `Cataclysm.PlayerStats.EveryEngineSuppliedBaseReachesACharacter`
#: is the test that checks it is kept.
ENGINE_SUPPLIED_BASES = {
    # Issue #985, The Breaking Point: "The conversion lasts 3 seconds, increased
    # by 5% per point". The 3 seconds is a constant of the mechanic rather than
    # anything about a character, so it is not a class stat line -- the class
    # sheet mirrors `sim/cataclysm_sim/classes.py`, which is a statement about
    # what makes each class feel different, and a timing window for one node is
    # not that.
    "damage_to_bleeding_window":
        "UCataclysmDamageConversion::BaseWindowSeconds, put on the character by "
        "UCataclysmPlayerClassStats::EngineSuppliedBases",

    # Issue #1026. What share of a hit a character takes, at 100 for normal, and
    # the same again for a hit that is damage over time. Three Masochist nodes
    # move them and all three are written as a percentage of what would otherwise
    # arrive, so a base of zero would leave every one of them worth nothing.
    #
    # NOT A CLASS STAT LINE, THOUGH FIVE STATS OF THE SAME SHAPE ARE ONE.
    # `Cataclysm.Attributes.CharacterSheetIsComplete` gives the rule: a stat is
    # off the character sheet when no affix grants it, nothing scales it, it has
    # no baseline of its own, and one passive node is its only source. The five
    # 100-means-normal stats on the `Default` line all fail that rule -- affixes
    # grant them and the Ritualist starts at 110 area of effect -- and these two
    # meet it. `docs/DECISIONS.md` records that they should be promoted to a
    # class line the day an affix grants one or a class differs on one.
    "damage_taken":
        "UCataclysmDamageCalculation::NormalDamageTaken, put on the character by "
        "UCataclysmPlayerClassStats::EngineSuppliedBases",
    "damage_over_time_taken":
        "UCataclysmDamageCalculation::NormalDamageTaken, put on the character by "
        "UCataclysmPlayerClassStats::EngineSuppliedBases",

    # AND HOW LONG A LASTING HARMFUL EFFECT ON THE CHARACTER RUNS, at 100 for
    # normal. Issue #1033. The Masochist's Symphony of Pain adds 2% a point and
    # its Vessel of Plagues adds 50%, and BOTH ARE `increased` ROWS, so without
    # a base under them the stat would resolve to zero and every stun and every
    # burn in the game would end the instant it landed. Issue #1025 records that
    # exact failure happening to the conversion window.
    "debuff_duration_taken":
        "UCataclysmDebuffs::NormalDuration, put on the character by "
        "UCataclysmPlayerClassStats::EngineSuppliedBases",
}


def validate_passive_effects(tables: dict[str, list[dict]],
                             known: set[str]) -> list[str]:
    """Every passive effect names a real node, a real stat and declared tags.

    ALL THREE FAIL SILENTLY WITHOUT THIS, and none of them errors at run time:

      a node key that does not exist   the effect is never applied to anything
      a stat nothing else supplies     the increase multiplies a base of zero
      an undeclared tag                the modifier's required tag matches
                                       nothing, so it applies to nothing

    A `flat` ROW IN THIS SHEET SUPPLIES A STAT TOO, and that is why the middle
    check is not simply "some class line names it". The complaint it exists for
    is an INCREASE with no base to multiply. A flat row IS the base: the
    Masochist's starting node grants 1 Fervour gained per 1% of maximum health
    lost, and the three rates it supplies are zero for every class on purpose,
    so no class stat line names them and none should -- a class stat row zero in
    both columns says nothing, and `test_class_sheets_match_the_model.py` holds
    that rule. Issue #954.

    A `flat` IMPLICIT ON AN ITEM BASE SUPPLIES ONE TOO, for the same reason and
    with the same argument. Issue #958. `attack_damage` is the case: no class
    line names it and none should, because a character's damage comes from the
    weapon in its hands. `UCataclysmPlayerClassStats` says so in as many words --
    "A weapon's damage is an `attack_damage` implicit on its base, so it already
    arrives as a flat modifier and the base is correctly zero" -- and a node
    reading "+2% increased damage per point" has a real base under it the moment
    a weapon is held.

    AND SO DOES AN ITEM BASE COLUMN, which is a second route from a weapon and
    not the same one. Issue #1002. A swing rate is a COLUMN on the base rather
    than an implicit, because two weapons average theirs and an implicit cannot
    be averaged, and `UCataclysmPlayerClassStats::StatBasesFromWeapons` turns it
    into the base. `attack_speed` is the case: no class line names it and none
    should, and an `increased` row on it has a real base under it the moment a
    weapon is held. See `item_base_column_stats` above.

    A ROLLED AFFIX IS DELIBERATELY NOT COUNTED, and the difference is not
    pedantic. An implicit is on EVERY item of that base type, so a Sword always
    carries attack damage; an affix may never roll at all. A stat whose only
    supplier is an optional affix really is zero for most characters, and an
    increase on it really is worth nothing, which is exactly what this check is
    for. Counting affixes would admit twelve more stats and blunt it.

    WHAT STILL CATCHES A TYPO IN A STAT NAME IS THE ENGINE SIDE.
    `Cataclysm.Passives.EveryStatAPassiveNodeGrantsHasAnAttributeBehindIt` reads
    this file and fails when a name has no gameplay attribute behind it, which is
    the failure that really matters: a stat with no attribute is dropped by
    `UCataclysmPlayerClassStats::ApplyTo` in silence.
    """
    effects = tables.get("PassiveEffects")
    nodes = tables.get("PassiveNodes")
    class_rows = tables.get("ClassStats")
    attribute_rows = tables.get("Attributes")
    if not effects or not nodes:
        return []

    node_names = {row["Name"] for row in nodes}

    # WHICH NODES ARE CAPSTONES, for the option check below. Issue #1029. The
    # `Kind` column is the same one `UCataclysmPassiveTree::CapstoneKind` matches
    # against in the engine.
    capstones = {row["Name"] for row in nodes
                 if str(row.get("Kind", "")).strip().lower() == "capstone"}

    stats = ({row["Stat"] for row in class_rows} if class_rows else set()) | \
            ({row["Stat"] for row in attribute_rows} if attribute_rows else set()) | \
            {row["Stat"] for row in effects
             if str(row["ValueKind"]).lower() == "flat"} | \
            item_base_flat_stats(tables.get("ItemBases")) | \
            item_base_column_stats(tables.get("ItemBases")) | \
            set(ENGINE_SUPPLIED_BASES)

    problems = []
    for row in effects:
        # THE `Node` COLUMN AND NOT THE ROW NAME. They were the same string
        # until issue #953; the row name now carries a `#1` suffix so that one
        # node can have several rows, and checking it against the node table
        # would report every row as missing.
        if row["Node"] not in node_names:
            problems.append(
                f"PassiveEffects/{row['Name']}: no passive node is called "
                f"{row['Node']}, so the effect reaches nothing")

        if stats and row["Stat"] not in stats:
            problems.append(
                f"PassiveEffects/{row['Name']}: {row['Stat']!r} is not a stat "
                f"any class line or attribute names, and no flat row in this "
                f"sheet supplies it either")

        for tag in (t.strip() for t in row["RequiredTags"].split(",")):
            if tag and tag not in known:
                problems.append(
                    f"PassiveEffects/{row['Name']}: undefined tag {tag}")

        # THE CAPSTONE OPTION COLUMN, WHICH FAILS SILENTLY IN BOTH DIRECTIONS.
        # Issue #1029.
        #
        # AN OPTION ON A NODE THAT IS NOT A CAPSTONE grants nothing at all:
        # `AccumulateInto` compares it against `ChosenOptionIn`, which answers 0
        # for such a node, so the row is skipped every time and the node reads as
        # unauthored.
        #
        # A CAPSTONE ROW WITH NO OPTION applies whichever of the three the player
        # picked, which is the opposite failure and the worse one: the player is
        # given part of an option they refused. Refusing it here is the only
        # place that can be caught. If a capstone ever legitimately wants a row
        # that applies whatever is chosen, this is the check to relax, and the
        # relaxation should say which node needed it.
        option = int(row.get("Option", 0) or 0)
        is_capstone = row["Node"] in capstones
        if option and not is_capstone:
            problems.append(
                f"PassiveEffects/{row['Name']}: names capstone option {option} "
                f"and {row['Node']} is not a capstone, so the row would be "
                f"skipped every time and grant nothing")
        elif is_capstone and not option:
            problems.append(
                f"PassiveEffects/{row['Name']}: {row['Node']} is a capstone and "
                f"the row names no option, so it would apply whichever of the "
                f"three the player picked")

    return problems

TABLES = {
    "DungeonModifiers": dungeon_modifiers,
    "WeaponSkills": weapon_skills,
    "EnchantmentsPositive": lambda b: enchantments(b, negative=False),
    "EnchantmentsNegative": lambda b: enchantments(b, negative=True),
    "EnemyModifiers": enemy_modifiers,
    "StatusEffects": status_effects,
    "Gems": gems,
    "CityUpgrades": city_upgrades,
    "CraftingMaterials": crafting_materials,
    "ItemBases": item_bases,
    "Affixes": affixes,
    "GearRarity": gear_rarity,
    "ItemSockets": item_sockets,
    "AffixTiers": affix_tiers,
    "EnemyDrops": enemy_drops,
    "MaterialTiers": material_tiers,
    "ClassStats": class_stats,
    "MinionTypes": minion_types,
    "MinionScaling": minion_scaling,
    "Attributes": attributes,
    "PassiveEffects": passive_effects,
    "SkillSlots": skill_slots,
    "ElementVisuals": element_visuals,
    "WeaponMeshes": weapon_meshes,
}

#: Tables built from sim/cataclysm_sim/enemy_stats.py rather than the workbook.
#: Kept apart from TABLES only so a stale-file message can name the right source;
#: everything downstream treats the two the same.
MODEL_TABLES = {
    "EnemyArchetypes": enemy_archetypes,
    "EnemyRarities": enemy_rarities,
}

#: Built from the class tree node graphs in `docs/`, which are exported from a
#: separate authoring tool. A third source beside the workbook and the Python
#: model; everything after this point treats all three alike.
TREE_TABLES = {
    "PassiveNodes": passive_nodes,
    "PassiveEdges": passive_edges,
}


# --------------------------------------------------------------------------

def render_csv(rows: list[dict]) -> str:
    if not rows:
        raise DataError("no rows to write")
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=list(rows[0].keys()),
                            lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue()


def declared_tags(book) -> set[str]:
    """Exactly what the Tags sheet declares, with no implied parents."""
    return {clean(r[0]) for r in book["Tags"].iter_rows(values_only=True)
            if r and clean(r[0])} - {"Tag Name"}


def known_tags(book) -> set[str]:
    """Every declared tag plus the parents Unreal creates implicitly."""
    known: set[str] = set()
    for tag in declared_tags(book):
        parts = tag.split(".")
        for i in range(1, len(parts) + 1):
            known.add(".".join(parts[:i]))
    return known


def validate_tags(tables: dict[str, list[dict]], known: set[str]) -> list[str]:
    problems = []
    for table, rows in tables.items():
        for row in rows:
            for tag in (t.strip() for t in row.get("Tags", "").split(",")):
                if tag and tag not in known:
                    problems.append(f"{table}/{row['Name']}: undefined tag {tag}")
    return problems


def validate_affix_slots(tables: dict[str, list[dict]]) -> list[str]:
    """Every slot an affix allows must be a slot some item base occupies.

    A misspelled slot does not fail: the affix simply never rolls, on any drop,
    and nothing reports it. Cross-checking the two sheets against each other
    catches it without either needing a hard-coded list of slots, so adding a
    slot to the design needs no change here.
    """
    bases = tables.get("ItemBases")
    affix_rows = tables.get("Affixes")
    if not bases or not affix_rows:
        return []

    real = {row["Slot"] for row in bases}
    problems = []
    for row in affix_rows:
        for slot in (s.strip() for s in row["AllowedSlots"].split(",")):
            if slot and slot not in real:
                problems.append(
                    f"Affixes/{row['Name']}: allows slot {slot!r}, which no "
                    f"item base occupies")

    # And the other way: a slot with no affix at all could fill none of its
    # four slots, which the affix pool is required to avoid.
    reachable = {s.strip() for row in affix_rows
                 for s in row["AllowedSlots"].split(",") if s.strip()}
    for slot in sorted(real - reachable):
        problems.append(f"ItemBases: slot {slot!r} can roll no affix at all")

    return problems



def validate_socket_slots(tables: dict[str, list[dict]]) -> list[str]:
    """Every gear slot an item base occupies needs a socket maximum, and no
    socket maximum may name a slot no base occupies.

    Neither direction fails anywhere else. A slot missing from the Item Sockets
    sheet leaves a drop with nothing to say how many sockets the piece can hold;
    a slot in that sheet which nothing occupies is a maximum no drop will ever
    read.

    MATCHED ON THE SLOT AND THE HAND COUNT TOGETHER, because a weapon's maximum
    depends on how many hands it takes and the two hand counts are separate rows
    in both sheets.
    """
    bases = tables.get("ItemBases")
    sockets = tables.get("ItemSockets")
    if not bases or not sockets:
        return []

    occupied = {(row["Slot"], row["Hands"]) for row in bases}
    stated = {(row["Slot"], row["Hands"]) for row in sockets}

    problems = []
    for slot, hands in sorted(occupied - stated):
        problems.append(
            f"ItemSockets: nothing states a socket maximum for {slot} at "
            f"{hands} hand(s), which item bases occupy")
    for slot, hands in sorted(stated - occupied):
        problems.append(
            f"ItemSockets: a socket maximum for {slot} at {hands} hand(s), "
            "which no item base occupies")
    return problems



def validate_enemy_drop_rarities(tables: dict[str, list[dict]]) -> list[str]:
    """The Enemy Drops sheet and the enemy rarity ladder must name the same
    rarities, in the same order.

    THE TWO COME FROM DIFFERENT PLACES, which is why this is worth checking.
    `EnemyRarities.csv` is generated from `sim/cataclysm_sim/enemy_stats.py`,
    whose ladder is in turn a port of the external DungeonSimulator power model.
    `EnemyDrops.csv` is typed into the workbook. A rarity added to one and not
    the other is a rarity that either drops nothing or drops for a creature that
    cannot exist, and nothing else would report it.

    ORDER TOO, NOT ONLY MEMBERSHIP, because both tables carry a Step column
    saying where a rarity sits and the two have to agree about that as well.
    """
    drops = tables.get("EnemyDrops")
    rarities = tables.get("EnemyRarities")
    if not drops or not rarities:
        return []

    problems = []
    dropping = [row["Name"] for row in drops]
    known = [row["Name"] for row in rarities]
    if dropping != known:
        problems.append(
            f"EnemyDrops names {dropping} and EnemyRarities names {known}. "
            "They have to be the same rarities in the same order; the ladder "
            "comes from the simulation's enemy model and is not the workbook's "
            "to change.")
        return problems

    for drop, rarity in zip(drops, rarities, strict=True):
        if drop["Step"] != rarity["Step"]:
            problems.append(
                f"EnemyDrops/{drop['Name']}: Step {drop['Step']} against "
                f"{rarity['Step']} in EnemyRarities")
    return problems


#: `Tier 3 (Rare). Drop from Dungeon Bosses/Elites.` -- the tier and its name.
MATERIAL_TIER_AND_SOURCE = re.compile(r"Tier\s*(\d)\s*\(([^)]+)\)")

#: `Upgrade Stone +7` -- the level the stone takes a piece of gear to.
UPGRADE_STONE_NAME = re.compile(r"Upgrade\s+Stone\s*\+(\d+)$")

#: The highest a piece of gear goes. Section VII of the design document states
#: it twice: the Upgrade Item Level operation is "1 per tier, maximum 10", and
#: the stone that reaches it "Raises an item to +10, which is the maximum".
#: Nothing above that can be crafted, so nothing above it can drop.
MAX_UPGRADE_STONE_LEVEL = 10


def validate_material_tier_counts(tables: dict[str, list[dict]]) -> list[str]:
    """Every material tier's stated count must match how many materials it has.

    THE COUNT IS NOT DECORATION. How often a NAMED top-tier material drops is
    the tier's share divided by this number, and Purified Essence -- the only
    thing that clears the Consumption Threshold -- is one of three that share
    the top tier. A wrong count here makes that figure wrong and nothing else
    notices.

    Counted off the Crafting sheet through `CraftingMaterials`, which is where
    the materials actually are, rather than trusting either restatement.
    """
    tiers = tables.get("MaterialTiers")
    materials = tables.get("CraftingMaterials")
    if not tiers or not materials:
        return []

    counted: dict[str, int] = {}
    for row in materials:
        found = MATERIAL_TIER_AND_SOURCE.match(row["TierAndSource"])
        if found:
            counted[found.group(2)] = counted.get(found.group(2), 0) + 1

    if not counted:
        return ["MaterialTiers: no 'Tier N (Name)' cell could be read out of "
                "CraftingMaterials, so the counts could not be checked at all"]

    problems = []
    for row in tiers:
        real = counted.get(row["TierName"])
        if real is None:
            problems.append(
                f"MaterialTiers/{row['Name']}: no crafting material is in a "
                f"tier called {row['TierName']!r}")
        elif real != row["Materials"]:
            problems.append(
                f"MaterialTiers/{row['Name']}: says {row['Materials']} "
                f"materials, and CraftingMaterials has {real}")

    for name in sorted(set(counted) - {row["TierName"] for row in tiers}):
        problems.append(
            f"MaterialTiers: CraftingMaterials has {counted[name]} material(s) "
            f"in a tier called {name!r}, and MaterialTiers has no such row")
    return problems


def validate_upgrade_stone_levels(tables: dict[str, list[dict]]) -> list[str]:
    """The ten upgrade stones must publish levels +1 to +10, one stone each.

    THE LEVEL IS DERIVED FROM THE NAME, so a rename in the workbook silently
    turns it into a zero rather than into an error. A zero reads as "not a
    stone", and `UCataclysmDropRoll::RollMaterial` caps what may drop by that
    level -- so a stone with a zero would be droppable at every difficulty tier,
    which is the exact bug #863 is about, arriving the other way round.

    Checked as a set rather than as a count. Two stones both named "+7" would
    keep the count at ten and put a hole in the ladder.
    """
    materials = tables.get("CraftingMaterials")
    if not materials:
        return []

    levels: dict[int, list[str]] = {}
    for row in materials:
        if row["UpgradeLevel"]:
            levels.setdefault(row["UpgradeLevel"], []).append(row["MaterialName"])

    if not levels:
        return ["CraftingMaterials: no material is named 'Upgrade Stone +N', so "
                "no upgrade stone can drop at all. The ten stones were renamed "
                "or removed, or UPGRADE_STONE_NAME no longer matches them."]

    problems = []
    for level in range(1, MAX_UPGRADE_STONE_LEVEL + 1):
        if level not in levels:
            problems.append(
                f"CraftingMaterials: nothing upgrades gear to +{level}. Gear "
                f"runs +1 to +{MAX_UPGRADE_STONE_LEVEL} and each step needs a "
                f"stone.")
        elif len(levels[level]) > 1:
            problems.append(
                f"CraftingMaterials: {len(levels[level])} materials upgrade "
                f"gear to +{level}: {sorted(levels[level])}")
    return problems


#: The one weapon type in the Weapon Skills sheet that is not a weapon. It marks
#: a skill that does not depend on which weapon is held, which is how the single
#: Aura skill is written: one aura per damage type rather than one per weapon.
WEAPON_INDEPENDENT_SKILL = "All"


def validate_weapon_skill_types(tables: dict[str, list[dict]]) -> list[str]:
    """Every weapon type in the Weapon Skills sheet must be a real weapon base.

    The two sheets held different names for the same three weapons: the skills
    sheet said "2H Axe", "2h Sword" and "2H Warhammer" where the item bases sheet
    and the design document's weapon table said Greataxe, Greatsword and
    Warhammer. Nothing reported it, and a lookup keyed on weapon type would have
    found no skills at all for five of the fourteen bases.

    Cross-checked against the other sheet rather than a list written here, so
    adding a weapon to the design needs no change in this file.

    IT ONLY EVER LOOKED AT THE `WeaponType` COLUMN, which is why the same rename
    survived in the gameplay tags. `validate_weapon_tags` below is the other
    half. Issue #620.
    """
    bases = tables.get("ItemBases")
    skills = tables.get("WeaponSkills")
    if not bases or not skills:
        return []

    real = {row["WeaponType"] for row in bases if row["WeaponType"]}
    problems = []
    for name in sorted({row["WeaponType"] for row in skills if row["WeaponType"]}):
        if name != WEAPON_INDEPENDENT_SKILL and name not in real:
            problems.append(
                f"WeaponSkills: weapon type {name!r} is not a weapon base. The "
                f"bases are {sorted(real)}.")

    # And the other way. A weapon with no rows at all can never have a skill in
    # any slot, which is a hole rather than a design choice.
    covered = {row["WeaponType"] for row in skills}
    for name in sorted(real - covered):
        problems.append(f"WeaponSkills: the {name} has no rows at all")

    return problems


def validate_weapon_skill_damage_types(tables: dict[str, list[dict]],
                                       declared: set[str]) -> list[str]:
    """Every damage type in the Weapon Skills sheet must be one somebody has.

    THE SAME SHAPE OF FAILURE `validate_weapon_skill_types` ABOVE WAS WRITTEN
    FOR, one column over, and until issue #579 the Damage Type column was checked
    by nothing at all. `UCataclysmWeaponSkills::SkillsFor` in the engine matches
    it with `Row.DamageType.Equals(DamageType, ESearchCase::IgnoreCase)`, an exact
    comparison, so a row naming a damage type nobody has is offered to nobody: it
    generates cleanly, imports cleanly, fills no slot and produces no error. A
    single misspelling costs one skill and says nothing.

    THE WILDCARD CASE IS THE ONE THAT LOOKS LIKE IT SHOULD WORK. The WEAPON axis
    has one -- `UCataclysmWeaponSkills::WeaponIndependent` is `All`, and the eight
    Aura rows use it, because an aura is not a property of the weapon being held.
    The damage type axis has no such thing, so `All` in this column is dead data
    that reads like a feature. Refused by name below, with the reason, because
    somebody writing it has guessed at symmetry that is not there.

    NOTHING IN THE SHEET USES `All` TODAY and nothing needs to. Issue #579 was
    filed when the basic attack was expected to live in this matrix as fourteen
    damage-type-independent rows; issue #524 settled it the other way, and the
    basic attack now comes from the weapon base through
    `UCataclysmWeaponSkills::BasicAttackFor`. So the wildcard is not built,
    because building it would mean choosing a precedence rule between four
    classes of row with no data to check the choice against.

    Cross-checked against the Tags sheet rather than a list written here, so
    adding a ninth damage type to the design needs no change in this file.
    """
    skills = tables.get("WeaponSkills")
    if not skills:
        return []

    real = {tag[len(ELEMENT_TAG_PREFIX):] for tag in declared
            if tag.startswith(ELEMENT_TAG_PREFIX)}
    if not real:
        return []

    problems = []
    for name in sorted({row["DamageType"] for row in skills if row["DamageType"]}):
        if name in real:
            continue
        if name == WEAPON_INDEPENDENT_SKILL:
            problems.append(
                f"WeaponSkills: damage type {name!r} is not a wildcard. The "
                f"WeaponType column has one and this column does not: "
                f"UCataclysmWeaponSkills::SkillsFor matches the damage type "
                f"exactly, so these rows would be granted to nobody. Issue #579 "
                f"records what building one would take. Name one of "
                f"{sorted(real)} instead.")
        else:
            problems.append(
                f"WeaponSkills: damage type {name!r} is not declared on the Tags "
                f"sheet. A row naming a damage type nobody has is granted to "
                f"nobody and reports nothing. The declared ones are "
                f"{sorted(real)}.")

    # And the other way. A damage type with no rows at all is a Cataclysm whose
    # characters have no skills, which is a hole rather than a design choice.
    covered = {row["DamageType"] for row in skills}
    for name in sorted(real - covered):
        problems.append(
            f"WeaponSkills: the {name} damage type has no rows at all")

    return problems


#: The prefix every weapon type's gameplay tag carries. Declared on the Tags
#: sheet and generated into game/Config/Tags/CataclysmTags.ini.
WEAPON_TAG_PREFIX = "Item.Weapon."


def weapon_tag_leaf(weapon_type: str) -> str:
    """The tag leaf a weapon type must use, from the type's own name.

    ONE RULE, AND IT IS ONLY ABOUT SPACES. A gameplay tag cannot contain a space
    and one weapon type does: "2H Crossbow" in `game/Data/ItemBases.csv`. So the
    leaf is the weapon type with its spaces removed and nothing else changed,
    which makes `Item.Weapon.2HCrossbow` the only legal spelling for it and
    leaves the other thirteen identical to their weapon type.
    """
    return weapon_type.replace(" ", "")


def validate_weapon_tags(tables: dict[str, list[dict]],
                         declared: set[str]) -> list[str]:
    """Every weapon type's tag is named after the weapon. Issue #620.

    WHY THIS EXISTS. Three weapons carried a tag named after what they used to be
    called: a Greataxe's rows said `Item.Weapon.2hAxe`, a Greatsword's said
    `Item.Weapon.2hSword`, a Warhammer's said `Item.Weapon.2hWarhammer`, and the
    2H Crossbow's differed in letter case. That was the tail of a rename which
    corrected the `WeaponType` column and left the tags behind.

    NOTHING WAS BROKEN BY IT, which is exactly why it survived. The naming was
    consistent within the data, so every lookup worked and no test failed. What
    it cost was a reader having to know two names for one weapon.

    `validate_weapon_skill_types` above is the same shape of check for the
    `WeaponType` column, and it is what caught the first half of that rename. It
    could not catch this half because it never looks at a tag.

    BOTH DIRECTIONS. A weapon with no tag cannot be scoped to by an affix or a
    passive, and a weapon tag naming no weapon is a name for something that does
    not exist.
    """
    bases = tables.get("ItemBases")
    if not bases:
        return []

    weapons = {row["WeaponType"] for row in bases if row["WeaponType"]}
    wanted = {WEAPON_TAG_PREFIX + weapon_tag_leaf(name) for name in weapons}
    have = {tag for tag in declared if tag.startswith(WEAPON_TAG_PREFIX)}

    problems = [
        f"Tags: {tag} is not declared, and it is the tag for a weapon type in "
        f"the Item Bases sheet. A weapon with no tag cannot be scoped to."
        for tag in sorted(wanted - have)
    ]
    problems += [
        f"Tags: {tag} names no weapon type. The Item Bases sheet's weapon types "
        f"are {sorted(weapons)}, so the tags are {sorted(wanted)}."
        for tag in sorted(have - wanted)
    ]
    return problems


def validate_skill_effects(tables: dict[str, list[dict]]) -> list[str]:
    """A skill's Effect must name a status effect that exists.

    An Effect nobody defined generates cleanly and grants nothing at runtime,
    because the tag it would use is never declared: the skill runs, spends its
    mana, waits its cooldown and applies no debuff. Cross-checked against the
    Buffs, Debuffs and DoTs sheets rather than a list written here, so adding an
    effect to the design needs no change in this file.
    """
    skills = tables.get("WeaponSkills")
    effects = tables.get("StatusEffects")
    if not skills or not effects:
        return []

    known = {row["EffectName"] for row in effects if row["EffectName"]}
    problems = []
    for row in skills:
        if not row["Shape"] or not row["ShapeParams"]:
            continue
        params = parse_shape_params(row["ShapeParams"], row["Shape"],
                                    f"WeaponSkills/{row['Name']}")
        # MORE THAN ONE IS ALLOWED, comma separated, as `Minions` already is.
        # Anathema lays every hex the Wand knows at once and composes from the
        # two that exist rather than inventing a third, so it writes
        # "Shred, Madness". Checked one name at a time: a list validated whole
        # would fail on the comma and say nothing useful about which half is
        # wrong.
        for named in (n.strip() for n in (params.get("Effect") or "").split(",")):
            if named and named not in known:
                problems.append(
                    f"WeaponSkills/{row['Name']}: applies the effect {named!r}, "
                    f"which is not in the Buffs, Debuffs or DoTs sheets")
    return problems


def validate_minion_references(tables: dict[str, list[dict]]) -> list[str]:
    """A skill's `Minions` must name types the minion table defines. Issue #338.

    THE WHOLE POINT OF THE MINION TYPE TABLE is that a creature's numbers live in
    one place, so two skills producing the same creature cannot disagree about
    it. That only holds if the skill names the type; a name with no row is a
    skill that summons something with no stats.

    IT ALSO CHECKS COUNT AGREES. A skill naming one type states how many twice --
    in `Count` and in the `Minions` entry -- and this makes that redundancy a
    guard rather than a chance to drift. Iron Fortress is the reason `Minions`
    carries counts at all: it deploys two ballistae and three spike traps, which
    a single `Count` cannot express, so it states no `Count`.
    """
    skills = tables.get("WeaponSkills")
    types = tables.get("MinionTypes")
    if not skills or not types:
        return []

    known = {row["Name"] for row in types}
    problems = []
    for row in skills:
        if not row["Shape"] or not row["ShapeParams"]:
            continue
        where = f"WeaponSkills/{row['Name']}"
        params = parse_shape_params(row["ShapeParams"], row["Shape"], where)
        written = params.get("Minions")
        if not written:
            continue

        produced = parse_minions(written, where)
        for name in produced:
            if name not in known:
                problems.append(
                    f"{where}: produces {name!r}, which has no row in the "
                    f"Minion Types sheet. It reads {sorted(known)}.")

        count = params.get("Count")
        if count and len(produced) == 1:
            only = next(iter(produced.values()))
            if float(count) != float(only):
                problems.append(
                    f"{where}: Count is {count} and Minions says {only}. One "
                    "of the two is wrong, and nothing would notice at runtime.")
    return problems


def validate_hybrid_parts(tables: dict[str, list[dict]]) -> list[str]:
    """A hybrid affix must name two affixes that exist.

    A hybrid grants two stats at a reduced share each. If a part names nothing,
    the hybrid grants half of what it says and the shortfall is invisible.
    """
    affix_rows = tables.get("Affixes")
    if not affix_rows:
        return []

    names = {row["AffixName"] for row in affix_rows}
    problems = []
    for row in affix_rows:
        if row["AffixKind"] != "Hybrid":
            continue
        for field in ("HybridPart1", "HybridPart2"):
            part = row[field]
            if not part:
                problems.append(f"Affixes/{row['Name']}: {field} is empty")
            elif part not in names:
                problems.append(
                    f"Affixes/{row['Name']}: {field} names {part!r}, which is "
                    "not an affix")
    return problems


#: Stats whose base is not a quantity a class holds.
#:
#: Cooldown reduction is the accumulated sum of increases rather than a value:
#: a skill's cooldown is its own base divided by one plus this. A class base of
#: zero is correct for it, so it is not worth reporting.
RATE_STATS = frozenset({"cooldown_reduction"})


def validate_stat_names(tables: dict[str, list[dict]]) -> list[str]:
    """Report attribute effects whose stat no class supplies a base for.

    ATTRIBUTES ONLY EVER SCALE. A point adds to a stat's sum of increases and
    the sum multiplies a base, so a point does nothing until something supplies
    that base. The class is not the only thing that can: gear implicits and
    affixes supply block chance, critical strike chance and evasion, and a
    weapon supplies attack speed. So this is information rather than an error.

    It is worth printing because the alternative is nobody noticing. Attack
    speed had no base anywhere at all for some time, which made every attack
    speed affix on every item worth exactly nothing; see issue #120.

    The two sheets are checked against each other rather than against a
    hard-coded list of the 33 stats, so adding a stat to the character sheet
    needs no change here.
    """
    class_rows = tables.get("ClassStats")
    attribute_rows = tables.get("Attributes")
    if not class_rows or not attribute_rows:
        return []

    from_classes = {row["Stat"] for row in class_rows}
    from_attributes = {row["Stat"] for row in attribute_rows}

    return [f"{stat!r} is scaled by an attribute, and no class supplies its "
            f"base. It does nothing until gear, a weapon or a skill does."
            for stat in sorted(from_attributes - from_classes - RATE_STATS)]


def validate_element_visuals(tables: dict[str, list[dict]],
                             declared: set[str]) -> list[str]:
    """Every damage type has exactly one effect palette row, and no row invents
    a damage type.

    BOTH DIRECTIONS MATTER AND THE FIRST IS THE QUIET ONE. A damage type with no
    row is a skill whose effects fall back to whatever the Niagara system was
    authored with, so every Void skill would look like whichever damage type the
    artist happened to build the system against. Nothing would report it: the
    table would load, the lookup would miss, and the effect would still play.

    Cross-checked against the Tags sheet rather than against a list written
    here, so adding a ninth damage type to the design needs no change in this
    file -- it fails generation until the palette row exists.
    """
    rows = tables.get("ElementVisuals")
    if not rows:
        return []

    wanted = {tag for tag in declared if tag.startswith(ELEMENT_TAG_PREFIX)}
    have = {row["ElementTag"] for row in rows}

    problems = [
        f"ElementVisuals: {tag} is a declared damage type with no effect "
        f"palette row, so its effects would take whatever colour the Niagara "
        f"system was authored with"
        for tag in sorted(wanted - have)
    ]
    problems += [
        f"ElementVisuals/{row['Name']}: ElementTag {row['ElementTag']!r} is not "
        f"declared on the Tags sheet. The declared damage types are "
        f"{sorted(wanted)}."
        for row in rows if row["ElementTag"] not in wanted
    ]
    return problems


def validate_weights(tables: dict[str, list[dict]]) -> list[str]:
    problems = []
    for table, rows in tables.items():
        for row in rows:
            if "Weight" in row and not 0 < float(row["Weight"]) <= 100:
                problems.append(f"{table}/{row['Name']}: weight "
                                f"{row['Weight']} is outside 0 to 100")
    return problems


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true",
                        help="verify without writing; exit 1 if out of date")
    parser.add_argument("--workbook", type=pathlib.Path, default=WORKBOOK)
    parser.add_argument("--output-dir", type=pathlib.Path, default=OUTPUT_DIR)
    args = parser.parse_args(argv)

    import openpyxl

    try:
        book = openpyxl.load_workbook(args.workbook, data_only=True)
        tables = {name: builder(book) for name, builder in TABLES.items()}
        tables.update({name: builder()
                       for name, builder in MODEL_TABLES.items()})
        tables.update({name: builder()
                       for name, builder in TREE_TABLES.items()})
    except DataError as error:
        print(f"FAIL: {error}", file=sys.stderr)
        return 1
    except KeyError as error:
        print(f"FAIL: the workbook is missing sheet {error}", file=sys.stderr)
        return 1

    # Reported and not fatal. An attribute scaling a stat no class supplies a
    # base for is the design's own stated case: attributes only ever scale, so
    # a point does nothing until gear, a weapon or a skill supplies the base.
    # Worth printing, because the alternative is nobody noticing.
    notes = validate_stat_names(tables)
    if notes:
        print(f"NOTE: {len(notes)} attribute effect(s) scale a stat no class "
              f"supplies:", file=sys.stderr)
        for line in notes:
            print(f"  {line}", file=sys.stderr)

    problems = (validate_tags(tables, known_tags(book))
                + validate_weights(tables)
                + validate_affix_slots(tables)
                + validate_socket_slots(tables)
                + validate_enemy_drop_rarities(tables)
                + validate_material_tier_counts(tables)
                + validate_upgrade_stone_levels(tables)
                + validate_weapon_skill_types(tables)
                + validate_weapon_skill_damage_types(tables, declared_tags(book))
                + validate_weapon_tags(tables, declared_tags(book))
                + validate_skill_effects(tables)
                + validate_minion_references(tables)
                + validate_hybrid_parts(tables)
                + validate_element_visuals(tables, declared_tags(book))
                + validate_passive_effects(tables, known_tags(book)))
    if problems:
        print(f"FAIL: {len(problems)} validation problem(s):", file=sys.stderr)
        for line in problems[:40]:
            print(f"  {line}", file=sys.stderr)
        if len(problems) > 40:
            print(f"  ... and {len(problems) - 40} more", file=sys.stderr)
        return 1

    stale = []
    for name, rows in tables.items():
        contents = render_csv(rows)
        path = args.output_dir / f"{name}.csv"
        if args.check:
            if not path.is_file() or path.read_text(encoding="utf-8") != contents:
                stale.append(path.name)
        else:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(contents, encoding="utf-8")

    if args.check:
        if stale:
            print(f"FAIL: {len(stale)} DataTable CSV(s) out of date. "
                  "Run without --check.", file=sys.stderr)
            for name in sorted(stale):
                source = ("sim/cataclysm_sim/enemy_stats.py"
                          if pathlib.Path(name).stem in MODEL_TABLES
                          else args.workbook.name)
                print(f"  {name:<28}behind {source}", file=sys.stderr)
            return 1
        print(f"All {len(tables)} DataTable CSVs are up to date.")
        return 0

    for name, rows in sorted(tables.items()):
        print(f"  {name + '.csv':<28}{len(rows):>5} rows")
    print(f"Wrote {len(tables)} CSVs to game/Data/")
    return 0


if __name__ == "__main__":
    sys.exit(main())

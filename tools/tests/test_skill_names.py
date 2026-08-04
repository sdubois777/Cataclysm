"""No skill may be named after a character class.

WHY THIS MATTERS. Skills come from WEAPONS, not from classes: the equipped
weapon's type and damage type decide which skill fills each of the six slots,
and the class tree is a separate choice made from the same damage type. A skill
named after a class implies it belongs to that class, which is a rule the game
does not have.

Issue #157 is the case that prompted this: the War Greataxe Heavy skill was
called "Ravager's Cleave", and Ravager is a DEMONIC class. So a player holding a
War greataxe was granted a skill named after a class they cannot spec into at
all.

BOTH LISTS ARE READ FROM THE DESIGN, NOT WRITTEN HERE. The class names come from
the "Classes by Damage Type" section of docs/Cataclysm_GDD_v2.md, which lists all
24, and the skill names come from the Weapon Skills sheet of
docs/All_Things_Cataclysm.xlsx. So adding a class or a skill needs no change in
this file, and renaming a class cannot leave a stale copy behind here.
"""

from __future__ import annotations

import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
DESIGN_DOC = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"

#: The heading the class tables sit under, and the one that ends them.
CLASS_SECTION_START = "## **Classes by Damage Type**"

#: Class names to ignore. "Default" is the shared stat line every class inherits
#: rather than a class anyone plays, and it appears in the Class Stats sheet
#: rather than in the design document's class tables.
NOT_REALLY_CLASS_NAMES = frozenset({"Default"})

#: The four skills renamed for issue #157, and what each must still describe.
#:
#: RENAMING MUST NOT QUIETLY CHANGE WHAT A SKILL DOES. The issue asked only for
#: new names, so each description has to keep saying what it said.
RENAMED_FOR_ISSUE_157 = {
    ("Greataxe", "Heavy"): ("Sundering Arc", ("wide cone", "bleed", "20%")),
    ("Axe", "Movement"): ("Headlong Charge", ("straight line", "distance", "bleed")),
    ("Dagger", "Movement"): ("Vanishing Trap", ("teleport", "trap", "12 meters")),
    ("Shield", "Support"): ("Braced Guard", ("block chance", "100%", "3 seconds")),
}


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def class_names() -> set[str]:
    """Every class name in the design document's Classes by Damage Type tables."""
    document = DESIGN_DOC.read_text(encoding="utf-8")
    start = document.find(CLASS_SECTION_START)
    assert start != -1, (
        f"{DESIGN_DOC.name} has no {CLASS_SECTION_START!r} section, so this test "
        f"has nothing to check against")

    # The section runs until the next SECOND-level heading, not the next
    # top-level one. Stopping at "\n# " scoops up every table between here and
    # the next part of the document -- affixes, effects, skills -- and their
    # first columns are not class names.
    end = document.find("\n## ", start + len(CLASS_SECTION_START))
    section = document[start:end if end != -1 else len(document)]

    names: set[str] = set()
    for line in section.splitlines():
        match = re.match(r"^\|\s*([A-Z][A-Za-z' ]+?)\s*\|", line.strip())
        if not match:
            continue
        name = match.group(1).strip()
        # The header row of every table reads "**Class**", which the escaping in
        # this document renders as \*\*Class\*\* and the pattern above rejects.
        if name and name not in NOT_REALLY_CLASS_NAMES:
            names.add(name)

    assert len(names) >= 20, (
        f"only found {len(names)} class names in {DESIGN_DOC.name}: "
        f"{sorted(names)}. The tables' format probably changed and this test "
        f"stopped checking.")
    return names


@pytest.fixture(scope="module")
def skill_names() -> list[tuple[str, str, str, str]]:
    """(weapon, damage type, slot, skill name) for every designed skill."""
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    out = []
    for raw in list(book["Weapon Skills"].iter_rows(values_only=True))[1:]:
        if not raw or not text(raw[0]) or not text(raw[3]):
            continue
        out.append((text(raw[0]), text(raw[1]), text(raw[2]), text(raw[3])))
    return out


def test_the_design_document_still_lists_the_classes(class_names):
    """The three Demonic classes are the ones the vertical slice ships."""
    for expected in ("Ravager", "Ritualist", "Masochist"):
        assert expected in class_names, (
            f"{expected} is not in the design document's class tables, so this "
            f"test would not catch a skill named after it")


def test_no_skill_is_named_after_a_class(class_names, skill_names):
    """Issue #157's acceptance criterion, across all 77 designed skills."""
    assert skill_names, "no skill is designed at all, so this checks nothing"

    offenders = []
    for weapon, damage, slot, name in skill_names:
        for class_name in sorted(class_names):
            # Substring rather than equality, because the case that prompted
            # this was "Ravager's Cleave" rather than a bare "Ravager".
            if class_name.lower() in name.lower():
                offenders.append(
                    f"{damage} {weapon}/{slot}: {name!r} contains the class "
                    f"name {class_name!r}")

    assert not offenders, (
        "skills are granted by weapons, not by classes, so a skill must not be "
        "named after one:\n  " + "\n  ".join(offenders))


def test_the_renamed_skills_kept_their_names_and_descriptions():
    """Renaming a skill must not quietly change what it does.

    Issue #157 asked only for new names. Each of the four still has to describe
    what it described before, or the rename took the behaviour with it.
    """
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    rows = {}
    for raw in book["Weapon Skills"].iter_rows(values_only=True):
        if raw and text(raw[1]) == "War":
            rows[(text(raw[0]), text(raw[2]))] = raw

    problems = []
    for key, (expected_name, phrases) in RENAMED_FOR_ISSUE_157.items():
        raw = rows.get(key)
        if raw is None:
            problems.append(f"the War {key[0]}/{key[1]} row is missing")
            continue

        actual = text(raw[3])
        if actual != expected_name:
            problems.append(
                f"War {key[0]}/{key[1]} is named {actual!r}, expected "
                f"{expected_name!r}")

        description = text(raw[4])
        for phrase in phrases:
            if phrase not in description:
                problems.append(
                    f"War {key[0]}/{key[1]} ({actual}) no longer mentions "
                    f"{phrase!r}: {description}")

    assert not problems, "\n".join(problems)

"""The Affix Tiers sheet must agree with the simulation's drop weights.

WHY THIS EXISTS. How heavily each affix tier is weighted on a drop lives in two
places: the Drop Weight column of the Affix Tiers sheet in
`docs/All_Things_Cataclysm.xlsx`, and `AFFIX_TIER_DROP_WEIGHT` in
`sim/cataclysm_sim/affixes.py`.

Two copies of the same seven numbers drift. This project has been bitten by
exactly that twice with `sim/cataclysm_sim/scoring.py`, which is why
`sim/verify_scoring_port.py` exists.

WHICH IS AUTHORITATIVE. The workbook, as everywhere else. When this fails, the
usual fix is to change the Python to match the sheet.

WHAT WOULD GO UNNOTICED WITHOUT IT. Nothing errors when they disagree. A weight
changed in the workbook to make top-tier affixes rarer would leave every drop
rolling on the old shape, and every figure derived in the simulation would
describe a different game from the one being played.
"""

from __future__ import annotations

import pathlib
import sys

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
SHEET = "Affix Tiers"


@pytest.fixture(scope="module")
def sheet() -> dict[int, float]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if SHEET not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {SHEET!r}")

    rows = list(book[SHEET].iter_rows(values_only=True))
    headers = [("" if h is None else str(h).strip()) for h in rows[0]]
    tier_at = headers.index("Tier")
    weight_at = headers.index("Drop Weight")

    out: dict[int, float] = {}
    for raw in rows[1:]:
        if not raw or raw[tier_at] is None:
            continue
        out[int(raw[tier_at])] = float(raw[weight_at])
    return out


@pytest.fixture(scope="module")
def model():
    if str(REPO_ROOT / "sim") not in sys.path:
        sys.path.insert(0, str(REPO_ROOT / "sim"))
    from cataclysm_sim import affixes

    return affixes


def test_the_sheet_was_found_and_lists_all_seven_tiers(sheet, model) -> None:
    """Read a real sheet, so the comparison below cannot pass on an empty one."""
    assert sorted(sheet) == list(model.AFFIX_TIERS), (
        f"the {SHEET} sheet lists tiers {sorted(sheet)}, and there are "
        f"{list(model.AFFIX_TIERS)}")


def test_every_drop_weight_matches(sheet, model) -> None:
    wrong = []
    for affix_tier, in_sheet in sheet.items():
        in_model = float(model.AFFIX_TIER_DROP_WEIGHT[affix_tier])
        if abs(in_sheet - in_model) > 1e-9:
            wrong.append(f"T{affix_tier}: sheet {in_sheet}, model {in_model}")

    assert not wrong, (
        "the Drop Weight column and AFFIX_TIER_DROP_WEIGHT in "
        f"sim/cataclysm_sim/affixes.py disagree: {'; '.join(wrong)}. The "
        "workbook is authoritative, so change the Python."
    )


def test_the_sheet_halves_at_every_step(sheet) -> None:
    """The shape the project owner set, read from the sheet rather than the
    model. Every other test here compares two copies to each other, which passes
    happily if both were changed together by mistake.
    """
    for lower in range(1, max(sheet)):
        assert sheet[lower] / sheet[lower + 1] == pytest.approx(2.0), (
            f"T{lower} weighs {sheet[lower]} against T{lower + 1}'s "
            f"{sheet[lower + 1]}, which is not a halving")


def test_a_top_tier_affix_is_one_in_a_hundred_and_twenty_seven(sheet) -> None:
    """What the sheet adds up to, at the difficulty tier where every tier is
    reachable."""
    assert sum(sheet.values()) == pytest.approx(127.0)

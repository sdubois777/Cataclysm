"""The Name Word column of the Affixes sheet must agree with the naming table.

WHY THIS EXISTS. The word each suffix affix contributes to an item's name lives in
two places: the Name Word column of the Affixes sheet in
`docs/All_Things_Cataclysm.xlsx`, and `AFFIX_NAME_WORD` in
`sim/cataclysm_sim/naming.py`.

Two copies of the same values drift. This project has been bitten by exactly that
twice with `sim/cataclysm_sim/scoring.py`, which is why `sim/verify_scoring_port.py`
exists.

WHICH IS AUTHORITATIVE. The workbook, as everywhere else. When this fails, the
usual fix is to change the Python to match the sheet.

WHAT WOULD GO UNNOTICED WITHOUT IT. Nothing errors when they disagree. A word
changed in the workbook would simply never appear on an item, and the old one
would keep being used with nothing to say so.

WHAT THIS DOES NOT CHECK. Whether a word is a good name for its affix. That is a
judgement and it belongs to the project owner.
"""

from __future__ import annotations

import pathlib
import sys

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
SHEET = "Affixes"


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def sheet() -> dict[str, dict[str, str]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if SHEET not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {SHEET!r}")

    rows = list(book[SHEET].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    if "Name Word" not in headers:
        pytest.fail(
            f"the {SHEET} sheet has no Name Word column. Its columns are "
            f"{headers}.")

    out: dict[str, dict[str, str]] = {}
    for raw in rows[1:]:
        if not raw or not text(raw[0]):
            continue
        record = {headers[i]: text(raw[i])
                  for i in range(len(headers)) if i < len(raw)}
        out[text(record["Affix Name"])] = record
    return out


@pytest.fixture(scope="module")
def model():
    if str(REPO_ROOT / "sim") not in sys.path:
        sys.path.insert(0, str(REPO_ROOT / "sim"))
    from cataclysm_sim import naming

    return naming


def test_the_sheet_was_found_and_holds_words(sheet) -> None:
    """Read a real sheet, so the comparisons below cannot pass on an empty one."""
    with_words = [name for name, record in sheet.items() if record["Name Word"]]
    assert len(with_words) > 40, (
        f"only {len(with_words)} rows of the {SHEET} sheet carry a Name Word, "
        "which is too few to be the real column")


def test_every_word_in_the_sheet_matches_the_model(sheet, model) -> None:
    wrong = []
    for name, record in sheet.items():
        in_sheet = record["Name Word"]
        if not in_sheet:
            continue
        in_model = model.AFFIX_NAME_WORD.get(name)
        if in_sheet != in_model:
            wrong.append(f"{name}: sheet {in_sheet!r}, model {in_model!r}")

    assert not wrong, (
        "the Name Word column and AFFIX_NAME_WORD in "
        f"sim/cataclysm_sim/naming.py disagree: {'; '.join(wrong)}. The workbook "
        "is authoritative, so change the Python."
    )


def test_the_model_names_nothing_the_sheet_leaves_blank(sheet, model) -> None:
    """The other direction, which the test above cannot see.

    A word added to the Python and not to the workbook would be used on real
    items while the sheet said the affix had no name.
    """
    stray = []
    for name, word in model.AFFIX_NAME_WORD.items():
        in_sheet = sheet.get(name, {}).get("Name Word", "")
        if not in_sheet:
            stray.append(f"{name} is {word!r} in the model and blank in the sheet")

    assert not stray, "; ".join(stray)


def test_only_suffix_affixes_carry_a_word(sheet) -> None:
    """An item's first word is its rarity, so a prefix contributes nothing.

    A word on a prefix row could never appear on an item, so it is a mistake
    rather than an unused option.
    """
    stray = {name: record["Name Word"] for name, record in sheet.items()
             if record["Name Word"] and record["Position"] != "suffix"}
    assert not stray, (
        f"these non-suffix affixes carry a Name Word: {stray}. The first word of "
        "an item's name is its rarity, so only a suffix contributes one."
    )


def test_every_suffix_affix_carries_a_word(sheet) -> None:
    """Or an item rolling it would have nothing to take its name from."""
    missing = [name for name, record in sheet.items()
               if record["Position"] == "suffix" and not record["Name Word"]]
    assert not missing, (
        f"these suffix affixes have no Name Word: {sorted(missing)}"
    )

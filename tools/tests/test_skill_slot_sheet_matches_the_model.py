"""The Skill Slots sheet must agree with the simulation's slot table.

WHY THIS EXISTS. What a skill in each slot is worth, how long it waits and what
it costs now lives in two places: the Skill Slots sheet of
`docs/All_Things_Cataclysm.xlsx`, which is what the game reads by way of
`game/Data/SkillSlots.csv` and the imported DataTable, and `SKILL_SLOTS` in
`sim/cataclysm_sim/character.py`, where the tuning work happens.

Two copies of the same numbers drift. This project has been bitten by exactly
that twice with `sim/cataclysm_sim/scoring.py`, which is why
`sim/verify_scoring_port.py` exists, and once more with the design document's War
Skill Examples table, which still disagrees with the workbook about what the
Ballista and Fortress do.

WHICH IS AUTHORITATIVE. The workbook, as everywhere else in this project. When
this fails, the usual fix is to change the Python to match the sheet.

WHAT WOULD GO UNNOTICED WITHOUT IT. Nothing errors when they disagree. The game
would use one set of cooldowns and every balance figure derived in the
simulation would describe a different game.
"""

from __future__ import annotations

import pathlib

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
SHEET = "Skill Slots"


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def sheet() -> dict[str, dict[str, float]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if SHEET not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {SHEET!r}")

    rows = list(book[SHEET].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out: dict[str, dict[str, float]] = {}
    for raw in rows[1:]:
        if not raw or not text(raw[0]):
            continue
        record = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
        out[text(record["Slot"])] = record
    return out


@pytest.fixture(scope="module")
def model():
    import sys

    sim = str(REPO_ROOT / "sim")
    if sim not in sys.path:
        sys.path.insert(0, sim)
    from cataclysm_sim import character

    return character


def test_the_sheet_has_exactly_the_seven_slots(sheet, model):
    assert set(sheet) == set(model.SKILL_SLOTS), (
        f"sheet has {sorted(sheet)}, the model has {sorted(model.SKILL_SLOTS)}")


def test_every_damage_figure_matches(sheet, model):
    for name, slot in model.SKILL_SLOTS.items():
        row = sheet[name]
        assert row["Damage Percent"] == pytest.approx(slot.typical_damage), name
        assert row["Damage Lowest"] == pytest.approx(slot.lowest), name
        assert row["Damage Highest"] == pytest.approx(slot.highest), name


def test_every_cooldown_matches(sheet, model):
    for name, slot in model.SKILL_SLOTS.items():
        row = sheet[name]
        assert row["Cooldown"] == pytest.approx(slot.typical_cooldown), name
        assert row["Cooldown Lowest"] == pytest.approx(slot.cooldown_lowest), name
        assert row["Cooldown Highest"] == pytest.approx(slot.cooldown_highest), name


def test_every_mana_cost_matches(sheet, model):
    for name, slot in model.SKILL_SLOTS.items():
        assert sheet[name]["Mana Cost"] == pytest.approx(slot.mana_cost), name


def test_only_the_basic_attack_restores_mana_on_hit(sheet, model):
    """It is the automatic slot, so this is income for being in a fight rather
    than a generator the player has to press. Any other slot restoring mana
    would be a different design."""
    for name in model.SKILL_SLOTS:
        expected = (model.BASIC_ATTACK_MANA_ON_HIT
                    if name == model.BASIC_ATTACK_SLOT else 0.0)
        assert sheet[name]["Mana On Hit"] == pytest.approx(expected), name


def test_the_sheet_carries_the_reason_each_slot_is_what_it_is(sheet, model):
    """The note column is what a reader of the sheet alone has to go on."""
    for name, slot in model.SKILL_SLOTS.items():
        assert text(sheet[name]["Note"]) == slot.note.strip(), name

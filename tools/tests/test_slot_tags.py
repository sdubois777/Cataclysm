"""Every skill row carries the gameplay tag for the slot it fills.

WHY THIS MATTERS. `docs/Cataclysm_GDD_v2.md` says increases are scoped by tag and
lists the ability slot tags among the scopes, so an affix reading "increased
Heavy Attack damage" is a modifier scoped to `Slot.Heavy`. Before issue #156 the
Weapon Skills sheet carried only two of the six slot tags -- `Slot.Movement` on
every Movement row and `Slot.Ultimate` on every Ultimate row -- and
`Slot.Heavy`, `Slot.Special`, `Slot.Support` and `Slot.Aura` appeared on no skill
at all. A modifier scoped to any of those four would have applied to nothing, and
nothing would have reported it. That is the same silent failure as issues #120 and
#146.

WHERE THE TAG COMES FROM. The Slot column, derived by `tags_with_slot` in
`tools/generate_datatables.py` when the CSV is written. It is deliberately NOT
written into the sheet: a row states its slot in one place, so the tag cannot
disagree with the column. A slot tag typed into a Tags cell is refused outright
rather than merged, because allowing both would put the slot in two places again.

WHAT IS CHECKED HERE. The generated CSV, `game/Data/WeaponSkills.csv`, which is
what the game reads. Not the sheet, and not a copy of the rules.
"""

from __future__ import annotations

import csv
import pathlib

import pytest
from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
GENERATED = REPO_ROOT / "game" / "Data" / "WeaponSkills.csv"


def generated_rows() -> list[dict[str, str]]:
    with GENERATED.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def tags_of(row: dict[str, str]) -> list[str]:
    return [tag.strip() for tag in row.get("Tags", "").split(",") if tag.strip()]


def declared_tags() -> set[str]:
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True)["Tags"]
    return {
        str(cells[0]).strip()
        for cells in sheet.iter_rows(values_only=True)
        if cells and cells[0] and str(cells[0]).strip() != "Tag Name"
    }


def test_every_row_carries_exactly_one_slot_tag() -> None:
    rows = generated_rows()
    assert rows, f"{GENERATED.name} is empty."

    wrong = [
        (row["Name"], [tag for tag in tags_of(row) if tag.startswith("Slot.")])
        for row in rows
        if len([tag for tag in tags_of(row) if tag.startswith("Slot.")]) != 1
    ]
    assert not wrong, (
        "These rows do not carry exactly one Slot.* tag: "
        + "; ".join(f"{name} has {found}" for name, found in wrong[:10])
    )


def test_the_slot_tag_agrees_with_the_slot_column() -> None:
    """A row tagged Slot.Heavy in the Ultimate slot is the failure this stops."""
    disagreeing = [
        (row["Name"], row["Slot"], tag)
        for row in generated_rows()
        for tag in tags_of(row)
        if tag.startswith("Slot.") and tag != f"Slot.{row['Slot']}"
    ]
    assert not disagreeing, (
        "These rows carry a slot tag that is not their Slot column: "
        + "; ".join(
            f"{name} is in {slot} but tagged {tag}"
            for name, slot, tag in disagreeing[:10]
        )
    )


def test_every_slot_in_use_has_a_declared_tag() -> None:
    """A Slot column value with no tag would generate an undefined tag.

    The generator's own tag validation already refuses that, so this states the
    rule in one place a reader can find rather than leaving it implied.
    """
    declared = declared_tags()
    slots = {row["Slot"] for row in generated_rows() if row["Slot"]}
    missing = sorted(slot for slot in slots if f"Slot.{slot}" not in declared)
    assert not missing, (
        "These slots are used by skill rows but have no tag in the Tags sheet of "
        f"{WORKBOOK.name}: {', '.join(missing)}"
    )


def test_all_six_designed_slots_now_carry_their_tag() -> None:
    """The point of issue #156: four of the six had no tagged skill at all."""
    tagged = {
        tag
        for row in generated_rows()
        for tag in tags_of(row)
        if tag.startswith("Slot.")
    }
    for slot in ("Heavy", "Special", "Support", "Ultimate", "Movement", "Aura"):
        assert f"Slot.{slot}" in tagged, (
            f"No skill row carries Slot.{slot}, so a modifier scoped to it "
            "would apply to nothing."
        )


def test_the_sheet_does_not_write_slot_tags_by_hand() -> None:
    """The Slot column is the only place a row states its slot."""
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True)["Weapon Skills"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    try:
        tags_column = header.index("Tags")
        name_column = header.index("Skill Name")
    except ValueError:  # pragma: no cover - the sheet would be unrecognisable
        pytest.fail(f"The Weapon Skills sheet of {WORKBOOK.name} has no Tags column.")

    offenders = [
        str(row[name_column] or "(unnamed)")
        for row in rows
        if row
        and len(row) > tags_column
        and any(
            part.strip().startswith("Slot.")
            for part in str(row[tags_column] or "").split(",")
        )
    ]
    assert not offenders, (
        "These rows write a slot tag into their Tags cell. It is derived from the "
        f"Slot column instead: {', '.join(offenders[:10])}"
    )

"""The Item Sockets sheet must agree with the simulation's socket maxima.

WHY THIS EXISTS. The most sockets a piece in each slot can have lives in three
places now: the socket table in section VI of `docs/Cataclysm_GDD_v2.md`, the Item
Sockets sheet of `docs/All_Things_Cataclysm.xlsx`, and `MAX_SOCKETS_BY_SLOT` and
`MAX_SOCKETS_BY_WEAPON_HANDS` in `sim/cataclysm_sim/loot.py`.

Three copies of the same twelve numbers drift. This project has been bitten by
exactly that twice with `sim/cataclysm_sim/scoring.py`, which is why
`sim/verify_scoring_port.py` exists.

WHICH IS AUTHORITATIVE. The workbook, as everywhere else. When this fails, the
usual fix is to change the Python to match the sheet.

AND ONE TEST HERE COMPARES NEITHER OF THEM TO THE OTHER. The totals check adds the
sheet up and compares it against the 45 the design document states separately, so
a single mistyped entry is caught even if the sheet and the Python were changed
together.
"""

from __future__ import annotations

import pathlib
import sys

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
SHEET = "Item Sockets"


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def sheet() -> list[dict]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if SHEET not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {SHEET!r}")

    rows = list(book[SHEET].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if not raw or not text(raw[0]):
            continue
        out.append({headers[i]: raw[i]
                    for i in range(len(headers)) if i < len(raw)})
    return out


@pytest.fixture(scope="module")
def model():
    if str(REPO_ROOT / "sim") not in sys.path:
        sys.path.insert(0, str(REPO_ROOT / "sim"))
    from cataclysm_sim import loot

    return loot


def test_the_sheet_was_found_and_covers_every_slot(sheet, model) -> None:
    """Read a real sheet, so the comparisons below cannot pass on an empty one."""
    import cataclysm_sim.affixes as af

    named = {text(row["Slot"]) for row in sheet}
    assert named == set(af.GEAR_SLOTS), (
        f"the {SHEET} sheet names {sorted(named)}, and the gear slots are "
        f"{sorted(af.GEAR_SLOTS)}")


def test_every_socket_maximum_matches(sheet, model) -> None:
    wrong = []
    for row in sheet:
        slot = text(row["Slot"])
        hands = int(row["Hands"])
        in_sheet = int(row["Max Sockets"])

        if slot == "Weapon":
            in_model = model.MAX_SOCKETS_BY_WEAPON_HANDS.get(hands)
            where = f"Weapon ({hands} hands)"
        else:
            in_model = model.MAX_SOCKETS_BY_SLOT.get(slot)
            where = slot

        if in_sheet != in_model:
            wrong.append(f"{where}: sheet {in_sheet}, model {in_model}")

    assert not wrong, (
        f"the {SHEET} sheet and the socket tables in sim/cataclysm_sim/loot.py "
        f"disagree: {'; '.join(wrong)}. The workbook is authoritative, so change "
        "the Python."
    )


def test_the_sheet_adds_up_to_the_total_the_design_states(sheet) -> None:
    """41 across the gear, plus four potion slots, is 45.

    THE ONE TEST HERE THAT IS NOT TWO COPIES COMPARED TO EACH OTHER. Every number
    on this sheet is restated from the design document, and a restated number can
    be mistyped. The 45 is stated separately in the same document, so adding the
    sheet up catches a single wrong entry that matching the Python would not.
    """
    rings_worn = 4 * 2       # eight rings
    potion_sockets = 4

    worn = 0
    for row in sheet:
        slot = text(row["Slot"])
        hands = int(row["Hands"])
        most = int(row["Max Sockets"])

        if slot == "Weapon" and hands == 1:
            continue          # a pair of these equals the two-hander counted below
        worn += most * (rings_worn if slot == "Ring" else 1)

    assert worn + potion_sockets == 45, (
        f"the sheet adds up to {worn} worn sockets plus {potion_sockets} potion "
        f"slots, which is {worn + potion_sockets}. The design document says the "
        "total across all equipment is 45."
    )


def test_two_one_handed_weapons_match_a_two_hander(sheet) -> None:
    """The design's rule, and why a one-hander has three rather than any other
    number: no loadout may be worth more sockets than another."""
    by_hands = {int(row["Hands"]): int(row["Max Sockets"])
                for row in sheet if text(row["Slot"]) == "Weapon"}
    assert by_hands[1] * 2 == by_hands[2], (
        f"two one-handed weapons carry {by_hands[1] * 2} sockets against a "
        f"two-hander's {by_hands[2]}")

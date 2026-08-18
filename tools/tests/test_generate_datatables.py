"""Tests for the DataTable CSV generator.

Most run against fixture workbooks built in the test, so a design change cannot
break them. The last group checks the real workbook, because the committed CSVs
going stale is the drift this tool exists to prevent.
"""

from __future__ import annotations

import pathlib
import sys

import openpyxl
import pytest

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

import generate_datatables as gen  # noqa: E402


def workbook_with(path: pathlib.Path, sheets: dict[str, list[list]]) -> pathlib.Path:
    """Build a workbook from {sheet name: rows}, always including a Tags sheet."""
    book = openpyxl.Workbook()
    book.remove(book.active)
    if "Tags" not in sheets:
        sheets = dict(sheets)
        sheets["Tags"] = [["Tag Name", "Description"],
                          ["Element.War", "Physical"],
                          ["Slot.Ultimate", "Ultimates"]]
    for name, rows in sheets.items():
        sheet = book.create_sheet(name)
        for row in rows:
            sheet.append(row)
    book.save(path)
    return path


class TestRowNames:
    def test_row_names_are_safe_for_fnames(self):
        assert gen.row_name("Fallen City", "Edict of Silence!") == \
            "Fallen_City_Edict_of_Silence"

    def test_empty_parts_are_skipped(self):
        assert gen.row_name("", "Thing", "") == "Thing"

    def test_a_name_that_reduces_to_nothing_is_an_error(self):
        with pytest.raises(gen.DataError, match="could not build a row name"):
            gen.row_name("!!!", "###")

    def test_duplicate_row_names_get_suffixes(self):
        rows = [{"Name": "A"}, {"Name": "A"}, {"Name": "A"}, {"Name": "B"}]
        gen.unique(rows, "test")
        assert [r["Name"] for r in rows] == ["A", "A_1", "A_2", "B"]


class TestNamedCells:
    def test_splits_name_from_description(self):
        assert gen.split_named("Hellfire Aura: burns things") == \
            ("Hellfire Aura", "burns things")

    def test_a_cell_with_no_colon_keeps_the_whole_text(self):
        assert gen.split_named("just a description") == ("", "just a description")


class TestDungeonModifiers:
    def test_reads_rows(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Dungeon Modifiers": [["Cataclysm Type", "Modifier Name", "Weight", "Description"],
                                  ["Demonic", "Hellfire", 20, "burns"]]}))
        rows = gen.dungeon_modifiers(book)
        assert rows == [{"Name": "Demonic_Hellfire", "CataclysmType": "Demonic",
                         "ModifierName": "Hellfire", "Weight": 20.0,
                         "Description": "burns"}]

    def test_rejects_an_unknown_cataclysm(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Dungeon Modifiers": [["Cataclysm Type", "Modifier Name", "Weight", "Description"],
                                  ["Sparkly", "Thing", 1, "d"]]}))
        with pytest.raises(gen.DataError, match="not a Cataclysm type"):
            gen.dungeon_modifiers(book)

    def test_rejects_a_non_numeric_weight(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Dungeon Modifiers": [["Cataclysm Type", "Modifier Name", "Weight", "Description"],
                                  ["Demonic", "Thing", "heavy", "d"]]}))
        with pytest.raises(gen.DataError, match="not a number"):
            gen.dungeon_modifiers(book)


class TestEnchantments:
    """The sheet holds two independent tables side by side."""

    SHEET = [
        ["Positives", "Type", "Weight", "Column 4", None,
         "Negatives", "Type", "Weight", "Tags"],
        ["More damage", "Generic", 1, "Element.War", None,
         "No basic attack", "Generic", 2, "Slot.Ultimate"],
        ["Only a positive", "Generic", 3, "Element.War", None, None, None, None, None],
    ]

    def test_positives_and_negatives_are_read_separately(self, tmp_path):
        book = openpyxl.load_workbook(
            workbook_with(tmp_path / "w.xlsx", {"Enchantments": self.SHEET}))
        positives = gen.enchantments(book, negative=False)
        negatives = gen.enchantments(book, negative=True)
        assert len(positives) == 2
        assert len(negatives) == 1
        assert positives[0]["IsNegative"] == "False"
        assert negatives[0]["IsNegative"] == "True"

    def test_the_negatives_tag_column_is_read(self, tmp_path):
        """Regression: an off-by-one once pointed this at the Weight column,
        so no negative enchantment's tags were ever checked."""
        book = openpyxl.load_workbook(
            workbook_with(tmp_path / "w.xlsx", {"Enchantments": self.SHEET}))
        assert gen.enchantments(book, negative=True)[0]["Tags"] == "Slot.Ultimate"

    def test_a_short_negative_column_does_not_invent_rows(self, tmp_path):
        book = openpyxl.load_workbook(
            workbook_with(tmp_path / "w.xlsx", {"Enchantments": self.SHEET}))
        assert all(r["Effect"] for r in gen.enchantments(book, negative=True))


class TestReshapedSheets:
    def test_enemy_modifiers_matrix_becomes_rows(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Enemy Modifiers": [["Demonic Modifiers", "Death Modifiers"],
                                ["Hellfire: burns", "Haunting: copies"],
                                ["Brute: tanky", None]]}))
        rows = gen.enemy_modifiers(book)
        assert len(rows) == 3
        assert {r["CataclysmType"] for r in rows} == {"Demonic", "Death"}
        assert rows[0]["ModifierName"] == "Hellfire"

    def test_status_effects_treat_the_first_row_as_data(self, tmp_path):
        """These sheets have no header. Skipping row one loses an effect."""
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Buffs": [["Mana Surge: more damage"], ["Warhound: a minion"]],
            "Debuffs": [["Wither: reduces things"]],
            "DoTs": [["Bleed: hurts"]]}))
        rows = gen.status_effects(book)
        assert len(rows) == 4
        assert {r["EffectKind"] for r in rows} == {"Buff", "Debuff", "DoT"}
        assert any(r["EffectName"] == "Mana Surge" for r in rows)


class TestCityUpgradeTiers:
    """The tier cells use four notations and the sheet has no kind column."""

    @pytest.mark.parametrize("cell,effect,kind,value,days", [
        ("0.3",    "Increase max defense by 20%",         "Percent",         0.30, 0),
        ("1",      "Remove 25% of dungeons",              "Percent",         1.00, 0),
        ("10",     "no more than 15 dungeons",            "Flat",           10.00, 0),
        ("8",      "take 4 less days to beat",            "Flat",            8.00, 0),
        ("3x",     "provide 2x more experience",          "Multiplier",      3.00, 0),
        ("10/10%", "Every 20 days ... heal 5%.",          "IntervalPercent", 0.10, 10),
        ("5/15%",  "Every 20 days ... heal 5%.",          "IntervalPercent", 0.15, 5),
        ("",       "anything",                            "",                0.00, 0),
    ])
    def test_each_notation_parses(self, cell, effect, kind, value, days):
        assert gen.parse_tier(cell, effect, "Tier 2", 1) == (kind, value, days)

    def test_the_days_half_of_the_pair_is_not_lost(self):
        """Regression: this notation was once flattened to just the percentage,
        which discarded the change to the trigger interval."""
        _, value, days = gen.parse_tier("5/15%", "Every 20 days ... 5%", "T", 1)
        assert (value, days) == (0.15, 5)

    def test_a_bare_number_is_a_percentage_only_when_the_effect_says_so(self):
        assert gen.parse_tier("10", "increase by 5%", "T", 1)[0] == "Percent"
        assert gen.parse_tier("10", "5 more floors", "T", 1)[0] == "Flat"

    def test_an_unreadable_cell_is_an_error(self):
        with pytest.raises(gen.DataError, match="none of a percentage"):
            gen.parse_tier("banana", "effect", "Tier 2", 7)


class TestOneTimeUseUpgrades:
    SHEET = [["Type", "Tier 1", "Tier 2", "Tier 3"],
             ["Architect", "Increase max defense by 20%", 0.3, 0.4],
             ["Architect*", "Restore defenses by 50%", 0.75, 1],
             ["", "A last resort with no tiers", None, None]]

    def _rows(self, tmp_path):
        book = openpyxl.load_workbook(
            workbook_with(tmp_path / "w.xlsx", {"City Upgrades": self.SHEET}))
        return gen.city_upgrades(book)

    def test_the_asterisk_marks_one_time_use(self, tmp_path):
        rows = self._rows(tmp_path)
        assert [r["IsOneTimeUse"] for r in rows] == ["False", "True", "True"]

    def test_the_asterisk_is_stripped_from_the_branch(self, tmp_path):
        """Nothing downstream should have to parse punctuation to know this."""
        assert self._rows(tmp_path)[1]["Branch"] == "Architect"

    def test_the_unbranched_upgrade_is_kept_and_marked(self, tmp_path):
        row = self._rows(tmp_path)[2]
        assert row["Branch"] == ""
        assert row["BranchUndecided"] == "True"
        assert row["IsOneTimeUse"] == "True"

    def test_a_normal_upgrade_is_not_marked(self, tmp_path):
        row = self._rows(tmp_path)[0]
        assert row["IsOneTimeUse"] == "False"
        assert row["BranchUndecided"] == "False"


class TestGems:
    def test_the_everyday_value_is_read_from_the_effect_text(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Gems": [["Column 1", "Everyday Gemstone", "Quality Gemstone",
                      "Superb Gemstone", "Masterful Gemstone", "Legendary Gemstone",
                      "Mythical Gemstone", "Ascendant Gemstone",
                      "Cataclysmic Gemstone", "Type"],
                     ["Of The Abyss", "10% chance to apply void splinter",
                      0.3, 0.5, 0.75, 1.0, 1.25, 1.5, 2.0, "Attack"]]}))
        row = gen.gems(book)[0]
        assert row["Everyday"] == pytest.approx(0.10)
        assert row["Quality"] == pytest.approx(0.30)
        assert row["Cataclysmic"] == pytest.approx(2.0)

    def test_a_gem_with_no_percentage_in_its_text_is_an_error(self, tmp_path):
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Gems": [["Column 1", "Everyday Gemstone", "Quality Gemstone",
                      "Superb Gemstone", "Masterful Gemstone", "Legendary Gemstone",
                      "Mythical Gemstone", "Ascendant Gemstone",
                      "Cataclysmic Gemstone", "Type"],
                     ["Of Nothing", "does a thing", 0.3, 0.5, 0.75, 1.0, 1.25,
                      1.5, 2.0, "Attack"]]}))
        with pytest.raises(gen.DataError, match="states no percentage"):
            gen.gems(book)

    def test_two_gems_cannot_share_a_name(self, tmp_path):
        """Issue #211. Two different gems were both called Of Recovery.

        One raised health regeneration and one reduced cooldowns. `unique` gave
        the second the row key `Gem_Of_Recovery_1`, so both imported and both
        showed the player the same name, with nothing to tell them apart and
        nothing anywhere reporting a problem. It stood for a month.
        """
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Gems": [["Column 1", "Everyday Gemstone", "Quality Gemstone",
                      "Superb Gemstone", "Masterful Gemstone", "Legendary Gemstone",
                      "Mythical Gemstone", "Ascendant Gemstone",
                      "Cataclysmic Gemstone", "Type"],
                     ["Of Recovery", "Increases hp regen by 10%",
                      0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.6, "Defense"],
                     ["Of Recovery", "Increases CDR by 10%",
                      0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.9, "Utility"]]}))
        with pytest.raises(gen.DataError, match="already the name of the gem"):
            gen.gems(book)

    def test_two_gems_with_different_names_are_fine(self, tmp_path):
        """The other side of it: the check must not refuse ordinary rows.

        A guard that rejected two gems sharing an effect, or two gems in the
        same type, would pass the test above and break the sheet.
        """
        book = openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Gems": [["Column 1", "Everyday Gemstone", "Quality Gemstone",
                      "Superb Gemstone", "Masterful Gemstone", "Legendary Gemstone",
                      "Mythical Gemstone", "Ascendant Gemstone",
                      "Cataclysmic Gemstone", "Type"],
                     ["Of Recovery", "Increases hp regen by 10%",
                      0.15, 0.2, 0.25, 0.3, 0.35, 0.4, 0.6, "Defense"],
                     ["Of Urgency", "Increases CDR by 10%",
                      0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.9, "Utility"]]}))
        names = [row["Name"] for row in gen.gems(book)]
        assert names == ["Gem_Of_Recovery", "Gem_Of_Urgency"], (
            "a gem row key must be readable, not a numeric suffix on someone "
            "else's name")

    @staticmethod
    def one_gem(tmp_path, effect: str, *values):
        """A workbook holding a single gem, so a value can be read back."""
        return openpyxl.load_workbook(workbook_with(tmp_path / "w.xlsx", {
            "Gems": [["Column 1", "Everyday Gemstone", "Quality Gemstone",
                      "Superb Gemstone", "Masterful Gemstone",
                      "Legendary Gemstone", "Mythical Gemstone",
                      "Ascendant Gemstone", "Cataclysmic Gemstone", "Type"],
                     ["Of Testing", effect, *values, "Utility"]]}))

    def test_a_percentage_written_without_a_leading_zero_is_read_whole(
            self, tmp_path):
        """Issue #246. The pattern was `\\d+(?:\\.\\d+)?`, which needs a digit
        before the decimal point, so ".5%" matched only its "5%" and the gem Of
        The Goblin shipped at 0.05 -- ten times its stated value, and larger
        than the six rarity tiers above it."""
        book = self.one_gem(tmp_path, "Increases Magic Find by .5%",
                            0.01, 0.015, 0.02, 0.025, 0.03, 0.035, 0.05)
        assert gen.gems(book)[0]["Everyday"] == pytest.approx(0.005)

    @pytest.mark.parametrize("effect,expected", [
        ("Increases Magic Find by .5%", 0.005),
        ("Increases Magic Find by 0.5%", 0.005),
        ("Increases Mana by 5%", 0.05),
        ("Increases AoE by 10%", 0.10),
        ("Increases movespeed by 1%", 0.01),
        ("20% chance to apply poison", 0.20),
        ("Increases something by 2.5 %", 0.025),
    ])
    def test_every_way_a_percentage_gets_written(self, tmp_path, effect,
                                                 expected):
        """The wider pattern must still read the forms that already worked. A
        fix that only handled the leading dot would be as wrong as the original.
        Values below rise from the largest of these so the ladder check passes
        whichever effect is used."""
        book = self.one_gem(tmp_path, effect,
                            0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9)
        assert gen.gems(book)[0]["Everyday"] == pytest.approx(expected)

    def test_a_gem_that_gets_worse_as_it_gets_rarer_is_an_error(self, tmp_path):
        """The check that would have caught issue #246 whatever caused it.

        Gear and gem rarity equal the difficulty tier, so this ladder is the
        whole of a gem's progression. A tier paying less than the one below it
        means finding a rarer gem is a downgrade, and nothing in the interface
        would say so.
        """
        book = self.one_gem(tmp_path, "Increases Magic Find by 5%",
                            0.01, 0.015, 0.02, 0.025, 0.03, 0.035, 0.05)
        with pytest.raises(gen.DataError, match="does not get better as it gets"):
            gen.gems(book)

    def test_a_gem_that_stalls_at_one_tier_is_also_an_error(self, tmp_path):
        """Equal is not rising. A tier worth exactly what the one below is worth
        is a rarity step a player pays for and gets nothing from."""
        book = self.one_gem(tmp_path, "Increases Magic Find by 1%",
                            0.02, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07)
        with pytest.raises(gen.DataError, match="does not get better as it gets"):
            gen.gems(book)

    def test_the_error_names_the_gem_and_shows_all_eight_values(self, tmp_path):
        """A message giving only 'a gem is wrong' costs whoever reads it a
        search of the sheet."""
        book = self.one_gem(tmp_path, "Increases Magic Find by 5%",
                            0.01, 0.015, 0.02, 0.025, 0.03, 0.035, 0.05)
        with pytest.raises(gen.DataError) as caught:
            gen.gems(book)
        message = str(caught.value)
        assert "Of Testing" in message
        assert "0.05" in message and "0.01" in message

    def test_a_rising_ladder_is_accepted(self, tmp_path):
        """The other side of it. A check that refused ordinary gems would pass
        every test above and break the whole sheet."""
        book = self.one_gem(tmp_path, "Increases Magic Find by .5%",
                            0.01, 0.015, 0.02, 0.025, 0.03, 0.035, 0.05)
        row = gen.gems(book)[0]
        assert row["Everyday"] == pytest.approx(0.005)
        assert row["Cataclysmic"] == pytest.approx(0.05)


class TestValidation:
    def test_an_undefined_tag_is_reported(self):
        tables = {"T": [{"Name": "row", "Tags": "Element.War, Type.Nonexistent"}]}
        problems = gen.validate_tags(tables, {"Element", "Element.War"})
        assert len(problems) == 1
        assert "Type.Nonexistent" in problems[0]

    def test_an_implicit_parent_is_accepted(self):
        tables = {"T": [{"Name": "row", "Tags": "Item.Weapon"}]}
        assert gen.validate_tags(tables, {"Item", "Item.Weapon", "Item.Weapon.Sword"}) == []

    @pytest.mark.parametrize("weight", [0, -1, 101])
    def test_a_weight_outside_the_range_is_reported(self, weight):
        tables = {"T": [{"Name": "row", "Weight": weight}]}
        assert len(gen.validate_weights(tables)) == 1

    def test_a_valid_weight_passes(self):
        assert gen.validate_weights({"T": [{"Name": "r", "Weight": 20}]}) == []


class TestASkillRowCannotNameADamageTypeNobodyHas:
    """ISSUE #579. The Damage Type column of the Weapon Skills sheet was checked
    by nothing at all, one column over from a WeaponType column that has been
    checked since a rename left five weapons with no skills.

    WHY A BAD ONE IS SILENT. `UCataclysmWeaponSkills::SkillsFor` in the engine
    compares the damage type exactly, so a row naming one nobody has is offered
    to nobody: it generates cleanly, imports cleanly, fills no slot and reports
    nothing. One misspelling costs one skill and says so nowhere.
    """

    DECLARED = {"Element.Demonic", "Element.War", "Item.Weapon.Sword"}

    @staticmethod
    def skills(*damage_types: str) -> dict[str, list[dict]]:
        return {"WeaponSkills": [
            {"Name": f"{d}_Sword_Heavy", "WeaponType": "Sword",
             "DamageType": d, "Slot": "Heavy"} for d in damage_types]}

    def test_declared_damage_types_pass(self):
        assert gen.validate_weapon_skill_damage_types(
            self.skills("Demonic", "War"), self.DECLARED) == []

    def test_a_damage_type_nobody_declared_is_reported(self):
        problems = gen.validate_weapon_skill_damage_types(
            self.skills("Demonic", "War", "Demonc"), self.DECLARED)
        assert len(problems) == 1
        assert "'Demonc' is not declared" in problems[0]

    def test_the_wildcard_is_refused_by_name_and_says_why(self):
        """`All` is what the WEAPON column means by every weapon. Somebody
        writing it in this column has guessed at a symmetry that is not there,
        so the refusal says that rather than only listing the legal values."""
        problems = gen.validate_weapon_skill_damage_types(
            self.skills("Demonic", "War", "All"), self.DECLARED)
        assert len(problems) == 1
        assert "is not a wildcard" in problems[0]
        assert "granted to nobody" in problems[0]

    def test_a_damage_type_with_no_rows_at_all_is_reported(self):
        """The other direction. A Cataclysm whose characters have no skills is a
        hole rather than a design choice, which is the same check the weapon
        column has carried since the rename that produced it."""
        problems = gen.validate_weapon_skill_damage_types(
            self.skills("Demonic"), self.DECLARED)
        assert len(problems) == 1
        assert "the War damage type has no rows at all" in problems[0]

    def test_it_reads_the_tags_sheet_rather_than_a_list_written_here(self):
        """So adding a ninth damage type to the design needs no change in
        tools/generate_datatables.py."""
        assert gen.validate_weapon_skill_damage_types(
            self.skills("Rust"), {"Element.Rust"}) == []

    def test_the_real_workbook_passes(self):
        """The check above is worth nothing if it does not run against the
        shipping data. All 398 rows name one of the eight declared types."""
        import openpyxl
        if not gen.WORKBOOK.is_file():
            pytest.skip("the design workbook is not present")
        book = openpyxl.load_workbook(gen.WORKBOOK, data_only=True)
        tables = {"WeaponSkills": gen.weapon_skills(book)}
        assert gen.validate_weapon_skill_damage_types(
            tables, gen.declared_tags(book)) == []


class TestWeaponTagsAreNamedAfterTheWeapon:
    """ISSUE #620. Three weapons carried a tag named after what they used to be
    called -- a Greataxe's rows said `Item.Weapon.2hAxe` -- and nothing compared
    the two, so the old vocabulary survived a rename that corrected everything
    else.

    NOTHING WAS BROKEN BY IT, which is why it lasted. The naming was consistent
    within the data, so every lookup worked and no test failed.
    """

    @staticmethod
    def bases(*weapon_types: str) -> dict[str, list[dict]]:
        return {"ItemBases": [{"Name": f"Weapon_{t}", "WeaponType": t}
                              for t in weapon_types]}

    def test_a_tag_named_after_the_weapon_passes(self):
        assert gen.validate_weapon_tags(
            self.bases("Greataxe"), {"Item.Weapon.Greataxe"}) == []

    def test_the_old_name_is_reported_from_both_sides(self):
        problems = gen.validate_weapon_tags(
            self.bases("Greataxe"), {"Item.Weapon.2hAxe"})
        assert len(problems) == 2
        assert any("Item.Weapon.Greataxe is not declared" in p for p in problems)
        assert any("Item.Weapon.2hAxe names no weapon type" in p
                   for p in problems)

    def test_a_weapon_with_no_tag_at_all_is_reported(self):
        problems = gen.validate_weapon_tags(
            self.bases("Sword", "Whip"), {"Item.Weapon.Sword"})
        assert len(problems) == 1
        assert "Item.Weapon.Whip is not declared" in problems[0]

    def test_a_space_in_a_weapon_type_is_removed_and_nothing_else_is(self):
        """A gameplay tag cannot contain a space and "2H Crossbow" does. The
        letter case is not touched, so the tag is `Item.Weapon.2HCrossbow` and
        not the `2hCrossbow` it used to be."""
        assert gen.weapon_tag_leaf("2H Crossbow") == "2HCrossbow"
        assert gen.validate_weapon_tags(
            self.bases("2H Crossbow"), {"Item.Weapon.2HCrossbow"}) == []
        assert gen.validate_weapon_tags(
            self.bases("2H Crossbow"), {"Item.Weapon.2hCrossbow"}) != []

    def test_tags_outside_the_weapon_prefix_are_ignored(self):
        """It reads `Item.Weapon.*` and nothing else, so the other 166 declared
        tags are not weapons that went missing."""
        assert gen.validate_weapon_tags(
            self.bases("Sword"),
            {"Item.Weapon.Sword", "Item.Slot.Weapon", "Element.War"}) == []

    def test_the_real_workbook_passes(self):
        """The check above is worth nothing if it does not run against the
        shipping data. This is what the four renames were for."""
        import openpyxl
        workbook = gen.WORKBOOK
        if not workbook.is_file():
            pytest.skip("the design workbook is not present")
        book = openpyxl.load_workbook(workbook, data_only=True)
        tables = {"ItemBases": gen.TABLES["ItemBases"](book)}
        assert gen.validate_weapon_tags(tables, gen.declared_tags(book)) == []


class TestShapeParams:
    """A skill's shape names which template runs; its params are that template's numbers.

    EVERY ONE OF THESE GUARDS EXISTS BECAUSE THE SILENT FAILURE IS THE BAD ONE.
    A misspelled parameter would read as zero, and a shape with a radius of zero
    hits nothing: it activates, spends mana, starts its cooldown, and does
    nothing. That is the same failure as issue #155's cooldown of zero, which
    went unnoticed across 77 skills.
    """

    def test_a_well_formed_cell_parses(self):
        assert gen.parse_shape_params(
            "Radius=4; Angle=120; Burn=1", "Strike", "here") == {
                "Radius": "4", "Angle": "120", "Burn": "1"}

    def test_an_empty_cell_is_no_parameters(self):
        assert gen.parse_shape_params("", "Strike", "here") == {}

    def test_a_parameter_the_shape_does_not_read_is_refused(self):
        # Pierce belongs to Projectile. On a Strike it would be ignored, and the
        # designer would have no way to tell it had been.
        with pytest.raises(gen.DataError, match="no parameter 'Pierce'"):
            gen.parse_shape_params("Pierce=3", "Strike", "here")

    def test_a_misspelled_parameter_is_refused(self):
        with pytest.raises(gen.DataError, match="no parameter 'Radiuss'"):
            gen.parse_shape_params("Radiuss=4", "Strike", "here")

    def test_a_rider_is_accepted_on_every_shape(self):
        for shape in gen.SHAPE_PARAMS:
            assert gen.parse_shape_params(
                "GroundRadius=3; GroundDuration=6", shape, "here") == {
                    "GroundRadius": "3", "GroundDuration": "6"}

    def test_a_non_numeric_value_is_refused(self):
        with pytest.raises(gen.DataError, match="not a number"):
            gen.parse_shape_params("Radius=wide", "Strike", "here")

    def test_a_missing_equals_is_refused(self):
        with pytest.raises(gen.DataError, match="not Key=Value"):
            gen.parse_shape_params("Radius 4", "Strike", "here")

    def test_a_repeated_parameter_is_refused(self):
        with pytest.raises(gen.DataError, match="given twice"):
            gen.parse_shape_params("Radius=4; Radius=6", "Strike", "here")

    def test_an_empty_value_is_refused(self):
        with pytest.raises(gen.DataError, match="has no value"):
            gen.parse_shape_params("Radius=", "Strike", "here")

    def test_mode_takes_only_the_three_movement_kinds(self):
        assert gen.parse_shape_params("Mode=Blink", "Movement", "here") == {
            "Mode": "Blink"}
        with pytest.raises(gen.DataError, match="not one of"):
            gen.parse_shape_params("Mode=Teleport", "Movement", "here")

    def test_an_unknown_shape_fails_generation(self, tmp_path):
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            ["Weapon Type", "Damage Type", "Slot", "Skill Name",
             "Skill Description", "Tags", "Shape", "Shape Params"],
            ["Sword", "War", "Heavy", "Cut", "Cuts.", "", "Wiggle", ""]]})
        book = openpyxl.load_workbook(path, data_only=True)
        with pytest.raises(gen.DataError, match="shape 'Wiggle' is not one of"):
            gen.weapon_skills(book)

    def test_parameters_without_a_shape_fail_generation(self, tmp_path):
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            ["Weapon Type", "Damage Type", "Slot", "Skill Name",
             "Skill Description", "Tags", "Shape", "Shape Params"],
            ["Sword", "War", "Heavy", "Cut", "Cuts.", "", "", "Radius=4"]]})
        book = openpyxl.load_workbook(path, data_only=True)
        with pytest.raises(gen.DataError, match="shape parameters but no shape"):
            gen.weapon_skills(book)

    def test_a_shape_without_a_skill_name_fails_generation(self, tmp_path):
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            ["Weapon Type", "Damage Type", "Slot", "Skill Name",
             "Skill Description", "Tags", "Shape", "Shape Params"],
            ["Sword", "War", "Heavy", "", "", "", "Strike", "Radius=4"]]})
        book = openpyxl.load_workbook(path, data_only=True)
        with pytest.raises(gen.DataError, match="shape but no skill name"):
            gen.weapon_skills(book)

    def test_a_sheet_without_the_two_columns_still_generates(self, tmp_path):
        """The 61 War rows predate shapes and must keep generating."""
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            ["Weapon Type", "Damage Type", "Slot", "Skill Name",
             "Skill Description", "Tags"],
            ["Sword", "War", "Heavy", "Cut", "Cuts.", ""]]})
        book = openpyxl.load_workbook(path, data_only=True)
        rows = gen.weapon_skills(book)
        assert rows[0]["Shape"] == "" and rows[0]["ShapeParams"] == ""


class TestASkillsOwnCriticalStrikeChance:
    """The Crit Chance column of the Weapon Skills sheet. Issue #657.

    WHAT WAS MISSING. The design says critical strike chance belongs to the skill
    being used -- its stat source table names "the skill being used" and adds "A
    character has no critical strike chance in the abstract" -- and the sheet had
    nowhere to say it. Every one of the 398 rows took the 5% default and a skill
    designed to critically strike more or less often than average could not be
    built.

    THE SENTINEL IS -1 AND NOT 0, which is what most of these check. The decision
    of 2026-08-04 says the 5% is "a default and not a floor: a skill that states
    1% gets 1%", so a skill built never to critically strike states 0 and must get
    0 rather than silently getting 5.
    """

    HEADERS = ["Weapon Type", "Damage Type", "Slot", "Skill Name",
               "Skill Description", "Tags", "Shape", "Shape Params",
               "Crit Chance"]

    def rows(self, tmp_path, crit):
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            self.HEADERS,
            ["Sword", "War", "Heavy", "Cut", "Cuts.", "", "", "", crit]]})
        book = openpyxl.load_workbook(path, data_only=True)
        return gen.weapon_skills(book)

    def test_a_blank_cell_states_nothing(self, tmp_path):
        assert self.rows(tmp_path, None)[0]["CritChancePercent"] == -1.0

    def test_a_stated_chance_is_carried(self, tmp_path):
        assert self.rows(tmp_path, 12.5)[0]["CritChancePercent"] == 12.5

    def test_zero_is_a_real_answer_and_not_a_blank(self, tmp_path):
        """A skill designed never to critically strike states 0 and gets 0."""
        assert self.rows(tmp_path, 0)[0]["CritChancePercent"] == 0.0

    def test_a_chance_written_as_text_is_read(self, tmp_path):
        """A spreadsheet cell formatted as text still holds a number."""
        assert self.rows(tmp_path, "7")[0]["CritChancePercent"] == 7.0

    def test_something_that_is_not_a_number_fails_generation(self, tmp_path):
        with pytest.raises(gen.DataError, match="is not a number"):
            self.rows(tmp_path, "often")

    def test_a_chance_above_the_hard_cap_fails_generation(self, tmp_path):
        """Otherwise the engine clamps it to 100 and the row lies."""
        with pytest.raises(gen.DataError, match="outside 0 to 100"):
            self.rows(tmp_path, 150)

    def test_a_negative_chance_fails_generation(self, tmp_path):
        """-1 is the generator's own sentinel, not something a row may state."""
        with pytest.raises(gen.DataError, match="outside 0 to 100"):
            self.rows(tmp_path, -1)

    def test_a_sheet_without_the_column_still_generates(self, tmp_path):
        """Every row predates this column, so a short sheet must still work."""
        path = workbook_with(tmp_path / "b.xlsx", {"Weapon Skills": [
            ["Weapon Type", "Damage Type", "Slot", "Skill Name",
             "Skill Description", "Tags"],
            ["Sword", "War", "Heavy", "Cut", "Cuts.", ""]]})
        book = openpyxl.load_workbook(path, data_only=True)
        assert gen.weapon_skills(book)[0]["CritChancePercent"] == -1.0

    def test_every_shipped_row_states_nothing(self):
        """No skill is designed to differ yet, and that is the owner's call.

        This is not a rule -- it is a record of where the data stands. When a
        skill is deliberately given its own chance, change this test with it.
        """
        book = openpyxl.load_workbook(gen.WORKBOOK, data_only=True)
        stated = {r["Name"]: r["CritChancePercent"]
                  for r in gen.weapon_skills(book)
                  if r["CritChancePercent"] != -1.0}
        assert not stated, (
            "these skill rows state a critical strike chance of their own: "
            f"{stated}. That is allowed; update this test to say so.")


class TestBasicAttacks:
    """The basic attack on the weapon base. Issue #524.

    THE SILENT FAILURE HERE IS A CHARACTER WITH NOTHING BETWEEN ITS COOLDOWNS.
    Before this, `game/Data/WeaponSkills.csv` had 398 rows across six slots and
    not one Basic row anywhere, so every character was granted five abilities and
    no ordinary attack. The project owner found it by playing. An empty column is
    exactly as invisible as an empty sheet, so the generator refuses one.
    """

    HEADERS = ["Base Name", "Slot", "Hands", "Sub-Type", "Weapon Type",
               "Max Damage Types", "Implicit 1 Stat", "Implicit 1 Kind",
               "Implicit 1 Value", "Attack Speed", "Basic Shape",
               "Basic Shape Params"]

    def sheet(self, tmp_path, *rows) -> "openpyxl.Workbook":
        path = workbook_with(tmp_path / "b.xlsx",
                             {"Item Bases": [self.HEADERS, *rows]})
        return openpyxl.load_workbook(path, data_only=True)

    def armed(self, **changes) -> list:
        row = ["Sword", "Weapon", 1, "Slashing", "Sword", 4,
               "attack_damage", "flat", 40.0, 1.3,
               "Strike", "Radius=1.8; Angle=90; MaxTargets=1"]
        for column, value in changes.items():
            row[self.HEADERS.index(column)] = value
        return row

    def test_an_armed_weapon_keeps_its_basic_attack(self, tmp_path):
        rows = gen.item_bases(self.sheet(tmp_path, self.armed()))
        assert rows[0]["BasicShape"] == "Strike"
        assert rows[0]["BasicShapeParams"] == "Radius=1.8; Angle=90; MaxTargets=1"

    def test_an_armed_weapon_with_no_basic_attack_is_refused(self, tmp_path):
        """The exact hole issue #524 reported, moved to where it now lives."""
        with pytest.raises(gen.DataError, match="no basic attack shape"):
            gen.item_bases(self.sheet(
                tmp_path, self.armed(**{"Basic Shape": "",
                                        "Basic Shape Params": ""})))

    def test_a_weapon_granting_no_attack_damage_gets_none_and_needs_none(
            self, tmp_path):
        """The Shield. It is a one-handed weapon that grants no attack damage,
        so there is no hit to compose from it. Issue #619."""
        shield = self.armed(**{"Base Name": "Shield", "Weapon Type": "Shield",
                               "Implicit 1 Stat": "block_chance",
                               "Basic Shape": "", "Basic Shape Params": ""})
        rows = gen.item_bases(self.sheet(tmp_path, shield))
        assert rows[0]["BasicShape"] == ""

    def test_a_weapon_granting_no_attack_damage_may_not_state_one(self, tmp_path):
        with pytest.raises(gen.DataError, match="100% of nothing"):
            gen.item_bases(self.sheet(tmp_path, self.armed(
                **{"Base Name": "Shield", "Weapon Type": "Shield",
                   "Implicit 1 Stat": "block_chance"})))

    def test_something_that_is_not_a_weapon_may_not_state_one(self, tmp_path):
        """A glove can grant flat attack damage -- the Vambraces base grants 12 --
        and still have no swing to describe."""
        with pytest.raises(gen.DataError, match="not a weapon"):
            gen.item_bases(self.sheet(tmp_path, self.armed(
                **{"Base Name": "Vambraces", "Slot": "Gloves", "Hands": "",
                   "Weapon Type": "", "Attack Speed": ""})))

    def test_a_shape_a_swing_cannot_take_is_refused(self, tmp_path):
        with pytest.raises(gen.DataError, match="basic attack shape is 'Summon'"):
            gen.item_bases(self.sheet(
                tmp_path, self.armed(**{"Basic Shape": "Summon"})))

    def test_a_misspelled_parameter_is_refused(self, tmp_path):
        with pytest.raises(gen.DataError, match="no parameter 'Radiuss'"):
            gen.item_bases(self.sheet(tmp_path, self.armed(
                **{"Basic Shape Params": "Radiuss=1.8"})))

    def test_a_rider_is_refused_even_though_every_other_slot_may_carry_one(
            self, tmp_path):
        """A basic attack is 100% weapon damage and nothing else, which is what
        makes it the anchor every other slot is a percentage of. `Burn=1` is
        legal on any weapon skill and is not legal here."""
        with pytest.raises(gen.DataError, match=r"carries \['Burn'\]"):
            gen.item_bases(self.sheet(tmp_path, self.armed(
                **{"Basic Shape Params":
                   "Radius=1.8; Angle=90; MaxTargets=1; Burn=1"})))

    def test_parameters_without_a_shape_are_refused(self, tmp_path):
        with pytest.raises(gen.DataError, match="parameters but no shape"):
            gen.item_bases(self.sheet(
                tmp_path, self.armed(**{"Basic Shape": ""})))

    def test_the_two_basic_attack_shape_vocabularies_agree(self):
        """`BASIC_ATTACK_SHAPES` exists in this generator and again in
        `sim/cataclysm_sim/affixes.py`, which enforces the same rule on the
        model. Two copies of a vocabulary drift, so they are compared here the
        same way the two copies of the full shape vocabulary are."""
        from cataclysm_sim import affixes as af

        assert af.BASIC_ATTACK_SHAPES == gen.BASIC_ATTACK_SHAPES, (
            "the basic attack shapes differ between affixes.py and "
            "generate_datatables.py: "
            f"{sorted(af.BASIC_ATTACK_SHAPES ^ gen.BASIC_ATTACK_SHAPES)}")

    def test_the_basic_attack_shapes_are_shapes_the_generator_knows(self):
        """Both names in BASIC_ATTACK_SHAPES appear in SHAPE_PARAMS, so a basic
        attack is validated against a real parameter list rather than against a
        missing dictionary key.

        ONE VOCABULARY IS CHECKED HERE AND THAT IS ENOUGH, because
        test_the_two_basic_attack_shape_vocabularies_agree above proves the copy
        in affixes.py is the same set.

        WHAT THIS DOES NOT CHECK is that a C++ ability template implements the
        shape, which is the stronger property and is not knowable from Python.
        Cataclysm.WeaponSlots.EveryArmedWeaponGrantsABasicAttack asserts it
        against the real table, calling TemplateFor on every armed weapon's
        basic attack shape.
        """
        assert gen.BASIC_ATTACK_SHAPES <= set(gen.SHAPE_PARAMS)


class TestElementVisuals:
    """The eight damage types' effect palette. Issue #549.

    THE SILENT FAILURE HERE IS A COLOUR NOBODY ASKED FOR. `FColor::FromHex` does
    not report bad input, which is how a length-only check in
    ACataclysmTelegraphMarker accepted the word "nonsense" -- eight characters --
    and produced a colour out of it. So the hex is checked character by
    character, and these tests are what say that check works.
    """

    HEADERS = ["Element Tag", "Primary", "Secondary",
               "Emissive Multiplier", "Spawn Rate Scale", "Velocity Scale"]

    def sheet(self, tmp_path, *rows):
        path = workbook_with(tmp_path / "b.xlsx",
                             {"Element Visuals": [self.HEADERS, *rows]})
        return openpyxl.load_workbook(path, data_only=True)

    def test_reads_a_row_and_keys_it_on_the_tags_leaf(self, tmp_path):
        book = self.sheet(tmp_path,
                          ["Element.War", "#FFFFFF", "#000000", 1, 1, 1])
        assert gen.element_visuals(book) == [{
            "Name": "War",
            "ElementTag": "Element.War",
            "PrimaryColour": "(R=1.000000,G=1.000000,B=1.000000,A=1.000000)",
            "SecondaryColour": "(R=0.000000,G=0.000000,B=0.000000,A=1.000000)",
            "EmissiveMultiplier": 1.0,
            "SpawnRateScale": 1.0,
            "VelocityScale": 1.0,
        }]

    def test_the_design_documents_srgb_becomes_linear(self, tmp_path):
        """#FF7A2E is Demonic's primary. Its middle channel is 0x7A, which is
        122/255 = 0.478 in sRGB and 0.195 in linear. Writing the sRGB figure
        into an FLinearColor would render a visibly paler orange."""
        book = self.sheet(tmp_path,
                          ["Element.War", "#FF7A2E", "#000000", 1, 1, 1])
        primary = gen.element_visuals(book)[0]["PrimaryColour"]
        assert primary == "(R=1.000000,G=0.194618,B=0.027321,A=1.000000)"

    def test_a_leading_hash_is_optional(self, tmp_path):
        book = self.sheet(tmp_path,
                          ["Element.War", "FFFFFF", "#000000", 1, 1, 1])
        assert gen.element_visuals(book)[0]["PrimaryColour"].startswith("(R=1.0")

    def test_six_characters_that_are_not_hex_digits_are_refused(self, tmp_path):
        """The real bug this guards. "wrong!" is six characters, so a check on
        the length alone would accept it and produce some colour."""
        book = self.sheet(tmp_path,
                          ["Element.War", "wrong!", "#000000", 1, 1, 1])
        with pytest.raises(gen.DataError, match="not six hex digits"):
            gen.element_visuals(book)

    @pytest.mark.parametrize("text", ["#FFF", "#FFFFFFFF", ""])
    def test_a_hex_of_the_wrong_length_is_refused(self, tmp_path, text):
        book = self.sheet(tmp_path,
                          ["Element.War", text, "#000000", 1, 1, 1])
        with pytest.raises(gen.DataError, match="not six hex digits"):
            gen.element_visuals(book)

    def test_a_key_that_is_not_a_damage_type_tag_is_refused(self, tmp_path):
        book = self.sheet(tmp_path,
                          ["Slot.Ultimate", "#FFFFFF", "#000000", 1, 1, 1])
        with pytest.raises(gen.DataError, match="must start with 'Element.'"):
            gen.element_visuals(book)

    @pytest.mark.parametrize("column", [3, 4, 5])
    @pytest.mark.parametrize("scale", [0, -1])
    def test_a_scale_of_zero_or_less_is_refused(self, tmp_path, column, scale):
        """Each of the three fails as a broken-looking effect rather than as
        bad data: no particles, particles that never move, or a black effect."""
        row = ["Element.War", "#FFFFFF", "#000000", 1, 1, 1]
        row[column] = scale
        book = self.sheet(tmp_path, row)
        with pytest.raises(gen.DataError, match="makes the effect invisible"):
            gen.element_visuals(book)

    def test_an_empty_scale_is_refused(self, tmp_path):
        book = self.sheet(tmp_path,
                          ["Element.War", "#FFFFFF", "#000000", None, 1, 1])
        with pytest.raises(gen.DataError, match="Emissive Multiplier is empty"):
            gen.element_visuals(book)

    def test_a_damage_type_with_no_row_is_reported(self):
        tables = {"ElementVisuals": [{"Name": "War",
                                      "ElementTag": "Element.War"}]}
        problems = gen.validate_element_visuals(
            tables, {"Element.War", "Element.Void", "Slot.Ultimate"})
        assert len(problems) == 1
        assert "Element.Void" in problems[0]

    def test_a_row_naming_an_undeclared_tag_is_reported(self):
        tables = {"ElementVisuals": [{"Name": "Sparkly",
                                      "ElementTag": "Element.Sparkly"}]}
        assert gen.validate_element_visuals(tables, {"Element.Sparkly"}) == []

        problems = gen.validate_element_visuals(tables, {"Element.War"})
        assert len(problems) == 2, problems
        assert any("Element.War" in p and "no effect palette row" in p
                   for p in problems)
        assert any("Element.Sparkly" in p and "not declared" in p
                   for p in problems)

    def test_a_matching_set_reports_nothing(self):
        tables = {"ElementVisuals": [{"Name": "War",
                                      "ElementTag": "Element.War"}]}
        assert gen.validate_element_visuals(
            tables, {"Element.War", "Slot.Ultimate"}) == []


# --------------------------------------------------------------------------
# The three loot sheets, and the word an affix gives an item's name.
#
# EVERY GUARD BELOW IS SHOWN FAILING. A check that cannot fail is worthless, and
# none of these would fail anywhere else: a drop weight of zero, a gate above +10
# or a residue band running backwards all produce a table that loads cleanly and
# carries the wrong numbers.
# --------------------------------------------------------------------------

GEAR_RARITY_HEADER = ["Rarity", "Drop Weight", "Gear Level Gate",
                      "Residue On Drop Lowest", "Residue On Drop Highest"]

#: The real ladder, weakest first. Tests copy this and change one cell.
GEAR_RARITY_ROWS = [
    ["Everyday", 15625, 0, 38, 62],
    ["Quality", 6250, 0, 75, 125],
    ["Superb", 2500, 0, 112, 188],
    ["Masterful", 1000, 0, 150, 250],
    ["Legendary", 125, 4, 188, 312],
    ["Mythical", 25, 6, 225, 375],
    ["Ascendant", 5, 8, 262, 438],
    ["Cataclysmic", 1, 10, 300, 500],
]


def gear_rarity_book(tmp_path, rows):
    return openpyxl.load_workbook(workbook_with(
        tmp_path / "w.xlsx", {"Gear Rarity": [GEAR_RARITY_HEADER] + rows}))


def changed(rows, index, column, value):
    """A copy of `rows` with one cell replaced."""
    out = [list(row) for row in rows]
    out[index][GEAR_RARITY_HEADER.index(column)] = value
    return out


class TestGearRarity:
    def test_reads_the_ladder_weakest_first(self, tmp_path):
        rows = gen.gear_rarity(gear_rarity_book(tmp_path, GEAR_RARITY_ROWS))
        assert [row["Rarity"] for row in rows] == list(gen.RARITY_LADDER)
        assert rows[-1] == {
            "Name": "Cataclysmic", "Rarity": "Cataclysmic",
            "DropWeight": 1.0, "GearLevelGate": 10,
            "ResidueOnDropLowest": 300.0, "ResidueOnDropHighest": 500.0}

    def test_the_ladder_decides_the_order_not_the_sheet(self, tmp_path):
        """A sheet sorted some other way still generates in ladder order, so
        nothing downstream has to trust how the rows happen to be typed."""
        backwards = list(reversed(GEAR_RARITY_ROWS))
        rows = gen.gear_rarity(gear_rarity_book(tmp_path, backwards))
        assert [row["Rarity"] for row in rows] == list(gen.RARITY_LADDER)

    def test_rejects_a_rarity_that_is_not_on_the_ladder(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 0, "Rarity", "Sparkly")
        with pytest.raises(gen.DataError, match="is not a rarity"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_the_same_rarity_twice(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 1, "Rarity", "Everyday")
        with pytest.raises(gen.DataError, match="appears twice"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_ladder_with_a_rarity_missing(self, tmp_path):
        with pytest.raises(gen.DataError, match="has no row for"):
            gen.gear_rarity(gear_rarity_book(tmp_path, GEAR_RARITY_ROWS[:-1]))

    def test_rejects_a_rarity_that_can_never_drop(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 7, "Drop Weight", 0)
        with pytest.raises(gen.DataError, match="can never drop"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_gate_above_the_highest_upgrade_level(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 7, "Gear Level Gate", 11)
        with pytest.raises(gen.DataError, match="outside 0 to 10"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_rarer_thing_that_drops_more_often(self, tmp_path):
        """The one that would look like a working table. Making Cataclysmic
        weigh 2000 keeps every other check happy and inverts the ladder."""
        rows = changed(GEAR_RARITY_ROWS, 7, "Drop Weight", 2000)
        with pytest.raises(gen.DataError, match="drops more often"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_gate_that_falls_as_rarity_rises(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 7, "Gear Level Gate", 2)
        with pytest.raises(gen.DataError, match="is gated lower"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_residue_band_that_runs_backwards(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 0, "Residue On Drop Highest", 10)
        with pytest.raises(gen.DataError, match="runs from"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_drop_carrying_no_residue(self, tmp_path):
        rows = changed(GEAR_RARITY_ROWS, 0, "Residue On Drop Lowest", 0)
        with pytest.raises(gen.DataError, match="Every drop carries some"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))

    def test_rejects_a_residue_band_that_shrinks_as_rarity_rises(self, tmp_path):
        """Both ends are checked, so a band that starts higher and ends lower
        than the rarity below it is still caught."""
        rows = changed(GEAR_RARITY_ROWS, 7, "Residue On Drop Highest", 400)
        with pytest.raises(gen.DataError,
                           match="ResidueOnDropHighest is smaller"):
            gen.gear_rarity(gear_rarity_book(tmp_path, rows))


class TestItemSockets:
    HEADER = ["Slot", "Hands", "Max Sockets"]
    ROWS = [["Head", 0, 2], ["Chest", 0, 6],
            ["Weapon", 1, 3], ["Weapon", 2, 6]]

    def book(self, tmp_path, rows):
        return openpyxl.load_workbook(workbook_with(
            tmp_path / "w.xlsx", {"Item Sockets": [self.HEADER] + rows}))

    def test_a_weapon_gets_one_row_per_hand_count(self, tmp_path):
        rows = gen.item_sockets(self.book(tmp_path, self.ROWS))
        assert [row["Name"] for row in rows] == [
            "Head", "Chest", "Weapon_1H", "Weapon_2H"]
        assert rows[3] == {"Name": "Weapon_2H", "Slot": "Weapon",
                           "Hands": 2, "MaxSockets": 6}

    def test_rejects_a_weapon_taking_three_hands(self, tmp_path):
        with pytest.raises(gen.DataError, match="takes 3 hands"):
            gen.item_sockets(self.book(tmp_path, [["Weapon", 3, 4]]))

    def test_rejects_the_same_slot_and_hand_count_twice(self, tmp_path):
        with pytest.raises(gen.DataError, match="appears twice"):
            gen.item_sockets(self.book(
                tmp_path, [["Head", 0, 2], ["Head", 0, 3]]))

    def test_rejects_a_slot_that_holds_no_gem(self, tmp_path):
        with pytest.raises(gen.DataError, match="no gem could ever go in one"):
            gen.item_sockets(self.book(tmp_path, [["Head", 0, 0]]))


class TestAffixTiers:
    HEADER = ["Tier", "Drop Weight"]
    ROWS = [[1, 64], [2, 32], [3, 16]]

    def book(self, tmp_path, rows):
        return openpyxl.load_workbook(workbook_with(
            tmp_path / "w.xlsx", {"Affix Tiers": [self.HEADER] + rows}))

    def test_reads_a_tier_per_row(self, tmp_path):
        rows = gen.affix_tiers(self.book(tmp_path, self.ROWS))
        assert rows == [{"Name": "T1", "Tier": 1, "DropWeight": 64.0},
                        {"Name": "T2", "Tier": 2, "DropWeight": 32.0},
                        {"Name": "T3", "Tier": 3, "DropWeight": 16.0}]

    def test_rejects_a_gap_in_the_tiers(self, tmp_path):
        """A drop draws from every tier at or below its cap, so a missing tier
        is one the draw has no weight for rather than one that never rolls."""
        with pytest.raises(gen.DataError, match="every tier from 1 upward"):
            gen.affix_tiers(self.book(tmp_path, [[1, 64], [3, 16]]))

    def test_rejects_tiers_that_do_not_start_at_one(self, tmp_path):
        with pytest.raises(gen.DataError, match="every tier from 1 upward"):
            gen.affix_tiers(self.book(tmp_path, [[2, 32], [3, 16]]))

    def test_rejects_a_tier_that_can_never_roll(self, tmp_path):
        with pytest.raises(gen.DataError, match="can never roll"):
            gen.affix_tiers(self.book(tmp_path, [[1, 64], [2, 0]]))

    def test_rejects_a_higher_tier_that_rolls_more_often(self, tmp_path):
        with pytest.raises(gen.DataError, match="rolls more often"):
            gen.affix_tiers(self.book(tmp_path, [[1, 64], [2, 128]]))


class TestTheWordAnAffixGivesAnItemsName:
    """An item is called `<rarity> <base> of <word>`, so only a suffix has one.

    None of these would fail anywhere else. A suffix with no word leaves an item
    that rolled it with nothing to be named after; a word on a prefix can never
    be read, because the first word of the name is the rarity.
    """

    HEADER = ["Affix Name", "Affix Kind", "Position", "Stat", "Value Kind",
              "Top Value", "Breadth", "Ailment", "Gem", "Hybrid Part 1",
              "Hybrid Part 2", "Allowed Slots", "Name Word"]

    def row(self, name, position, word):
        return [name, "Stat", position, "max_health", "flat", 100, None,
                None, None, None, None, "Chest", word]

    def book(self, tmp_path, rows):
        return openpyxl.load_workbook(workbook_with(
            tmp_path / "w.xlsx", {"Affixes": [self.HEADER] + rows}))

    def test_a_suffix_carries_its_word_into_the_table(self, tmp_path):
        rows = gen.affixes(self.book(tmp_path, [
            self.row("Flat life leech", "suffix", "the Leech"),
            self.row("Flat maximum health", "prefix", None)]))
        assert [row["NameWord"] for row in rows] == ["the Leech", ""]

    def test_rejects_a_suffix_with_no_word(self, tmp_path):
        with pytest.raises(gen.DataError, match="has no name word"):
            gen.affixes(self.book(
                tmp_path, [self.row("Flat life leech", "suffix", None)]))

    def test_rejects_a_prefix_that_carries_one(self, tmp_path):
        with pytest.raises(gen.DataError, match="could never appear"):
            gen.affixes(self.book(
                tmp_path, [self.row("Flat maximum health", "prefix", "Vigour")]))

    def test_rejects_two_affixes_sharing_a_word(self, tmp_path):
        """A player reading "of Warding" should find one thing."""
        with pytest.raises(gen.DataError, match="is carried by both"):
            gen.affixes(self.book(tmp_path, [
                self.row("Single resistance", "suffix", "Warding"),
                self.row("Two resistances", "suffix", "Warding")]))


class TestSocketMaximaCoverEverySlotThatIsWorn:
    """The Item Sockets and Item Bases sheets have to describe the same slots.

    A slot missing from Item Sockets leaves a drop with nothing to say how many
    sockets the piece can hold. A slot in Item Sockets that no base occupies is a
    maximum nothing will ever read. Neither is reported anywhere else.
    """

    def tables(self, bases, sockets):
        return {"ItemBases": [{"Name": f"b{i}", "Slot": slot, "Hands": hands}
                              for i, (slot, hands) in enumerate(bases)],
                "ItemSockets": [{"Name": f"s{i}", "Slot": slot, "Hands": hands}
                                for i, (slot, hands) in enumerate(sockets)]}

    def test_a_matching_set_reports_nothing(self):
        pairs = [("Head", 0), ("Weapon", 1), ("Weapon", 2)]
        assert gen.validate_socket_slots(self.tables(pairs, pairs)) == []

    def test_a_worn_slot_with_no_maximum_is_reported(self):
        problems = gen.validate_socket_slots(
            self.tables([("Head", 0), ("Relic", 0)], [("Head", 0)]))
        assert len(problems) == 1, problems
        assert "Relic" in problems[0] and "which item bases occupy" in problems[0]

    def test_a_maximum_for_a_slot_nobody_wears_is_reported(self):
        problems = gen.validate_socket_slots(
            self.tables([("Head", 0)], [("Head", 0), ("Wings", 0)]))
        assert len(problems) == 1, problems
        assert "Wings" in problems[0] and "no item base occupies" in problems[0]

    def test_the_two_hand_counts_of_a_weapon_are_matched_separately(self):
        """A one-handed maximum does not stand in for a two-handed one, which is
        the whole reason a weapon has two rows."""
        problems = gen.validate_socket_slots(
            self.tables([("Weapon", 1), ("Weapon", 2)], [("Weapon", 1)]))
        assert len(problems) == 1, problems
        assert "at 2 hand(s)" in problems[0]


class TestAgainstTheRealWorkbook:
    def test_the_committed_csvs_are_current(self):
        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")
        assert gen.main(["--check"]) == 0, (
            "game/Data/*.csv are out of date. "
            "Run: python tools/generate_datatables.py"
        )

    def test_every_table_has_rows(self):
        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")
        book = openpyxl.load_workbook(gen.WORKBOOK, data_only=True)
        for name, builder in gen.TABLES.items():
            assert len(builder(book)) > 0, f"{name} produced no rows"

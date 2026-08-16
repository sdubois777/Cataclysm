"""An effect the player can apply must be reachable from gear.

WHY THIS FILE EXISTS. Issue #152. Burn was the only one of the ten player-
applicable effects with no gem and no affix behind it. Every one of the sixteen
designed Demonic skills applies burn, and those worked, because a skill applies
its effect outright with no chance roll. But `docs/Cataclysm_GDD_v2.md` says
chance to apply caps at 100% and everything above it becomes magnitude instead,
summed across affixes, gems, keystones and enchantments. With no affix and no
gem, a Demonic character's burn chance from gear was always zero and its
magnitude could never rise above the base. A War character could build around
bleed and a Demonic character could not build around burn.

It went unnoticed for a month because the check that a gem's effect has an affix
was written as a hand-maintained set of names in `sim/cataclysm_sim/affixes.py`.
Burn was in neither the set nor the sheet, so nothing disagreed with anything.

WHAT IS CHECKED. Three rules, all read out of `docs/All_Things_Cataclysm.xlsx`,
which is the authoritative source the DataTables are generated from.

1. Every effect a gem applies has an affix that applies the same effect.
2. Every effect whose own description calls it player-applied has both.
3. An ailment affix's Top Value is exactly five points above the starting chance
   its gem states. That is the rule the nine values that predate this file all
   follow: a 20% gem has a 25 affix, a 15% gem a 20, and a 10% gem a 15. It is
   what makes the value for a new ailment a derivation rather than a free choice.

Rule 2 is the one that would have caught burn.
"""

from __future__ import annotations

import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"

#: What a gem's effect text looks like when it applies an effect on hit:
#: `10% chance to apply bleed`. Gems that grant a stat instead say
#: `Increases Attack Speed by 10%` and match nothing here.
GEM_APPLIES = re.compile(r"^(\d+(?:\.\d+)?)% chance to apply (.+?)\s*$",
                         re.IGNORECASE)

#: How much higher an ailment affix's Top Value is than its gem's chance.
AFFIX_IS_ABOVE_GEM_BY = 5.0

#: The phrase an effect's own description uses to say the player can apply it,
#: as opposed to it only ever being inflicted on the player.
PLAYER_APPLIED = "player-applied"

#: Sheets holding effect descriptions, in the shape `Name: Description`.
EFFECT_SHEETS = ("DoTs", "Debuffs", "Buffs")


def load_sheet(title: str) -> list[list[object]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")
    return [list(row) for row in book[title].iter_rows(values_only=True)]


@pytest.fixture(scope="module")
def gems() -> dict[str, tuple[float, str]]:
    """Gem name to the chance it states and the effect it applies.

    Only gems that apply an effect. A gem granting a stat is not in here.
    """
    rows = load_sheet("Gems")
    found: dict[str, tuple[float, str]] = {}
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        name, effect = str(row[0]).strip(), str(row[1] or "").strip()
        match = GEM_APPLIES.match(effect)
        if match:
            found[name] = (float(match.group(1)), match.group(2).strip())
    assert found, "no gem in the Gems sheet applies an effect on hit"
    return found


@pytest.fixture(scope="module")
def ailment_affixes() -> list[dict[str, object]]:
    """The Ailment rows of the Affixes sheet, keyed by column heading."""
    rows = load_sheet("Affixes")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    found = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        record = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        if str(record.get("Affix Kind", "")).strip() == "Ailment":
            found.append(record)
    assert found, "no Ailment rows in the Affixes sheet"
    return found


@pytest.fixture(scope="module")
def player_applied_effects() -> dict[str, str]:
    """Effect name to description, for effects the player can apply.

    Each row of the Buffs, Debuffs and DoTs sheets is `Name: Description`.
    """
    found: dict[str, str] = {}
    for sheet in EFFECT_SHEETS:
        for row in load_sheet(sheet):
            if not row or row[0] is None:
                continue
            cell = str(row[0]).strip()
            if ":" not in cell:
                continue
            name, description = cell.split(":", 1)
            if PLAYER_APPLIED in description.lower():
                found[name.strip()] = description.strip()
    assert found, (
        f"no effect in {', '.join(EFFECT_SHEETS)} describes itself as "
        f"{PLAYER_APPLIED}. The wording changed; update this test.")
    return found


def test_every_effect_a_gem_applies_has_an_affix(gems, ailment_affixes) -> None:
    from_gems = {effect.casefold() for _, effect in gems.values()}
    from_affixes = {str(a.get("Ailment", "")).strip().casefold()
                    for a in ailment_affixes}
    missing = sorted(from_gems - from_affixes)
    assert not missing, (
        "these effects have a gem in the Gems sheet but no affix in the Affixes "
        f"sheet, so they can only be built by spending a socket: {missing}")


#: The one ailment affix whose Gem cell is empty, and why.
#:
#: Nothing in the Gems sheet applies stun, so there is no gem for the chance to
#: stun to name. Added with issue #298 on 2026-08-16. Written out here so a
#: second empty cell has to be deliberate rather than a row somebody half filled
#: in, and `test_an_affix_with_no_gem_really_has_none_available` below checks the
#: claim rather than trusting it.
AFFIXES_WITH_NO_GEM = {"Chance to stun"}


def test_every_ailment_affix_names_a_gem_that_applies_the_same_effect(
        gems, ailment_affixes) -> None:
    for affix in ailment_affixes:
        name = str(affix.get("Affix Name", "")).strip()
        gem_name = str(affix.get("Gem", "")).strip()
        if name in AFFIXES_WITH_NO_GEM:
            assert not gem_name or gem_name.lower() == "none", (
                f"the {name} affix is listed as having no gem and names "
                f"{gem_name!r}. If a gem now applies it, take it out of "
                f"AFFIXES_WITH_NO_GEM.")
            continue
        assert gem_name in gems, (
            f"the {name} affix names the gem {gem_name!r}, which either does not "
            "exist in the Gems sheet or does not apply an effect on hit")
        _, effect = gems[gem_name]
        ailment = str(affix.get("Ailment", "")).strip()
        assert effect.casefold() == ailment.casefold(), (
            f"the {name} affix applies {ailment!r} but names the {gem_name} gem, "
            f"which applies {effect!r}")


def test_an_ailment_affix_is_five_points_above_its_gem(gems, ailment_affixes) -> None:
    """The rule every value that predates issue #152 follows.

    Without it, the chance on a new ailment affix is a free choice. With it, a
    gem retuned without its affix fails here rather than quietly making one
    ailment cheaper to reach than the rest.
    """
    for affix in ailment_affixes:
        name = str(affix.get("Affix Name", "")).strip()
        gem_name = str(affix.get("Gem", "")).strip()
        if gem_name not in gems:
            continue  # the test above reports this properly
        gem_chance, _ = gems[gem_name]
        top_value = float(affix.get("Top Value") or 0.0)
        assert top_value == gem_chance + AFFIX_IS_ABOVE_GEM_BY, (
            f"{name} has Top Value {top_value} against the {gem_name} gem's "
            f"{gem_chance}%. Every other ailment affix is exactly "
            f"{AFFIX_IS_ABOVE_GEM_BY} points above its gem.")


def test_every_player_applied_effect_has_a_gem(player_applied_effects, gems) -> None:
    """The check that would have caught burn.

    An effect a skill applies but that no gem applies cannot be built around.
    The skill still works, because a skill applies its effect outright with no
    chance roll, which is why this was invisible.
    """
    applied_by_gems = {effect.casefold() for _, effect in gems.values()}
    missing = sorted(name for name in player_applied_effects
                     if name.casefold() not in applied_by_gems)
    assert not missing, (
        "these effects describe themselves as player-applied but no gem applies "
        f"them, so a build cannot raise their chance from gear: {missing}")


def test_every_player_applied_effect_has_an_affix(
        player_applied_effects, ailment_affixes) -> None:
    from_affixes = {str(a.get("Ailment", "")).strip().casefold()
                    for a in ailment_affixes}
    missing = sorted(name for name in player_applied_effects
                     if name.casefold() not in from_affixes)
    assert not missing, (
        "these effects describe themselves as player-applied but no affix "
        f"applies them, so the only gear route to them is a socket: {missing}")


def test_burn_in_particular_is_reachable(gems, ailment_affixes) -> None:
    """Named on its own because it is the one that was missing.

    A test over the general rule passes the moment the general rule is satisfied
    by any means, including by deleting the word player-applied from burn's
    description. This one says burn specifically.
    """
    gem_effects = {effect.casefold(): name for name, (_, effect) in gems.items()}
    assert "burn" in gem_effects, (
        "no gem applies burn. Every one of the sixteen designed Demonic skills "
        "applies it, so a Demonic build has no gear route to it. See issue #152.")

    ailments = {str(a.get("Ailment", "")).strip().casefold(): a
                for a in ailment_affixes}
    assert "burn" in ailments, (
        "no affix applies burn, so burn chance from gear is always zero and its "
        "magnitude can never rise above the base. See issue #152.")


def test_an_affix_with_no_gem_really_has_none_available(gems, ailment_affixes) -> None:
    """What stops AFFIXES_WITH_NO_GEM being a way to skip the check above.

    If a gem is ever added that applies the same effect, this fails and the
    affix has to name it like every other one.
    """
    by_name = {str(a.get("Affix Name", "")).strip(): a for a in ailment_affixes}
    applied = {effect.casefold() for _, effect in gems.values()}

    for name in AFFIXES_WITH_NO_GEM:
        assert name in by_name, (
            f"{name!r} is listed as having no gem and is not an ailment affix "
            f"at all any more, so the exception should go.")
        effect = str(by_name[name].get("Ailment", "")).strip()
        assert effect.casefold() not in applied, (
            f"a gem now applies {effect!r}, so the {name} affix should name it "
            f"rather than being listed as having none.")

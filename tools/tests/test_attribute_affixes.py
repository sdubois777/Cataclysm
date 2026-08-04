"""What an attribute affix is, pinned against every copy of the affix pool.

WHY THIS EXISTS. Gear granting primary attributes was forbidden until
2026-08-04, when the project owner reversed it. `test_what_affixes_do_not_grant.py`
used to enforce the old rule and no longer does; `test_primary_attribute_names.py`
proves the eight names are real. Neither says what an attribute affix *is*, and
that is what this file holds.

THE FOUR RULES, all stated by the owner on the day of the reversal:

    Each of the eight attributes has exactly one affix.
    It is a percentage increase, never flat.
    It is a suffix.
    No hybrid grants an attribute.

WHY PERCENTAGE ONLY IS THE ONE WORTH GUARDING. A flat grant is worth the same to
every character. A percentage increase is worth little to a character spread
across several attributes and a great deal to one that has specialised, so the
affix pays out on a decision the player already made. A flat version added later
"for consistency with the other stats" would quietly undo the design, and it would
look like an improvement while doing it.

WHERE THE SLOTS COME FROM, and why that is checked rather than listed. An
attribute affix rolls wherever the stats that attribute drive already roll. Nothing
was chosen by hand. That is what keeps a weapon offensive without a rule about
weapons: Ferocity drives critical strike and Efficacy drives area of effect, both
of which already roll on a weapon, while Vitality drives health and Constitution
drives armour, which do not. Listing the slots here instead would let the derived
rule and the data drift apart without anything noticing.

WHAT IT CHECKS AND WHERE. Both copies of the pool that a person or the game reads:

    docs/All_Things_Cataclysm.xlsx, Affixes sheet   authoritative, hand-edited
    game/Data/Affixes.csv                           generated, what the game loads

`sim/cataclysm_sim/affixes.py` is the third copy and is checked by
`sim/tests/test_affixes.py`, which can reach its objects directly.
"""

from __future__ import annotations

import collections
import csv
import pathlib

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
AFFIX_CSV = REPO_ROOT / "game" / "Data" / "Affixes.csv"
ATTRIBUTE_CSV = REPO_ROOT / "game" / "Data" / "Attributes.csv"

#: The two attributes whose driven stats already roll on a weapon. Asserted
#: rather than assumed, because it is the part of the derived rule a reader is
#: most likely to doubt.
ATTRIBUTES_ON_A_WEAPON = frozenset({"ferocity", "efficacy"})


def text(value) -> str:
    return "" if value is None else str(value).strip()


def read_csv(path: pathlib.Path) -> list[dict[str, str]]:
    if not path.is_file():
        pytest.skip(f"{path.name} not present")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_sheet(title: str) -> list[dict[str, object]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")
    rows = list(book[title].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if not raw or raw[0] is None or not text(raw[0]):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers))
                    if i < len(raw)})
    return out


@pytest.fixture(scope="module")
def attributes() -> frozenset[str]:
    """The eight names, read from the data rather than listed here.

    `test_primary_attribute_names.py` is what proves these are the names the
    game actually runs on.
    """
    return frozenset(row["Attribute"].strip().lower()
                     for row in read_csv(ATTRIBUTE_CSV))


@pytest.fixture(scope="module")
def affix_rows() -> list[dict[str, str]]:
    return read_csv(AFFIX_CSV)


@pytest.fixture(scope="module")
def affix_sheet() -> list[dict[str, object]]:
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Affixes")


@pytest.fixture(scope="module")
def attribute_affixes(affix_rows, attributes) -> list[dict[str, str]]:
    return [row for row in affix_rows
            if row["Stat"].strip().lower() in attributes]


class TestEveryAttributeIsGrantedExactlyOnce:
    """An attribute with no affix cannot be geared for. One with two doubles how
    much gear can give it without anyone deciding to."""

    def test_in_the_table_the_game_loads(self, attribute_affixes, attributes):
        granted = {row["Stat"].strip().lower() for row in attribute_affixes}
        assert granted == attributes, (
            f"{AFFIX_CSV.name} is missing affixes for "
            f"{sorted(attributes - granted)}")

    def test_in_the_workbook(self, affix_sheet, attributes):
        granted = {text(row["Stat"]).lower() for row in affix_sheet} & attributes
        assert granted == attributes, (
            f"the Affixes sheet of {WORKBOOK.name} is missing affixes for "
            f"{sorted(attributes - granted)}")

    def test_none_is_granted_twice(self, attribute_affixes):
        counts = collections.Counter(row["Stat"].strip().lower()
                                     for row in attribute_affixes)
        repeated = sorted(name for name, n in counts.items() if n > 1)
        assert not repeated, f"{AFFIX_CSV.name} grants {repeated} more than once"


class TestTheShapeOfAnAttributeAffix:

    def test_it_is_always_an_increase_and_never_flat(self, attribute_affixes):
        """The rule the whole design rests on. See the module docstring."""
        for row in attribute_affixes:
            assert row["ValueKind"].strip() == "increased", (
                f"{row['AffixName']} grants an attribute as "
                f"{row['ValueKind']!r}. Attribute affixes are percentage "
                "increases only. A flat one hands every character the same "
                "value instead of rewarding specialisation, which is the "
                "design.")

    def test_it_is_always_a_suffix(self, attribute_affixes):
        for row in attribute_affixes:
            assert row["Position"].strip() == "suffix", (
                f"{row['AffixName']} is a {row['Position']}; attribute affixes "
                "are suffixes")

    def test_they_all_share_one_top_value(self, attribute_affixes):
        """Eight affixes of the same kind with different ceilings would make one
        attribute better to gear for than another for no stated reason."""
        values = {row["TopValue"].strip() for row in attribute_affixes}
        assert len(values) == 1, (
            f"attribute affixes have different top values: {sorted(values)}")

    def test_no_hybrid_reaches_an_attribute_through_its_halves(self, affix_rows,
                                                               attributes):
        """A hybrid names two other affixes rather than a stat, so reading the
        `Stat` column alone would miss one reaching an attribute sideways."""
        by_name = {row["AffixName"].strip(): row for row in affix_rows}
        for row in affix_rows:
            for column in ("HybridPart1", "HybridPart2"):
                part = row[column].strip()
                if not part:
                    continue
                assert part in by_name, (
                    f"{row['AffixName']} names a half, {part!r}, that is not an "
                    "affix")
                stat = by_name[part]["Stat"].strip().lower()
                assert stat not in attributes, (
                    f"{row['AffixName']} reaches the primary attribute {stat!r} "
                    f"through its half {part!r}. Hybrids granting an attribute "
                    "were ruled out on 2026-08-04.")


class TestTheSlotsAreDerivedFromTheStatsTheAttributeDrives:

    def test_each_one_rolls_where_its_driven_stats_roll(self, affix_rows,
                                                        attribute_affixes,
                                                        attributes):
        drives: dict[str, list[str]] = collections.defaultdict(list)
        for row in read_csv(ATTRIBUTE_CSV):
            drives[row["Attribute"].strip().lower()].append(
                row["Stat"].strip().lower())

        slots_for_stat: dict[str, set[str]] = collections.defaultdict(set)
        for row in affix_rows:
            stat = row["Stat"].strip().lower()
            if not stat or stat in attributes:
                continue
            slots_for_stat[stat].update(
                part.strip() for part in row["AllowedSlots"].split(",")
                if part.strip())

        for row in attribute_affixes:
            attribute = row["Stat"].strip().lower()
            expected: set[str] = set()
            for stat in drives[attribute]:
                expected |= slots_for_stat.get(stat, set())
            actual = {part.strip() for part in row["AllowedSlots"].split(",")
                      if part.strip()}
            assert actual == expected, (
                f"{row['AffixName']} rolls on {sorted(actual)} but the stats it "
                f"drives roll on {sorted(expected)}. The slot list is derived, "
                "not chosen; change the driven stats or the derivation, not "
                "this row.")

    def test_only_the_two_offensive_attributes_reach_a_weapon(self,
                                                              attribute_affixes):
        """A consequence of the rule above, asserted directly because it is the
        part of it a reader is most likely to doubt."""
        on_weapon = {row["Stat"].strip().lower() for row in attribute_affixes
                     if "Weapon" in row["AllowedSlots"]}
        assert on_weapon == ATTRIBUTES_ON_A_WEAPON, (
            f"attributes reaching a weapon are {sorted(on_weapon)}, expected "
            f"{sorted(ATTRIBUTES_ON_A_WEAPON)}")

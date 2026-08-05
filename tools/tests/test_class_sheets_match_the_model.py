"""The Class Stats and Attributes sheets must agree with the simulation.

Same reason as `test_affix_sheets_match_the_model.py`. The class stat lines and
the attribute effects now live in two places: the workbook, which is edited and
which the Unreal DataTables are generated from, and
`sim/cataclysm_sim/classes.py` and `character.py`, where the tuning happens. Two
copies of the same numbers drift.

WHICH IS AUTHORITATIVE. The workbook. When this fails, the usual fix is to change
the Python to match the sheet.

UNITS DIFFER ON PURPOSE, and that is the one thing to be careful of. The sheet
stores an attribute effect as PERCENT PER POINT, so Vitality reads 2. The model
stores the same figure as a fraction, 0.02. The conversion is done here in the
open rather than hidden in either.

WHAT IS NOT COMPARED. `ClassDefinition.spends_health` is declared in
`character.py` and read nowhere, so it was not exported and there is nothing to
compare. If it ever gains a meaning it needs a column and a case here.
"""

from __future__ import annotations

import pathlib

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"

DEFAULT_CLASS = "Default"


def read_sheet(title: str) -> list[dict[str, object]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")

    rows = list(book[title].iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [{headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
            for raw in rows[1:]
            if raw and raw[0] is not None and str(raw[0]).strip()]


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def model():
    from cataclysm_sim import character
    return character


@pytest.fixture(scope="module")
def classes():
    from cataclysm_sim.classes import DEMONIC_CLASSES
    return DEMONIC_CLASSES


@pytest.fixture(scope="module")
def class_sheet():
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Class Stats")


@pytest.fixture(scope="module")
def attribute_sheet():
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Attributes")


class TestTheDefaultStatLine:
    def test_the_sheet_and_the_model_name_the_same_stats(self, class_sheet, model):
        in_sheet = {text(r["Stat"]) for r in class_sheet
                    if text(r["Class"]) == DEFAULT_CLASS}
        in_model = {stat for stat, scaling in model.DEFAULT_STAT_LINE.items()
                    if scaling.base or scaling.per_level}
        assert in_sheet == in_model, (
            f"only in the sheet: {sorted(in_sheet - in_model)}; "
            f"only in character.py: {sorted(in_model - in_sheet)}")

    def test_every_default_value_matches(self, class_sheet, model):
        for row in class_sheet:
            if text(row["Class"]) != DEFAULT_CLASS:
                continue
            scaling = model.DEFAULT_STAT_LINE[text(row["Stat"])]
            assert float(row["Base"]) == pytest.approx(scaling.base), row["Stat"]
            assert float(row["Per Level"]) == pytest.approx(scaling.per_level), row["Stat"]

    def test_a_stat_of_all_zeroes_is_left_out(self, class_sheet):
        """A row of zeroes says nothing, and every stat not named resolves to
        zero anyway."""
        for row in class_sheet:
            assert float(row["Base"]) or float(row["Per Level"]), (
                f"{row['Class']} {row['Stat']} is zero in both columns")


class TestClassOverrides:
    def test_the_same_classes_exist_in_both(self, class_sheet, classes):
        in_sheet = {text(r["Class"]) for r in class_sheet} - {DEFAULT_CLASS}
        assert in_sheet == set(classes), (
            f"only in the sheet: {sorted(in_sheet - set(classes))}; "
            f"only in classes.py: {sorted(set(classes) - in_sheet)}")

    def test_each_class_overrides_the_same_stats(self, class_sheet, classes):
        for name, definition in classes.items():
            in_sheet = {text(r["Stat"]) for r in class_sheet
                        if text(r["Class"]) == name}
            assert in_sheet == set(definition.overrides), (
                f"{name}: only in the sheet "
                f"{sorted(in_sheet - set(definition.overrides))}; only in "
                f"classes.py {sorted(set(definition.overrides) - in_sheet)}")

    def test_every_override_value_matches(self, class_sheet, classes):
        for row in class_sheet:
            name = text(row["Class"])
            if name == DEFAULT_CLASS:
                continue
            scaling = classes[name].overrides[text(row["Stat"])]
            assert float(row["Base"]) == pytest.approx(scaling.base), row
            assert float(row["Per Level"]) == pytest.approx(scaling.per_level), row

    def test_no_class_sets_the_same_stat_twice(self, class_sheet):
        seen = set()
        for row in class_sheet:
            key = (text(row["Class"]), text(row["Stat"]))
            assert key not in seen, f"{key} appears twice"
            seen.add(key)


class TestAttributeEffects:
    def test_the_same_attributes_exist_in_both(self, attribute_sheet, model):
        in_sheet = {text(r["Attribute"]) for r in attribute_sheet}
        assert in_sheet == set(model.ATTRIBUTE_NAMES)

    def test_each_attribute_raises_the_same_stats(self, attribute_sheet, model):
        for attribute in model.ATTRIBUTE_NAMES:
            in_sheet = {text(r["Stat"]) for r in attribute_sheet
                        if text(r["Attribute"]) == attribute}
            assert in_sheet == set(model.ATTRIBUTE_EFFECTS[attribute]), attribute

    def test_every_value_matches_after_converting_units(self, attribute_sheet, model):
        """The sheet stores percent per point, the model stores a fraction."""
        for row in attribute_sheet:
            attribute = text(row["Attribute"])
            stat = text(row["Stat"])
            as_fraction = float(row["Percent Per Point"]) / 100.0
            assert as_fraction == pytest.approx(
                model.ATTRIBUTE_EFFECTS[attribute][stat]), (attribute, stat)

    def test_no_attribute_effect_is_zero_or_negative(self, attribute_sheet):
        """A point that gives nothing, or takes something away, is not an
        attribute the design has."""
        for row in attribute_sheet:
            assert float(row["Percent Per Point"]) > 0, row


class TestTheCountsThatAreAssertedInUnreal:
    """`game/Source/Cataclysm/Tests/CataclysmDataTableTests.cpp` pins these row
    counts by hand. No Python test can catch a stale number there, and it has
    been missed before, so both are stated here as well."""

    def test_the_class_stat_row_count(self, class_sheet):
        # 30 until issue #243 added a Default row for loot_quantity, which
        # needed a baseline of 100 rather than zero. 33 since issue #205 added
        # Default rows for dot_damage and dot_duration, which need a baseline of
        # 100 for the same reason.
        assert len(class_sheet) == 33

    def test_the_attribute_effect_row_count(self, attribute_sheet):
        assert len(attribute_sheet) == 17

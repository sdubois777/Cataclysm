"""The two things the design says an affix must never be.

WHY THIS EXISTS. Issue #204 reported that none of the eight primary attributes
can be found on gear, and read that as an omission. It is not: the design
document already decides it. `docs/Cataclysm_GDD_v2.md` has a section headed
"What Affixes Do Not Grant" holding exactly two rules, and `docs/DECISIONS.md`
records the reasoning behind the first of them.

    There are no attribute affixes. The design gives one attribute point per
    level, and the Maw consumes items and enemies for more. Gear granting
    attribute points appears nowhere, so an affix for it would be a new mechanic
    rather than a filled gap.

    No ordinary affix is a "more" multiplier. An affix is flat or increased.
    Multiplicative sources come from gems, passive tree keystones and
    enchantments, as section IV states.

THE POINT IS DISCOVERABILITY. Both rules were already written down in prose when
#204 was filed. The audit that produced #204 cross-referenced the attribute sets
against the `Stat` column of `game/Data/Affixes.csv`, which is a data-level
reading that never touches the design document, so it found a hole where the
design had put a decision. A rule that lives only in prose gets re-reported. This
turns both rules into something a data-level reading trips over.

IT ALSO GUARDS THE DECISION ITSELF. If either rule is ever reversed, this test
fails, and the fix is to change the design document and this file together rather
than to let the data and the prose disagree.

WHAT IT CHECKS AND WHERE. Three copies of the affix pool exist and all three are
checked, because an affix added to one and not the others is the other failure
mode:

    docs/All_Things_Cataclysm.xlsx, Affixes sheet   authoritative, hand-edited
    game/Data/Affixes.csv                           generated, what the game loads
    sim/cataclysm_sim/affixes.py                    the model the tuning uses

WHY THE ATTRIBUTE NAMES ARE READ RATHER THAN LISTED. A hand-written list of the
eight would keep passing if an attribute were renamed, because the new name would
appear in no list and so could never be found in an affix. The names come from
`game/Data/Attributes.csv`, the generated table saying what a point of each
attribute does, and are cross-checked against
`game/Source/Cataclysm/AbilitySystem/CataclysmPrimaryAttributeSet.h`, the C++
attribute set the game actually runs on. A rename in either place fails the
cross-check instead of silently emptying the guard.
"""

from __future__ import annotations

import csv
import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
AFFIX_CSV = REPO_ROOT / "game" / "Data" / "Affixes.csv"
ATTRIBUTE_CSV = REPO_ROOT / "game" / "Data" / "Attributes.csv"
ATTRIBUTE_SET_HEADER = (REPO_ROOT / "game" / "Source" / "Cataclysm" /
                        "AbilitySystem" / "CataclysmPrimaryAttributeSet.h")
GDD = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"

#: The section of the design document these rules come from. Named here so a
#: failure says where to go, and asserted below so renaming the section does not
#: leave this file pointing at nothing.
GDD_SECTION = "What Affixes Do Not Grant"

#: The two value kinds an ordinary affix may have. "more" is deliberately absent.
ALLOWED_VALUE_KINDS = frozenset({"flat", "increased"})


def text(value) -> str:
    return "" if value is None else str(value).strip()


def read_sheet(title: str) -> list[dict[str, object]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")

    rows = list(book[title].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if not raw or raw[0] is None or not text(raw[0]):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers))
                    if i < len(raw)})
    return out


def read_csv(path: pathlib.Path) -> list[dict[str, str]]:
    if not path.is_file():
        pytest.skip(f"{path.name} not present")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


@pytest.fixture(scope="module")
def affix_sheet():
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Affixes")


@pytest.fixture(scope="module")
def affix_rows():
    return read_csv(AFFIX_CSV)


@pytest.fixture(scope="module")
def primary_attributes() -> frozenset[str]:
    """The eight primary attributes, lower-cased, read from the data.

    Read rather than listed so that renaming one cannot quietly empty this
    file's checks. `test_the_attribute_names_come_from_the_game` is what makes
    that claim hold.
    """
    rows = read_csv(ATTRIBUTE_CSV)
    return frozenset(row["Attribute"].strip().lower() for row in rows)


class TestTheAttributeNamesAreReal:
    """Everything below is worthless if these eight names are wrong."""

    def test_there_are_eight_of_them(self, primary_attributes):
        assert len(primary_attributes) == 8, sorted(primary_attributes)

    def test_the_attribute_names_come_from_the_game(self, primary_attributes):
        """The names in the generated table are the ones the C++ runs on.

        `CataclysmPrimaryAttributeSet.h` declares each attribute as an
        `FGameplayAttributeData`. If the two lists ever disagree, one of them has
        been renamed and this whole file is checking for names that no longer
        exist.
        """
        if not ATTRIBUTE_SET_HEADER.is_file():
            pytest.skip("the Unreal attribute set header is not present")
        declared = {name.lower() for name in re.findall(
            r"FGameplayAttributeData\s+(\w+);",
            ATTRIBUTE_SET_HEADER.read_text(encoding="utf-8"))}
        assert declared == set(primary_attributes), (
            f"only in {ATTRIBUTE_SET_HEADER.name}: "
            f"{sorted(declared - primary_attributes)}; "
            f"only in {ATTRIBUTE_CSV.name}: "
            f"{sorted(primary_attributes - declared)}")

    def test_no_attribute_shares_a_name_with_a_stat(self, primary_attributes):
        """An attribute and a stat must not collide.

        `Attributes.csv` maps each attribute to the stats it scales. If an
        attribute were ever named after one of those stats, an affix granting the
        stat would look like an affix granting the attribute and the checks below
        would fail for the wrong reason.
        """
        scaled = {row["Stat"].strip().lower() for row in read_csv(ATTRIBUTE_CSV)}
        assert not (scaled & primary_attributes), sorted(scaled & primary_attributes)


class TestNoAffixGrantsAPrimaryAttribute:
    """The first rule. See the module docstring for why it is a decision."""

    def test_not_in_the_workbook(self, affix_sheet, primary_attributes):
        granted = sorted({text(row["Stat"]).lower() for row in affix_sheet}
                         & primary_attributes)
        assert not granted, (
            f"the Affixes sheet of {WORKBOOK.name} grants primary attributes "
            f"{granted}. The design document forbids that under "
            f"'{GDD_SECTION}'. Reversing the decision means changing the design "
            "document and this test together.")

    def test_not_in_the_table_the_game_loads(self, affix_rows, primary_attributes):
        granted = sorted({row["Stat"].strip().lower() for row in affix_rows}
                         & primary_attributes)
        assert not granted, (
            f"{AFFIX_CSV.name} grants primary attributes {granted}")

    def test_not_in_the_hybrid_halves_either(self, affix_rows, primary_attributes):
        """A hybrid names two other affixes rather than a stat.

        Checking only the `Stat` column would miss an attribute reaching gear
        through a hybrid whose halves grant one. The halves are affix names, so
        this looks for an attribute name inside them.
        """
        by_name = {row["AffixName"].strip(): row for row in affix_rows}
        for row in affix_rows:
            for column in ("HybridPart1", "HybridPart2"):
                part = row[column].strip()
                if not part:
                    continue
                assert part in by_name, (
                    f"{row['AffixName']} names a half, {part!r}, that is not an "
                    "affix")
                stat = by_name[part]["Stat"].strip().lower()
                assert stat not in primary_attributes, (
                    f"{row['AffixName']} reaches the primary attribute {stat!r} "
                    f"through its half {part!r}")

    def test_the_model_would_reject_one(self, primary_attributes):
        """`affixes.py` validates the stat an affix names on construction.

        A stat outside `AFFIXABLE_STATS` raises, so the model cannot be given an
        attribute affix even by hand. That is the third copy of the pool and the
        one the tuning work reads.
        """
        from cataclysm_sim import affixes as af

        reachable = sorted(set(af.AFFIXABLE_STATS) & primary_attributes)
        assert not reachable, (
            f"affixes.AFFIXABLE_STATS admits primary attributes {reachable}")

        with pytest.raises(ValueError, match="character sheet"):
            af.StatAffix("Flat agility", sorted(primary_attributes)[0], "flat",
                         10.0, af.DEFENSIVE_SLOTS, af.PREFIX)


class TestNoAffixIsAMoreMultiplier:
    """The second rule. An affix is flat or increased, never multiplicative."""

    def test_every_stat_affix_in_the_workbook_is_flat_or_increased(self, affix_sheet):
        wrong = sorted({text(row["Value Kind"]) for row in affix_sheet
                        if text(row["Affix Kind"]) == "Stat"}
                       - ALLOWED_VALUE_KINDS)
        assert not wrong, (
            f"the Affixes sheet uses value kinds {wrong}. The design document "
            f"allows only {sorted(ALLOWED_VALUE_KINDS)} under '{GDD_SECTION}': "
            "multiplicative sources are gems, keystones and enchantments.")

    def test_every_stat_affix_the_game_loads_is_flat_or_increased(self, affix_rows):
        wrong = sorted({row["ValueKind"].strip() for row in affix_rows
                        if row["AffixKind"].strip() == "Stat"}
                       - ALLOWED_VALUE_KINDS)
        assert not wrong, f"{AFFIX_CSV.name} uses value kinds {wrong}"

    def test_there_are_stat_affixes_of_both_kinds_to_check(self, affix_rows):
        """A check over an empty set passes and proves nothing."""
        kinds = [row["ValueKind"].strip() for row in affix_rows
                 if row["AffixKind"].strip() == "Stat"]
        assert kinds.count("flat") >= 1 and kinds.count("increased") >= 1


def test_the_design_document_still_states_both_rules():
    """The rules this file pins must still be the design.

    Without this, deleting the section from `docs/Cataclysm_GDD_v2.md` would
    leave the tests enforcing a rule the design no longer holds, which is the
    same disagreement between prose and data that produced issue #204.
    """
    if not GDD.is_file():
        pytest.skip("the design document is not present")
    gdd = GDD.read_text(encoding="utf-8")
    assert GDD_SECTION in gdd, (
        f"{GDD.name} no longer has a '{GDD_SECTION}' section")
    assert "There are no attribute affixes." in gdd
    assert 'No ordinary affix is a "more" multiplier.' in gdd

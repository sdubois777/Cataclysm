"""A self buff's magnitude is written twice, and the two must agree.

WHY THIS FILE EXISTS. Burning Wrath's description says "4% more fire damage
for every enemy currently burning within 15 meters". Three of those numbers are
now also machine-readable, in the same row's Shape Params cell, because
`UCataclysmSelfBuffSkill` has to be given them as numbers rather than prose:
`Duration=10; Radius=15; Burn=1; MoreDamagePer=4; ScalingSource=Burning`.

Prose and parameters drifting apart is the failure this stops. A player reads the
description and the skill does what the parameters say, so a description saying
4% while the cell says 3 is a lie the game tells with a straight face and nothing
else would catch.

NOTHING HERE IS A COPY OF EITHER NUMBER. Both are read out of
`docs/All_Things_Cataclysm.xlsx`, the workbook the design lives in, and compared
against each other. Changing the design in the sheet is enough.
"""

from __future__ import annotations

import pathlib
import re

import pytest
from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"

#: The one designed skill whose buff has a magnitude the code applies.
BURNING_WRATH = "Burning Wrath"


def weapon_skill_rows() -> list[dict[str, str]]:
    """Every row of the Weapon Skills sheet, by column heading."""
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True)["Weapon Skills"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    # strict=False: openpyxl stops a row at its last filled cell, so a row whose
    # trailing columns are blank is shorter than the header. Those columns are
    # simply absent from the dictionary, which every reader here copes with.
    return [
        dict(zip(header, [str(cell or "") for cell in row], strict=False))
        for row in rows
    ]


def row_named(name: str) -> dict[str, str]:
    for row in weapon_skill_rows():
        if row.get("Skill Name") == name:
            return row
    pytest.fail(f"No row named {name!r} in the Weapon Skills sheet of {WORKBOOK.name}.")


def shape_params(row: dict[str, str]) -> dict[str, str]:
    """The `Key=Value; Key=Value` cell, as a dictionary."""
    params: dict[str, str] = {}
    for piece in row.get("Shape Params", "").split(";"):
        piece = piece.strip()
        if "=" in piece:
            key, value = (part.strip() for part in piece.split("=", 1))
            params[key] = value
    return params


def test_burning_wrath_carries_the_increase_its_description_states() -> None:
    """The 4% in the prose and the 4 in MoreDamagePer are one number.

    THE KEY CHANGED ON 2026-09-01 and so did the word in the prose. It was
    `IncreasePerBurning=4` against "4% increased fire damage", which put a self
    buff in the additive bucket where it competed with every gear affix the
    character wore. It is now `MoreDamagePer=4; ScalingSource=Burning` against
    "4% more fire damage", and `UCataclysmSelfBuffSkill` builds the modifier in
    the multiplicative bucket.

    THE SOURCE IS CHECKED AS WELL AS THE NUMBER. `MoreDamagePer` alone says how
    much per something without saying per what, and the template grants nothing
    unless the source is Burning -- so a row carrying the magnitude and no
    source would read correctly here and do nothing in the game.
    """
    row = row_named(BURNING_WRATH)
    params = shape_params(row)

    if "MoreDamagePer" not in params:
        pytest.fail(
            f"{BURNING_WRATH} has no MoreDamagePer in its Shape Params cell, "
            f"which reads {row.get('Shape Params')!r}. Without it "
            "UCataclysmSelfBuffSkill grants no increase and the buff is only a "
            "duration again. See issue #166."
        )

    if params.get("ScalingSource") != "Burning":
        pytest.fail(
            f"{BURNING_WRATH} states MoreDamagePer without ScalingSource=Burning; "
            f"its cell reads {row.get('Shape Params')!r}. "
            "UCataclysmSelfBuffSkill grants nothing unless the source is Burning, "
            "so the magnitude would sit in the data and never reach the player."
        )

    match = re.search(
        r"(\d+(?:\.\d+)?)% more fire damage for every enemy currently burning",
        row.get("Skill Description", ""),
    )
    if match is None:
        pytest.fail(
            f"{BURNING_WRATH}'s description no longer states a percent of "
            "more fire damage per burning enemy. It reads:\n"
            f"{row.get('Skill Description')}"
        )

    assert float(params["MoreDamagePer"]) == float(match.group(1)), (
        f"{BURNING_WRATH} says {match.group(1)}% in its description and "
        f"{params['MoreDamagePer']} in MoreDamagePer. A player reads "
        "the first and the game does the second."
    )


def test_burning_wrath_counts_over_the_radius_its_description_states() -> None:
    """The 15 metres in the prose is the Radius the count is taken over."""
    row = row_named(BURNING_WRATH)
    params = shape_params(row)

    match = re.search(r"burning within (\d+(?:\.\d+)?) meters", row.get("Skill Description", ""))
    if match is None:
        pytest.fail(
            f"{BURNING_WRATH}'s description no longer states the radius it counts "
            f"burning enemies over. It reads:\n{row.get('Skill Description')}"
        )

    assert float(params.get("Radius", 0)) == float(match.group(1)), (
        f"{BURNING_WRATH} counts burning enemies within {match.group(1)} metres by "
        f"its description and within {params.get('Radius')} by its Radius parameter."
    )


def test_burning_wrath_lasts_as_long_as_its_description_states() -> None:
    """The 10 seconds in the prose is the Duration the buff is held for."""
    row = row_named(BURNING_WRATH)
    params = shape_params(row)

    match = re.search(r"for (\d+(?:\.\d+)?) seconds", row.get("Skill Description", ""))
    if match is None:
        pytest.fail(
            f"{BURNING_WRATH}'s description no longer states how long it lasts. "
            f"It reads:\n{row.get('Skill Description')}"
        )

    assert float(params.get("Duration", 0)) == float(match.group(1)), (
        f"{BURNING_WRATH} lasts {match.group(1)} seconds by its description and "
        f"{params.get('Duration')} by its Duration parameter."
    )


def test_every_skill_with_an_increase_carries_an_element_tag() -> None:
    """An increase is scoped to the skill's damage type, so there must be one.

    `UCataclysmSelfBuffSkill::GrantIncrease` scopes the modifier it adds to the
    skill's own `Element.*` tag. A row carrying IncreasePerBurning and no element
    would grant an increase that applied to every skill the character owns,
    which is not what any description says.
    """
    offenders = [
        row["Skill Name"]
        for row in weapon_skill_rows()
        if "IncreasePerBurning" in shape_params(row)
        and not any(
            tag.strip().startswith("Element.")
            for tag in row.get("Tags", "").split(",")
        )
    ]
    assert not offenders, (
        "These skills grant an increase but carry no Element tag to scope it to, "
        f"so it would apply to everything: {', '.join(offenders)}"
    )

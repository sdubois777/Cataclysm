"""What a summoned minion hits for is stated once, in the design.

WHY THIS MATTERS. `ACataclysmMinion::DamagePercentOfSummoner` used to be 25 and
was labelled in its own comment as a judgement rather than a design figure,
because nothing in `docs/All_Things_Cataclysm.xlsx` or
`docs/Cataclysm_GDD_v2.md` said what a summoned imp hit for. Issue #165 asked
the design to state one. It now does, and this file is what stops the design and
the code drifting apart again.

BOTH NUMBERS ARE READ, NOT WRITTEN HERE. The design figures come out of the
"How a Skill Behaves: the Seven Shapes" section of `docs/Cataclysm_GDD_v2.md`,
and the code figures out of
`game/Source/Cataclysm/AbilitySystem/CataclysmMinion.h`. Neither is copied into
this file, so changing the design in one place is enough and a stale duplicate
cannot survive here.

WHY THE FIGURES ARE NOT SHAPE PARAMS. They do not vary between summoning skills.
Summon Imp, Open the Rift and Cinder Swarm differ in how many minions they make
and how long those last, which the Shape Params column already carries as
`Count`, `MaxActive` and `Duration`. A per-skill minion damage column would be
the same number written three times.
"""

from __future__ import annotations

import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
DESIGN_DOC = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"
MINION_HEADER = (
    REPO_ROOT / "game" / "Source" / "Cataclysm" / "AbilitySystem" / "CataclysmMinion.h"
)

#: The paragraph in the design that states both figures.
DESIGN_SENTENCE_START = "**What a summoned minion hits for"


def read_design_paragraph() -> str:
    """The design's own sentence about minion damage."""
    text = DESIGN_DOC.read_text(encoding="utf-8")
    start = text.find(DESIGN_SENTENCE_START)
    if start == -1:
        pytest.fail(
            f"{DESIGN_DOC.name} no longer contains a paragraph beginning "
            f"{DESIGN_SENTENCE_START!r}. The design must state what a summoned "
            f"minion hits for; see issue #165."
        )
    # To the end of that paragraph. Paragraphs in this document are separated by
    # a blank line.
    end = text.find("\n\n", start)
    return text[start : end if end != -1 else len(text)]


def design_damage_percent() -> float:
    """The percent of the summoner's weapon damage the design states."""
    paragraph = read_design_paragraph()
    match = re.search(
        r"\*\*(\d+(?:\.\d+)?)% of its summoner's weapon damage\*\*", paragraph
    )
    if match is None:
        pytest.fail(
            "The minion paragraph in "
            f"{DESIGN_DOC.name} no longer states a percent of the summoner's "
            f"weapon damage. Paragraph read:\n{paragraph}"
        )
    return float(match.group(1))


def design_attacks_per_second() -> float:
    """How often the design says a minion attacks."""
    paragraph = read_design_paragraph()
    if "**once per second**" in paragraph:
        return 1.0
    match = re.search(r"\*\*(\d+(?:\.\d+)?) times per second\*\*", paragraph)
    if match is None:
        pytest.fail(
            "The minion paragraph in "
            f"{DESIGN_DOC.name} no longer states how often a minion attacks. "
            f"Paragraph read:\n{paragraph}"
        )
    return float(match.group(1))


def code_constant(name: str) -> float:
    """The value of a `static constexpr float` in CataclysmMinion.h."""
    text = MINION_HEADER.read_text(encoding="utf-8")
    match = re.search(
        rf"static\s+constexpr\s+float\s+{re.escape(name)}\s*=\s*(-?\d+(?:\.\d+)?)f?\s*;",
        text,
    )
    if match is None:
        pytest.fail(
            f"{MINION_HEADER.name} no longer declares "
            f"static constexpr float {name}."
        )
    return float(match.group(1))


def test_the_design_states_what_a_minion_hits_for() -> None:
    """The figure exists at all. Issue #165 was that it did not."""
    assert design_damage_percent() > 0.0
    assert design_attacks_per_second() > 0.0


def test_the_code_deals_what_the_design_says() -> None:
    """The constant in the minion matches the design's percent."""
    assert code_constant("DamagePercentOfSummoner") == design_damage_percent()


def test_the_code_attacks_as_often_as_the_design_says() -> None:
    """Seconds between attacks is the reciprocal of attacks per second."""
    interval = code_constant("AttackIntervalSeconds")
    assert interval > 0.0
    assert 1.0 / interval == design_attacks_per_second()


def test_three_minions_deal_less_than_a_basic_attack() -> None:
    """The budget the figure was chosen against still holds.

    THE REASON THE NUMBER IS WHAT IT IS. An automatic basic attack is 128% to
    150% of weapon damage per second depending on weapon speed. Summon Imp caps
    at three active. If three minions out-damaged a basic attack, a summoner
    would be better off never attacking, which is not what a summon slot is for.

    The cap comes from the design's own Summon Imp description rather than being
    written here.
    """
    from openpyxl import load_workbook

    workbook = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
    sheet = load_workbook(workbook, read_only=True, data_only=True)["Weapon Skills"]

    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    name_column = header.index("Skill Name")
    params_column = header.index("Shape Params")

    max_active = None
    for row in rows:
        if row[name_column] == "Summon Imp":
            params = str(row[params_column] or "")
            match = re.search(r"MaxActive\s*=\s*(\d+)", params)
            if match:
                max_active = int(match.group(1))
            break

    if max_active is None:
        pytest.fail(
            "Could not read MaxActive from Summon Imp's Shape Params in "
            f"{workbook.name}."
        )

    per_second = (
        max_active * design_damage_percent() * design_attacks_per_second()
    )
    slowest_basic_attack = 128.0

    assert per_second < slowest_basic_attack, (
        f"{max_active} minions at {design_damage_percent()}% each, attacking "
        f"{design_attacks_per_second()} times a second, come to {per_second}% "
        f"of weapon damage per second. An automatic basic attack is only "
        f"{slowest_basic_attack}% per second at its slowest, so the minions "
        f"would be worth more than attacking."
    )

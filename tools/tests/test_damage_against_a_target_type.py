"""Increased damage against a target's damage type, pinned in every copy.

WHY THIS EXISTS. Issue #213. The project owner approved eight affixes, one per
damage type: increased damage against War, Demonic, Death, Pestilence, Famine,
Celestial, Chaos or Void enemies. The issue left four things to be decided, and
this file is where those decisions are held so a later change cannot quietly
undo one.

THE FOUR DECISIONS.

    How much larger than the generic Increased Damage affix?
        400% against 125%. Not invented: it is the ratio this project already
        pays for narrowing a modifier from all eight damage types to one. The
        resistance families give 20% per type at breadth 1 and 6% at breadth 8,
        so narrowing is worth 20/6, about 3.33 times. The generic damage affix
        is the breadth 8 case because it applies whatever the target is.
        125 x 3.33 is 416.7, rounded to 400.

    Is there a two-type or an all-type version?
        No. The all-type version IS the generic Increased Damage affix, so a
        second one would be the same affix twice. A two-type version is a middle
        rung that cannot be priced before the two ends have been played.

    Prefix or suffix?
        Prefix, and in the same slots as the generic Increased Damage affix, so
        the two compete for one slot on one piece. That competition is the
        choice. Splitting them across positions would let a player take both and
        remove it.

    Does it multiply, or add into the same bracket?
        It adds. The pipeline is (base + flat) x (1 + increases) x more1 x more2
        and a conditional increase joins the increases bracket. Diablo 4 states
        the same rule outright -- a damage bonus with a stated condition is
        additive -- and Last Epoch sums all compatible increased damage
        modifiers. A separate multiplier would be far stronger and much harder
        to balance.

WHY THE RATIO IS THE THING TO GUARD, not the two numbers on their own. 400 and
125 are both tuning numbers and either may move against real play. What must not
move without a decision is that the conditional affix is worth MORE than the
unconditional one, because a conditional affix paying the same is a strictly
worse roll and no player would ever swap gear for it. That is the property the
project owner stated, and it is checked here directly rather than inferred from
the two values.

WHAT IT CHECKS AND WHERE. All three copies of the affix pool, plus the two places
the stats themselves live:

    docs/All_Things_Cataclysm.xlsx, Affixes sheet   authoritative, hand-edited
    game/Data/Affixes.csv                           generated, what the game loads
    sim/cataclysm_sim/affixes.py                    the simulation's model
    sim/cataclysm_sim/character.py                  the character sheet
    game/Source/.../CataclysmCombatAttributeSet.h   what the game runs on
    docs/Cataclysm_GDD_v2.md                        the design document
"""

from __future__ import annotations

import csv
import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
AFFIX_CSV = REPO_ROOT / "game" / "Data" / "Affixes.csv"
GDD = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"
COMBAT_HEADER = (REPO_ROOT / "game" / "Source" / "Cataclysm" / "AbilitySystem"
                 / "CataclysmCombatAttributeSet.h")

#: The name of the unconditional affix these are measured against.
GENERIC = "Increased damage"

SECTION = "### **Damage Against a Target's Type**"


def text(value) -> str:
    return "" if value is None else str(value).strip()


def read_csv(path: pathlib.Path) -> list[dict[str, str]]:
    if not path.is_file():
        pytest.skip(f"{path.name} not present")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_sheet(title: str) -> list[dict[str, object]]:
    import openpyxl

    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")
    rows = list(book[title].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if not raw or raw[0] is None or not text(raw[0]):
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers))
                    if i < len(raw)})
    return out


@pytest.fixture(scope="module")
def model():
    from cataclysm_sim import affixes
    return affixes


@pytest.fixture(scope="module")
def sheet_model():
    from cataclysm_sim import character
    return character


@pytest.fixture(scope="module")
def damage_types(sheet_model) -> tuple[str, ...]:
    """The eight names, read from the model rather than listed here."""
    return sheet_model.DAMAGE_TYPES


@pytest.fixture(scope="module")
def affix_rows() -> list[dict[str, str]]:
    return read_csv(AFFIX_CSV)


@pytest.fixture(scope="module")
def affix_sheet() -> list[dict[str, object]]:
    return read_sheet("Affixes")


def expected_names(damage_types) -> list[str]:
    return [f"Increased damage against {d} enemies" for d in damage_types]


def expected_stats(damage_types) -> list[str]:
    return [f"damage_vs_{d.lower()}" for d in damage_types]


class TestTheEightStatsExist:
    def test_the_character_sheet_has_one_per_damage_type(
            self, sheet_model, damage_types):
        assert list(sheet_model.DAMAGE_VS_STATS) == expected_stats(damage_types)

    def test_they_are_offensive_stats(self, sheet_model):
        offense = set(sheet_model.STAT_GROUPS["Offense"])
        assert set(sheet_model.DAMAGE_VS_STATS) <= offense

    def test_every_class_starts_at_zero(self, sheet_model):
        """No class is born better against a Cataclysm. Which Cataclysms a run
        faces is drawn at run start and the class is chosen before that."""
        for stat in sheet_model.DAMAGE_VS_STATS:
            scaling = sheet_model.DEFAULT_STAT_LINE[stat]
            assert scaling.base == 0.0, stat
            assert scaling.per_level == 0.0, stat

    def test_the_combat_attribute_set_has_all_eight(self, damage_types):
        """The game runs on the attribute set, not on the Python model."""
        source = COMBAT_HEADER.read_text(encoding="utf-8")
        for damage_type in damage_types:
            declared = f"FGameplayAttributeData DamageVs{damage_type};"
            assert declared in source, (
                f"{COMBAT_HEADER.name} does not declare {declared}")


class TestTheEightAffixesExist:
    def test_the_model_has_one_per_damage_type(self, model, damage_types):
        assert [a.name for a in model.DAMAGE_VS_AFFIXES] == \
            expected_names(damage_types)
        assert [a.stat for a in model.DAMAGE_VS_AFFIXES] == \
            expected_stats(damage_types)

    def test_the_workbook_has_all_eight(self, affix_sheet, damage_types):
        names = {text(row["Affix Name"]) for row in affix_sheet}
        missing = [n for n in expected_names(damage_types) if n not in names]
        assert not missing, f"the Affixes sheet is missing {missing}"

    def test_the_table_the_game_loads_has_all_eight(
            self, affix_rows, damage_types):
        names = {text(row["AffixName"]) for row in affix_rows}
        missing = [n for n in expected_names(damage_types) if n not in names]
        assert not missing, f"{AFFIX_CSV.name} is missing {missing}"


class TestTheDecisionsThatWereMade:
    """Each of the four questions the issue left open, held where a later change
    has to notice it."""

    def test_they_are_prefixes(self, affix_rows, damage_types):
        wanted = set(expected_names(damage_types))
        for row in affix_rows:
            if text(row["AffixName"]) in wanted:
                assert text(row["Position"]) == "prefix", (
                    f"{row['AffixName']} is a {row['Position']}. These are "
                    "prefixes so they compete with the generic Increased "
                    "damage affix for one slot on one piece.")

    def test_they_add_into_the_increases_bracket(self, affix_rows, damage_types):
        """They add rather than multiplying, which is the decision above.

        THE COLUMN READS `flat` AND THAT IS WHAT ADDING LOOKS LIKE HERE. Two
        senses of the word point opposite ways and this is where they meet.

        In the design, "increased damage against Demonic enemies" is a
        percentage rather than a flat number of damage. That is the decision and
        it has not changed.

        In the data model, `Value Kind` says how a modifier combines with the
        STAT IT NAMES, and the stat named here is `damage_vs_<type>` -- which IS
        the bucket of conditional increases. Contributing percentage points to a
        bucket is a `flat` modifier. `increased` would multiply whatever the
        bucket already holds, and nothing else puts anything in it, so until
        2026-08-24 all eight multiplied zero and were worth nothing at all.

        THIS TEST ASSERTED `increased` UNTIL THEN, so it was holding the one
        state in which the decision it guards could not be true. The same
        correction was made to cooldown reduction the same day, for the same
        reason; `docs/DECISIONS.md` carries both.
        """
        wanted = set(expected_names(damage_types))
        for row in affix_rows:
            if text(row["AffixName"]) in wanted:
                assert text(row["ValueKind"]) == "flat", (
                    f"{row['AffixName']} is `{row['ValueKind']}`. The stat it "
                    "names holds the conditional increases themselves, so a "
                    "source adds into it. See this test's docstring.")

    def test_they_roll_where_the_generic_damage_affix_rolls(
            self, model, damage_types):
        """Same slots, or the choice between them would not exist on a piece."""
        generic = next(a for a in model.AFFIX_POOL if a.name == GENERIC)
        for affix in model.DAMAGE_VS_AFFIXES:
            assert affix.allowed_slots == generic.allowed_slots, affix.name
            assert affix.position == generic.position, affix.name

    def test_all_eight_are_worth_the_same(self, model):
        values = {a.top_value for a in model.DAMAGE_VS_AFFIXES}
        assert len(values) == 1, (
            f"the eight have different top values: {sorted(values)}. No damage "
            "type is harder to fight than another, so none of them is worth "
            "more.")

    def test_the_conditional_affix_is_worth_more_than_the_unconditional_one(
            self, model):
        """The property the project owner stated, checked directly.

        A conditional affix paying the same as an unconditional one is a
        strictly worse roll, and nobody would ever swap gear for it.
        """
        generic = next(a for a in model.AFFIX_POOL if a.name == GENERIC)
        assert model.DAMAGE_VS_TOP_VALUE > generic.top_value

    def test_the_ratio_is_what_narrowing_a_resistance_is_worth(self, model):
        """400 against 125 is 3.2. Narrowing a resistance from all eight types
        to one is worth 20/6, about 3.33. The damage ratio is derived from the
        resistance one and rounded, so it must stay near it."""
        generic = next(a for a in model.AFFIX_POOL if a.name == GENERIC)
        damage_ratio = model.DAMAGE_VS_TOP_VALUE / generic.top_value
        resistance_ratio = (model.SINGLE_RESISTANCE.top_value
                            / model.ALL_RESISTANCE.top_value)
        assert damage_ratio == pytest.approx(3.2)
        assert abs(damage_ratio - resistance_ratio) < 0.25, (
            f"the damage ratio is {damage_ratio:.2f} and the resistance ratio "
            f"is {resistance_ratio:.2f}. One was derived from the other; if "
            "either moves on purpose, move this bound and say why.")

    def test_there_is_no_two_type_or_all_type_version(self, affix_rows):
        """The all-type version is the generic affix. A second one would be the
        same affix twice, and strictly better than every single-type one."""
        for row in affix_rows:
            name = text(row["AffixName"])
            if not name.startswith("Increased damage against"):
                continue
            assert not text(row["Breadth"]) or text(row["Breadth"]) == "0", (
                f"{name} has a breadth of {row['Breadth']}. These cover exactly "
                "one damage type each; a broader version has not been designed.")


class TestTheDesignDocumentSaysTheSame:
    @pytest.fixture(scope="module")
    def section(self) -> str:
        if not GDD.is_file():
            pytest.skip("the design document is not present")
        body = GDD.read_text(encoding="utf-8")
        start = body.find(SECTION)
        assert start != -1, f"{GDD.name} has no section headed {SECTION!r}"
        after = body[start + len(SECTION):]
        ends = [m.start() for m in re.finditer(r"^#{1,6} ", after, re.MULTILINE)]
        return after[:ends[0]] if ends else after

    def test_it_states_the_value(self, section, model):
        match = re.search(
            r"Increased damage against one type\s*\|\s*(\d+(?:\.\d+)?)%",
            section)
        assert match, "the section's table has no row for the type-specific affix"
        assert float(match.group(1)) == model.DAMAGE_VS_TOP_VALUE

    def test_it_states_the_generic_value_beside_it(self, section, model):
        generic = next(a for a in model.AFFIX_POOL if a.name == GENERIC)
        match = re.search(r"Increased damage\s*\|\s*(\d+(?:\.\d+)?)%", section)
        assert match, "the section's table has no row for the generic affix"
        assert float(match.group(1)) == generic.top_value

    def test_it_states_the_crossover(self, section, model):
        """The number of active Cataclysms at which the two are equal. It is
        what the whole design rests on, so the document has to carry it and it
        has to follow from the two values rather than be typed in."""
        match = re.search(r"equal at C = (\d+(?:\.\d+)?)", section)
        assert match, "the section does not state where the two are equal"
        generic = next(a for a in model.AFFIX_POOL if a.name == GENERIC)
        stated = float(match.group(1))
        assert stated == pytest.approx(
            model.DAMAGE_VS_TOP_VALUE / generic.top_value, abs=0.05)

    def test_it_names_all_eight_damage_types(self, section, damage_types):
        for damage_type in damage_types:
            assert damage_type in section, (
                f"the section does not name {damage_type}")

    def test_it_says_they_add_rather_than_multiply(self, section):
        assert "additive" in section or "adds" in section or "add into" in section
        assert "more1" in section, (
            "the section should show the pipeline, so a reader can see which "
            "bracket a conditional increase joins")

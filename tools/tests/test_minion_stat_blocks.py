"""The minion type table: what is in it, and whether it plays the way it was set.

`tools/tests/test_minion_damage.py` checks that `docs/Cataclysm_GDD_v2.md` states
the minion rules and that the data agrees with the prose. This file checks the
stat blocks themselves -- the numbers in `game/Data/MinionTypes.csv`, the table
generated from the "Minion Types" sheet of `docs/All_Things_Cataclysm.xlsx`.

Issue #336.
"""

from __future__ import annotations

import csv
import pathlib
import sys

import openpyxl
import pytest

ROOT = pathlib.Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "tools"))
sys.path.insert(0, str(ROOT / "sim"))

import generate_datatables as gen  # noqa: E402
from cataclysm_sim import player_damage as pd  # noqa: E402
from cataclysm_sim import player_power as pp  # noqa: E402

#: Summon Imp and Open the Rift are both Staff skills, so the summoner holds a
#: Staff. Read from the skill table below rather than assumed.
SUMMONER_WEAPON = "Staff"

#: From the Shape Params of the skills that produce them.
IMP_MAX_ACTIVE = 3
MOTE_MAX_ACTIVE = 4


def minion_rows() -> dict[str, dict]:
    path = ROOT / "game" / "Data" / "MinionTypes.csv"
    with path.open(encoding="utf-8") as handle:
        return {r["Name"]: r for r in csv.DictReader(handle)}


def skill_rows() -> list[dict]:
    path = ROOT / "game" / "Data" / "WeaponSkills.csv"
    with path.open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def damage_at(row: dict, level: int) -> float:
    return float(row["BaseDamage"]) + float(row["DamagePerLevel"]) * level


def health_at(row: dict, level: int) -> float:
    return float(row["BaseHealth"]) + float(row["HealthPerLevel"]) * level


def squad_per_second(row: dict, level: int, count: int) -> float:
    return count * damage_at(row, level) / float(row["AttackIntervalSeconds"])


# --------------------------------------------------------------------------
# What is in the table
# --------------------------------------------------------------------------

def test_the_table_holds_every_minion_type_a_skill_produces():
    """Two summoned creatures and three deployed machines.

    The Demonic pair landed with issue #336, for the vertical slice. The three
    War deployables followed with issue #338, because their numbers lived only
    in prose and a skill deploying two ballistae and three spike traps cannot
    say so without them.
    """
    assert set(minion_rows()) == {"Imp", "Mote", "BoltTurret", "Ballista",
                                  "SpikeTrap"}


def test_every_demonic_minion_skill_produces_a_type_the_table_defines():
    """The whole reason this table exists: a skill names a creature and the
    creature's numbers live in one place, so two skills making the same creature
    cannot disagree about it."""
    # SUBJUGATE PRODUCES NO NAMED TYPE, because a thrall is the enemy it was
    # taken from and carries that enemy's own stat block. Open the Rift and
    # Cinder Swarm were retired on 2026-09-01 by the Demonic verb rewrite.
    produced = {"Summon Imp": "Imp"}
    demonic_minion_skills = {
        r["SkillName"] for r in skill_rows()
        if "Type.Minion" in r["Tags"] and r["DamageType"] == "Demonic"
        # A possessing skill produces a minion and names no type, because the
        # thrall keeps the stat block of the enemy it was taken from. There is
        # nothing in the Minion Types sheet for it to point at.
        and "Possess=1" not in (r["ShapeParams"] or "")}

    assert demonic_minion_skills == set(produced), (
        "the Demonic minion skills in game/Data/WeaponSkills.csv changed, so "
        "the mapping this test uses is stale")
    for skill, minion_type in produced.items():
        assert minion_type in minion_rows(), (
            f"{skill} produces a {minion_type} and no such row exists")


def test_two_skills_making_the_same_creature_share_one_stat_block():
    """Summon Imp is the one Staff skill that makes a lesser imp. There is one Imp row,
    so the numbers cannot exist twice and drift apart."""
    staff_minion_skills = [
        r["SkillName"] for r in skill_rows()
        if "Type.Minion" in r["Tags"] and r["WeaponType"] == "Staff"]
    assert sorted(staff_minion_skills) == ["Subjugate", "Summon Imp"]
    assert len([n for n in minion_rows() if n == "Imp"]) == 1


def scaling_rows() -> list[dict]:
    path = ROOT / "game" / "Data" / "MinionScaling.csv"
    with path.open(encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def tags_of(row: dict) -> set[str]:
    return {t.strip() for t in row["Tags"].split(",") if t.strip()}


def increase_for(minion: dict, stat: str, points: dict[str, int]) -> float:
    """What the summoner's attribute points add to one of a minion's stats.

    Only rows whose tag the minion carries count, which is the whole point of
    scoping by tag rather than by a shared stat name.
    """
    total = 0.0
    for row in scaling_rows():
        if row["Stat"] != stat or row["RequiresTag"] not in tags_of(minion):
            continue
        total += points.get(row["Attribute"], 0) * float(row["PercentPerPoint"])
    return total / 100.0


def test_a_creature_is_raised_by_spirit_and_a_machine_by_agility():
    """Settled in issue #335, and now expressed as a tag rather than a column."""
    by_tag = {(r["RequiresTag"], r["Stat"]): r for r in scaling_rows()}
    assert by_tag[("Minion.Creature", "damage")]["Attribute"] == "spirit"
    assert by_tag[("Minion.Machine", "damage")]["Attribute"] == "agility"
    for row in scaling_rows():
        assert float(row["PercentPerPoint"]) == 1.0, row


def test_every_minion_carries_its_familys_tag_and_only_its_familys_tag():
    """Carrying both would let both families' scaling reach one minion."""
    for name, row in minion_rows().items():
        tags = tags_of(row)
        mine = f"Minion.{row['Family']}"
        others = {f"Minion.{f}" for f in ("Creature", "Machine")} - {mine}
        assert "Type.Minion" in tags, name
        assert mine in tags, name
        assert not (others & tags), name


def test_a_deployed_machine_does_not_move_and_a_summoned_creature_does():
    """The design says move speed is how the two are told apart without a rule
    of its own: zero for a deployable, non-zero for a summon."""
    for name, row in minion_rows().items():
        speed = float(row["MoveSpeed"])
        if row["Family"] == "Machine":
            assert speed == 0.0, f"{name} is a machine and moves at {speed}"
            assert "Type.Deployable" in tags_of(row), name
        else:
            assert speed > 0.0, f"{name} is a creature and does not move"
            assert "Type.Summon" in tags_of(row), name


def test_a_turret_draws_far_less_attention_than_a_creature():
    """The design's own example of what the threat column is for: a turret near
    zero, an imp at 100."""
    imp = float(minion_rows()["Imp"]["ThreatPercent"])
    for name, row in minion_rows().items():
        if row["Family"] == "Machine":
            assert float(row["ThreatPercent"]) < imp / 4, name


def test_agility_does_nothing_for_a_creature():
    """THE DOUBLE COUNT THIS SHAPE EXISTS TO PREVENT.

    A single shared "increased minion damage" stat would let a summoner's
    Agility raise an imp, because the character sheet sums every attribute that
    names a stat. Scoping by tag means an Agility point reaches a machine and
    nothing else.
    """
    imp = minion_rows()["Imp"]
    assert increase_for(imp, "damage", {"spirit": 100}) == pytest.approx(1.0)
    assert increase_for(imp, "damage", {"agility": 100}) == pytest.approx(0.0)
    assert increase_for(imp, "damage", {"spirit": 50, "agility": 50}) == \
        pytest.approx(0.5)


def test_one_hundred_points_of_the_right_attribute_doubles_minion_damage():
    """The design document states "Each grants 1.0% increased minion damage per
    point, so 100 points doubles it"."""
    imp = minion_rows()["Imp"]
    plain = damage_at(imp, 100)
    raised = plain * (1.0 + increase_for(imp, "damage", {"spirit": 100}))
    assert raised == pytest.approx(plain * 2.0)


def test_only_damage_is_decided_and_health_is_expressible():
    """Health scaling has no figure yet. The table can carry it; nothing does."""
    assert {r["Stat"] for r in scaling_rows()} == {"damage"}
    assert "health" in gen.MINION_SCALABLE_STATS


def test_the_ballista_is_the_one_that_picks_the_furthest_enemy():
    """Which is the whole reason target mode is a column rather than one global
    rule. Its skill description says it "fires massive bolts at the furthest
    enemy within 15 meters"."""
    rows = minion_rows()
    furthest = [n for n, r in rows.items() if r["TargetMode"] == "Furthest"]
    assert furthest == ["Ballista"]
    for name, row in rows.items():
        if name != "Ballista":
            assert row["TargetMode"] == "Nearest", name


# --------------------------------------------------------------------------
# How it plays
# --------------------------------------------------------------------------

def test_three_imps_out_damage_the_summoners_own_basic_attack_at_every_tier():
    """DECIDED BY THE PROJECT OWNER on 2026-08-15, and it reverses the floor
    issue #336 originally proposed, which was that an uninvested summoner should
    get LESS from three imps than from their basic attack.

    Three imps are permanent in normal play: Summon Imp holds 3 for 20 seconds
    against a 5 second cooldown.
    """
    imp = minion_rows()["Imp"]
    for tier in range(1, 9):
        level = pp.reference_character(tier).level
        squad = squad_per_second(imp, level, IMP_MAX_ACTIVE)
        player = pd.damage_per_second(tier, SUMMONER_WEAPON)
        assert squad > player, (
            f"at difficulty tier {tier} three imps deal {squad:,.0f} a second "
            f"and the summoner's own basic attack deals {player:,.0f}")


def test_a_mote_is_the_weaker_body():
    """The project owner's words on 2026-08-15: a mote is a weaker body in
    larger numbers. So one mote hits for less than one imp, and there are more
    of them."""
    imp, mote = minion_rows()["Imp"], minion_rows()["Mote"]
    for level in (1, 25, 50, 100):
        assert damage_at(mote, level) < damage_at(imp, level), level
        assert health_at(mote, level) < health_at(imp, level), level
    assert MOTE_MAX_ACTIVE > IMP_MAX_ACTIVE


def test_a_full_swarm_of_motes_deals_less_than_a_full_squad_of_imps():
    """The project owner left the size of the difference open and it is set at
    about three quarters. The swarm is paid back by the burst each mote leaves
    when it expires, which an imp only gives when a fourth is summoned over the
    cap."""
    imp, mote = minion_rows()["Imp"], minion_rows()["Mote"]
    for tier in range(1, 9):
        level = pp.reference_character(tier).level
        motes = squad_per_second(mote, level, MOTE_MAX_ACTIVE)
        imps = squad_per_second(imp, level, IMP_MAX_ACTIVE)
        assert 0.70 <= motes / imps <= 0.80, (
            f"at difficulty tier {tier} four motes deal {motes / imps:.2f}x "
            f"what three imps deal")


def test_a_minion_keeps_pace_with_its_summoner_rather_than_running_away():
    """RECORDS A CONSEQUENCE, not a defect. REWRITTEN ON 2026-09-03.

    THIS TEST USED TO ASSERT THE OPPOSITE. It required the imp squad to be worth
    MORE at difficulty tier 1 than at tier 8, on the reasoning that a minion's
    damage rises with the summoner's LEVEL only while the summoner's own rises
    with level AND affix tier AND gear upgrade level, so the summoner climbs
    faster and the squad falls behind.

    That was true, and it was true because the gear ladder was steep. Issue
    #1179 eased both affix ladders -- the tier ladder from a span of 7.0 to 3.0
    and the gear ladder from 3.52 to 2.5 -- so the summoner's own damage now
    climbs more gently, and the squad no longer falls away from it. The fall-off
    was a consequence of the old numbers rather than something anybody chose.

    WHAT IS WORTH GUARDING NOW is that the two stay near each other, in both
    directions. A squad that runs away from its summoner makes an uninvested
    summoner strictly better off ignoring their own weapon, and a squad that
    falls far behind makes summoning pointless without the minion affixes of
    issue #337. The band below is wide because nothing here has been played yet;
    it is here to catch a change that breaks the relationship, not to pin a
    tuned number.
    """
    imp = minion_rows()["Imp"]
    ratios = {}
    for tier in range(1, 9):
        level = pp.reference_character(tier).level
        ratios[tier] = (squad_per_second(imp, level, IMP_MAX_ACTIVE)
                        / pd.damage_per_second(tier, SUMMONER_WEAPON))

    for tier, ratio in ratios.items():
        assert 1.0 < ratio < 2.0, (
            f"at difficulty tier {tier} three imps are worth {ratio:.2f}x the "
            f"summoner's own basic attack, outside the band this test holds. "
            f"Below 1.0 there is no reason to summon; above 2.0 there is no "
            f"reason to swing. Ratios across all eight tiers: "
            f"{ {t: round(r, 2) for t, r in ratios.items()} }")

    spread = max(ratios.values()) / min(ratios.values())
    assert spread < 2.0, (
        f"the squad's worth against the summoner varies {spread:.2f}x across "
        f"the eight difficulty tiers, which is too much for one relationship. "
        f"Ratios: { {t: round(r, 2) for t, r in ratios.items()} }")


def test_a_minion_is_fragile_enough_to_be_worth_killing():
    """A minion inherits no armour and no resistance, so it takes hits whole.
    Both should die to a handful of Common hits rather than being ignorable."""
    from cataclysm_sim import enemy_stats as es

    for name, row in minion_rows().items():
        level = pp.reference_character(8).level
        hit = es.stats_on_floor("Common", 8, "Cataclysm").average_damage_per_hit
        survives = health_at(row, level) / hit
        assert survives < 5.0, (
            f"a {name} survives {survives:.1f} Common hits at tier 8")


# --------------------------------------------------------------------------
# The generator's guards
# --------------------------------------------------------------------------

HEADERS = ["Minion Type", "Family", "Base Health", "Health Per Level",
           "Base Damage", "Damage Per Level", "Attack Interval Seconds",
           "Move Speed", "Threat Percent", "Reach Cm", "Notice Radius Cm",
           "Target Mode", "Tags"]

GOOD = ["Imp", "Creature", 200, 90, 10, 10.0, 1.0, 4.4, 100, 200, 1500,
        "Nearest", "Type.Minion, Type.Summon, Minion.Creature, Minion.Melee"]

SCALING_HEADERS = ["Attribute", "Requires Tag", "Stat", "Percent Per Point"]
GOOD_SCALING = ["spirit", "Minion.Creature", "damage", 1.0]


def book_with(tmp_path, row, sheet_name="Minion Types", headers=None):
    book = openpyxl.Workbook()
    book.remove(book.active)
    sheet = book.create_sheet(sheet_name)
    sheet.append(headers or HEADERS)
    sheet.append(row)
    path = tmp_path / "w.xlsx"
    book.save(path)
    return openpyxl.load_workbook(path)


def scaling_book(tmp_path, row):
    return book_with(tmp_path, row, "Minion Scaling", SCALING_HEADERS)


def test_the_good_row_is_accepted(tmp_path):
    """A guard that rejects everything proves nothing, so check the control."""
    rows = gen.minion_types(book_with(tmp_path, list(GOOD)))
    assert rows[0]["Name"] == "Imp"
    assert "Minion.Creature" in rows[0]["Tags"]


def test_a_creature_without_its_familys_tag_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Tags")] = "Type.Minion, Type.Summon, Minion.Melee"
    with pytest.raises(gen.DataError, match="does not carry Minion.Creature"):
        gen.minion_types(book_with(tmp_path, row))


def test_a_minion_carrying_both_family_tags_is_refused(tmp_path):
    """Both families' scaling would reach it, which is the double count again."""
    row = list(GOOD)
    row[HEADERS.index("Tags")] = ("Type.Minion, Minion.Creature, Minion.Machine")
    with pytest.raises(gen.DataError, match="also carries"):
        gen.minion_types(book_with(tmp_path, row))


def test_a_minion_with_no_tags_at_all_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Tags")] = ""
    with pytest.raises(gen.DataError, match="carries no tags"):
        gen.minion_types(book_with(tmp_path, row))


def test_an_unknown_family_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Family")] = "Elemental"
    with pytest.raises(gen.DataError, match="family"):
        gen.minion_types(book_with(tmp_path, row))


def test_an_unknown_target_mode_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Target Mode")] = "Weakest"
    with pytest.raises(gen.DataError, match="targets"):
        gen.minion_types(book_with(tmp_path, row))


def test_an_attack_interval_of_zero_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Attack Interval Seconds")] = 0
    with pytest.raises(gen.DataError, match="rather than at a rate"):
        gen.minion_types(book_with(tmp_path, row))


def test_no_health_at_all_is_refused(tmp_path):
    row = list(GOOD)
    row[HEADERS.index("Base Health")] = 0
    with pytest.raises(gen.DataError, match="base health"):
        gen.minion_types(book_with(tmp_path, row))


def test_a_good_scaling_row_is_accepted(tmp_path):
    rows = gen.minion_scaling(scaling_book(tmp_path, list(GOOD_SCALING)))
    assert rows[0]["Attribute"] == "spirit"
    assert rows[0]["RequiresTag"] == "Minion.Creature"


def test_a_scaling_row_with_no_tag_is_refused(tmp_path):
    """Without a tag it would reach every minion of every family."""
    row = list(GOOD_SCALING)
    row[SCALING_HEADERS.index("Requires Tag")] = ""
    with pytest.raises(gen.DataError, match="requires no tag"):
        gen.minion_scaling(scaling_book(tmp_path, row))


def test_a_scaling_row_naming_a_stat_a_minion_does_not_have_is_refused(tmp_path):
    row = list(GOOD_SCALING)
    row[SCALING_HEADERS.index("Stat")] = "crit_chance"
    with pytest.raises(gen.DataError, match="expected one of"):
        gen.minion_scaling(scaling_book(tmp_path, row))


def test_a_scaling_row_worth_nothing_per_point_is_refused(tmp_path):
    row = list(GOOD_SCALING)
    row[SCALING_HEADERS.index("Percent Per Point")] = 0
    with pytest.raises(gen.DataError, match="the point is wasted"):
        gen.minion_scaling(scaling_book(tmp_path, row))

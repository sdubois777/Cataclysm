"""The Demonic skills designed for the vertical slice must stay coherent.

WHY THIS EXISTS. Three separate things can drift apart here, and every one of
them has already drifted somewhere else in this project.

  1. A skill row can name a tag that does not exist. The generator already
     rejects that, but only for tags; nothing checked that a Demonic row uses
     the same weapon tag the War row for that weapon uses, so a Greataxe skill
     tagged `Item.Weapon.Axe` would generate cleanly and grant the wrong kit.

  2. A skill description can contradict a rule the design states. The design
     says an enemy carries at most one stack of any effect the player applies.
     Fifteen War skill descriptions were written before that rule and still say
     "applies 2 bleed stacks". The Demonic rows must not repeat it.

  3. The design document's own example table can drift from the workbook the
     tables are generated from. The War Skill Examples table has already done
     this: it gives the Ballista a 2.5 second interval where the workbook says
     2, and has Fortress reflecting 100% where the workbook says 500%.

WHICH IS AUTHORITATIVE. The workbook, for the same reason as everywhere else in
this project: it is what the project owner edits and what the generated data
tables come from.
"""

from __future__ import annotations

import pathlib
import re

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
DESIGN_DOC = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"

#: The three weapon types the vertical slice designs, one per Demonic class:
#: Greataxe for the Ravager, Fist for the Masochist, Staff for the Ritualist.
SLICE_WEAPONS = ("Greataxe", "Fist", "Staff")

#: Every slot a weapon supplies. The basic attack is automatic and has no row.
WEAPON_SLOTS = ("Heavy", "Special", "Support", "Aura", "Ultimate", "Movement")


def text(value) -> str:
    return "" if value is None else str(value).strip()


def tag_set(value) -> set[str]:
    return {t.strip() for t in text(value).split(",") if t.strip()}


@pytest.fixture(scope="module")
def skills() -> list[dict[str, str]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    out = []
    for raw in list(book["Weapon Skills"].iter_rows(values_only=True))[1:]:
        if not raw or not text(raw[0]):
            continue
        out.append({"weapon": text(raw[0]), "damage": text(raw[1]),
                    "slot": text(raw[2]), "name": text(raw[3]),
                    "description": text(raw[4]), "tags": text(raw[5]),
                    "shape": text(raw[6]) if len(raw) > 6 else "",
                    "params": text(raw[7]) if len(raw) > 7 else ""})
    return out


@pytest.fixture(scope="module")
def known_tags() -> set[str]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    return {text(r[0]) for r in book["Tags"].iter_rows(values_only=True)
            if r and text(r[0])} - {"Tag Name"}


@pytest.fixture(scope="module")
def demonic(skills) -> list[dict[str, str]]:
    return [row for row in skills if row["damage"] == "Demonic"]


@pytest.fixture(scope="module")
def designed(demonic) -> list[dict[str, str]]:
    return [row for row in demonic if row["name"]]


def test_the_slice_designs_every_slot_of_its_three_weapons(demonic):
    """Sixteen rows: three weapons across five slots, plus the shared Aura."""
    missing = []
    for weapon in SLICE_WEAPONS:
        for slot in WEAPON_SLOTS:
            if slot == "Aura":
                continue  # weapon-agnostic, carried on the "All" row
            found = [r for r in demonic
                     if r["weapon"] == weapon and r["slot"] == slot]
            if not found:
                missing.append(f"{weapon}/{slot}: no row in the sheet at all")
            elif not found[0]["name"]:
                missing.append(f"{weapon}/{slot}: row exists but is undesigned")

    aura = [r for r in demonic if r["slot"] == "Aura"]
    if not aura:
        missing.append("All/Aura: no Demonic aura row")
    elif not aura[0]["name"]:
        missing.append("All/Aura: row exists but is undesigned")

    assert not missing, "\n".join(missing)


def test_every_designed_row_carries_a_name_a_description_and_tags(designed):
    assert designed, "no Demonic skill is designed at all"
    incomplete = [f"{r['weapon']}/{r['slot']} ({r['name']})"
                  for r in designed
                  if not r["description"] or not r["tags"]]
    assert not incomplete, (
        "designed Demonic rows missing a description or tags: "
        + ", ".join(incomplete))


def test_every_tag_used_is_a_tag_that_exists(designed, known_tags):
    problems = []
    for row in designed:
        for tag in sorted(tag_set(row["tags"]) - known_tags):
            problems.append(f"{row['weapon']}/{row['slot']}: undefined {tag}")
    assert not problems, "\n".join(problems)


def test_every_designed_row_is_tagged_demonic(designed):
    wrong = [f"{r['weapon']}/{r['slot']}" for r in designed
             if "Element.Demonic" not in tag_set(r["tags"])]
    assert not wrong, ("Demonic rows without the Element.Demonic tag: "
                       + ", ".join(wrong))


def test_the_weapon_tag_matches_the_one_war_uses_for_that_weapon(
        skills, designed):
    """A Greataxe skill tagged Item.Weapon.Axe would generate cleanly.

    War is the one complete damage type, so it is the reference for which tag
    names which weapon. This catches a Demonic row that grants the wrong kit.
    """
    war_tag: dict[str, set[str]] = {}
    for row in skills:
        if row["damage"] != "War" or not row["tags"]:
            continue
        weapon_tags = {t for t in tag_set(row["tags"])
                       if t.startswith("Item.Weapon.")}
        war_tag.setdefault(row["weapon"], set()).update(weapon_tags)

    problems = []
    for row in designed:
        if row["weapon"] == "All":
            continue  # the aura is weapon-agnostic and carries no weapon tag
        mine = {t for t in tag_set(row["tags"])
                if t.startswith("Item.Weapon.")}
        if len(mine) != 1:
            problems.append(
                f"{row['weapon']}/{row['slot']}: expected exactly one "
                f"Item.Weapon tag, found {sorted(mine) or 'none'}")
            continue
        expected = war_tag.get(row["weapon"])
        if expected and mine != expected:
            problems.append(
                f"{row['weapon']}/{row['slot']}: tagged {sorted(mine)[0]} but "
                f"War tags the {row['weapon']} as {sorted(expected)[0]}")
    assert not problems, "\n".join(problems)


def test_the_aura_row_carries_no_weapon_tag(designed):
    aura = [r for r in designed if r["slot"] == "Aura"]
    assert aura, "no designed Demonic aura row"
    for row in aura:
        weapon_tags = {t for t in tag_set(row["tags"])
                       if t.startswith("Item.Weapon.")}
        assert not weapon_tags, (
            f"the Demonic aura is weapon-agnostic but is restricted to "
            f"{sorted(weapon_tags)}")


def test_no_description_counts_stacks(skills):
    """The design says an enemy carries at most one stack of any effect.

    EVERY DAMAGE TYPE, not only Demonic. This checked the Demonic rows alone
    until issue #153, because fifteen War descriptions predated the rule and
    still counted bleed stacks: failing on them here would have made this test
    unfixable by its own change. Those fifteen were rewritten, so the check now
    covers the whole sheet and a new row in any damage type is held to the rule.

    MATCHED ON THE WORD, NOT ON A LIST OF EFFECTS. "2 bleed stacks", "stacking
    up to 4 times" and "for each bleed stack within 15 meters" are three
    different shapes of the same mistake, and the only thing they share is the
    word. Anything a description genuinely needs to say can be said without it:
    a stronger application is magnitude, and counting other enemies is counting
    bleeding enemies.
    """
    offenders = []
    for row in skills:
        if not row["name"]:
            continue
        found = re.findall(r"[^.]*\bstack(?:s|ing|ed)?\b[^.]*\.",
                           row["description"], re.IGNORECASE)
        for sentence in found:
            offenders.append(
                f"[{row['damage']}] {row['weapon']}/{row['slot']} "
                f"({row['name']}): {sentence.strip()}")
    assert not offenders, (
        "Skill descriptions counting stacks, which the single-stack rule "
        "forbids:\n" + "\n".join(offenders))


def test_every_designed_row_names_a_shape_and_its_numbers(designed):
    """A skill with no shape has a name and a description and does nothing.

    That was true of all 61 War rows and all 16 Demonic ones until the shape
    columns arrived: the weapon slots component granted a placeholder, the key
    reached it, and it ended immediately. A Demonic row without a shape now
    means somebody wrote a skill and stopped short of making it real.
    """
    # Not vacuous: there are 51 Demonic rows in the sheet after issue #23 cut
    # the matrix from 558 to 398, and all of them are designed. A fixture that
    # silently returned nothing would pass every assertion below it.
    assert len(designed) == 51, (
        f"expected 51 designed Demonic rows, found {len(designed)}")

    missing = [f"{r['weapon']}/{r['slot']} ({r['name']})"
               for r in designed if not r["shape"]]
    assert not missing, ("designed Demonic rows with no shape, so no behaviour: "
                         + ", ".join(missing))

    # A shape with no parameters is a skill with a radius of zero, which hits
    # nothing. Every shape in use reads at least one number.
    bare = [f"{r['weapon']}/{r['slot']} ({r['name']})"
            for r in designed if r["shape"] and not r["params"]]
    assert not bare, ("designed Demonic rows with a shape but no numbers: "
                      + ", ".join(bare))


def test_every_designed_row_sets_its_target_alight(designed):
    """The design says every Demonic skill applies burn, as every War skill bleeds.

    Checked on the DATA rather than the prose, because the prose is what a
    reader sees and the Burn parameter is what the game acts on. A description
    saying "setting each one alight" with no Burn=1 beside it would read
    correctly and do nothing.

    The two Support skills that buff the caster rather than touching an enemy
    are the exception: there is nothing for them to set alight at the moment
    they are used.
    """
    missing = []
    for row in designed:
        if row["shape"] in ("SelfBuff", "Debuff"):
            continue
        if "Burn=1" not in row["params"]:
            missing.append(f"{row['weapon']}/{row['slot']} ({row['name']})")
    assert not missing, ("Demonic skills that do not set their target alight: "
                         + ", ".join(missing))


def test_a_closing_hit_is_only_written_on_a_skill_that_repeats(designed):
    """FinalHitPercent is landed by the timer that ends a repeating swing.

    A skill that does not repeat never reaches that code, so the number would
    sit in the data and nothing would read it -- a skill quietly weaker than its
    own description. See UCataclysmStrikeSkill::Finish.
    """
    stranded = []
    for row in designed:
        params = row["params"]
        if "FinalHitPercent" not in params:
            continue
        repeats = "Duration=" in params and "Interval=" in params
        if not repeats:
            stranded.append(f"{row['weapon']}/{row['slot']} ({row['name']})")
    assert not stranded, (
        "skills stating a closing hit that nothing will land, because they do "
        "not repeat: " + ", ".join(stranded))


def test_burn_is_an_effect_the_player_can_apply():
    """Every Demonic skill sets enemies alight, so burn has to be player-applied.

    Before this change the DoTs sheet described burn only as something the
    Infernal Brand and Hellfire Aura enemy modifiers do to the player, which
    made every one of those skills a percentage of nothing.
    """
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    rows = [text(r[0]) for r in book["DoTs"].iter_rows(values_only=True)
            if r and text(r[0])]
    burn = [r for r in rows if r.startswith("Burn:")]
    assert len(burn) == 1, f"expected one Burn row in the DoTs sheet, got {burn}"
    assert "player-applied" in burn[0].lower(), (
        "burn must be an effect the player can apply, because every Demonic "
        f"skill applies it. The sheet says: {burn[0]}")


def test_the_design_document_examples_match_the_workbook(designed):
    """The War examples table already drifted from the sheet. Stop the Demonic one.

    Checks the skill names only. The descriptions in the document are
    deliberately abbreviated, so comparing those would fail on formatting
    rather than on drift.
    """
    document = DESIGN_DOC.read_text(encoding="utf-8")
    start = document.find("## **Demonic Skill Examples**")
    assert start != -1, "the design document has no Demonic Skill Examples section"
    end = document.find("# **VI. Itemization**", start)
    assert end != -1, "could not find the end of the Demonic examples section"
    section = document[start:end]

    by_key = {(r["weapon"], r["slot"]): r["name"] for r in designed}
    checked = 0
    problems = []
    for line in section.splitlines():
        match = re.match(r"^\|\s*(\w+)\s*/\s*(\w+)\s*\|\s*([^|]+?)\s*\|",
                         line.strip())
        if not match:
            continue
        weapon, slot, name = match.groups()
        if weapon == "Weapon":  # the header row
            continue
        checked += 1
        actual = by_key.get((weapon, slot))
        if actual is None:
            problems.append(
                f"{weapon}/{slot} is in the document but not designed "
                f"in the workbook")
        elif actual != name:
            problems.append(
                f"{weapon}/{slot}: the document says {name!r}, the workbook "
                f"says {actual!r}")

    assert checked >= 5, (
        f"only parsed {checked} rows out of the Demonic examples table; "
        f"the table format probably changed and this test stopped checking")
    assert not problems, "\n".join(problems)

"""The affix sheets in the workbook must agree with the simulation's pool.

WHY THIS EXISTS. The affix pool now lives in two places: the Item Bases and
Affixes sheets of `docs/All_Things_Cataclysm.xlsx`, which is where it is edited
and where the Unreal DataTables are generated from, and
`sim/cataclysm_sim/affixes.py`, which is where the design rules are enforced and
where the tuning work happens. Two copies of the same numbers drift. This project
has been bitten by exactly that before: `sim/cataclysm_sim/scoring.py` is a copy
of a file in another repository and silently drifted twice, which is why
`sim/verify_scoring_port.py` exists.

WHICH IS AUTHORITATIVE. The workbook. It is what the project owner edits and
what the generated data tables come from. When this test fails, the usual fix is
to change the Python to match the sheet, not the other way round.

WHAT IT DOES NOT CHECK. Anything derived rather than stored: the seven-tier
curve, the roll band, the gear level multiplier and the two-handed multiplier are
formulas in `affixes.py` and appear in no sheet. Only the stored values are
compared.
"""

from __future__ import annotations

import pathlib

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"


def read_sheet(title: str) -> list[dict[str, object]]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if title not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {title!r}")

    rows = list(book[title].iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for raw in rows[1:]:
        if not raw or raw[0] is None or not str(raw[0]).strip():
            continue
        out.append({headers[i]: raw[i] for i in range(len(headers))
                    if i < len(raw)})
    return out


def text(value) -> str:
    return "" if value is None else str(value).strip()


def slot_set(value) -> frozenset[str]:
    return frozenset(s.strip() for s in text(value).split(",") if s.strip())


@pytest.fixture(scope="module")
def model():
    from cataclysm_sim import affixes as af
    return af


@pytest.fixture(scope="module")
def base_sheet():
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Item Bases")


@pytest.fixture(scope="module")
def affix_sheet():
    if not WORKBOOK.is_file():
        pytest.skip("design workbook not present")
    return read_sheet("Affixes")


class TestItemBases:
    def test_the_same_bases_exist_in_both(self, base_sheet, model):
        in_sheet = {text(r["Base Name"]) for r in base_sheet}
        in_model = {b.name for b in model.ITEM_BASES}
        assert in_sheet == in_model, (
            f"only in the sheet: {sorted(in_sheet - in_model)}; "
            f"only in affixes.py: {sorted(in_model - in_sheet)}")

    def test_every_base_is_in_the_same_slot(self, base_sheet, model):
        for row in base_sheet:
            base = model.base_named(text(row["Base Name"]))
            assert text(row["Slot"]) == base.slot, base.name

    def test_every_implicit_matches(self, base_sheet, model):
        for row in base_sheet:
            base = model.base_named(text(row["Base Name"]))
            stated = []
            for index in (1, 2):
                stat = text(row[f"Implicit {index} Stat"])
                if not stat:
                    continue
                stated.append((stat, text(row[f"Implicit {index} Kind"]),
                               float(row[f"Implicit {index} Value"])))
            # The sheet holds the value BEFORE the two-handed multiplier, which
            # is what `Implicit.value` holds too.
            modelled = [(i.stat, i.kind, i.value) for i in base.implicits]
            assert stated == modelled, base.name

    def test_weapon_bases_agree_on_hands_and_sub_type(self, base_sheet, model):
        for row in base_sheet:
            base = model.base_named(text(row["Base Name"]))
            if not isinstance(base, model.WeaponBase):
                assert not text(row["Hands"]), f"{base.name} is not a weapon"
                continue
            assert int(float(row["Hands"])) == base.hands, base.name
            assert text(row["Sub-Type"]) == base.sub_type, base.name
            assert text(row["Weapon Type"]) == base.weapon_type, base.name
            assert (int(float(row["Max Damage Types"]))
                    == base.max_damage_types_on_base), base.name
            # Checked for presence before conversion, because a blank cell would
            # otherwise fail as a TypeError about NoneType and say nothing about
            # which weapon is missing what.
            assert row["Attack Speed"] is not None, (
                f"{base.name} has no attack speed in the sheet")
            assert float(row["Attack Speed"]) == base.attack_speed, base.name

    def test_the_weapon_rates_still_average_to_what_the_two_handed_multiplier_assumed(
            self, base_sheet, model):
        """The two-handed multiplier of 2.0 is already shipped, and it was
        derived against a one-handed rate of 1.35 and a two-handed rate of 1.28.
        Per-weapon rates that do not average back to those move a multiplier
        nobody meant to move. See sim/analyse_two_handed_multiplier.py."""
        by_hands: dict[int, list[float]] = {1: [], 2: []}
        for row in base_sheet:
            base = model.base_named(text(row["Base Name"]))
            if isinstance(base, model.WeaponBase):
                by_hands[base.hands].append(base.attack_speed)

        assert by_hands[1] and by_hands[2], "no weapons found in the sheet"
        assert sum(by_hands[1]) / len(by_hands[1]) == pytest.approx(1.35)
        assert sum(by_hands[2]) / len(by_hands[2]) == pytest.approx(1.28)


class TestAffixes:
    def test_the_same_affixes_exist_in_both(self, affix_sheet, model):
        in_sheet = {text(r["Affix Name"]) for r in affix_sheet}
        in_model = (
            {a.name for a in model.AFFIX_POOL}
            | {f.name for f in model.RESISTANCE_FAMILIES}
            | {a.name for a in model.AILMENT_AFFIXES}
            | {h.name for h in model.HYBRID_AFFIXES})
        assert in_sheet == in_model, (
            f"only in the sheet: {sorted(in_sheet - in_model)}; "
            f"only in affixes.py: {sorted(in_model - in_sheet)}")

    def test_stat_affixes_match(self, affix_sheet, model):
        by_name = {a.name: a for a in model.AFFIX_POOL}
        seen = 0
        for row in affix_sheet:
            if text(row["Affix Kind"]) != "Stat":
                continue
            affix = by_name[text(row["Affix Name"])]
            assert text(row["Stat"]) == affix.stat, affix.name
            assert text(row["Value Kind"]) == affix.kind, affix.name
            assert float(row["Top Value"]) == pytest.approx(affix.top_value), affix.name
            assert text(row["Position"]) == affix.position, affix.name
            assert slot_set(row["Allowed Slots"]) == affix.allowed_slots, affix.name
            seen += 1
        assert seen == len(model.AFFIX_POOL)

    def test_resistance_families_match(self, affix_sheet, model):
        by_name = {f.name: f for f in model.RESISTANCE_FAMILIES}
        seen = 0
        for row in affix_sheet:
            if text(row["Affix Kind"]) != "Resistance":
                continue
            family = by_name[text(row["Affix Name"])]
            assert int(float(row["Breadth"])) == family.breadth, family.name
            assert float(row["Top Value"]) == pytest.approx(family.top_value), family.name
            assert text(row["Position"]) == family.position, family.name
            seen += 1
        assert seen == len(model.RESISTANCE_FAMILIES)

    def test_ailment_affixes_match(self, affix_sheet, model):
        by_name = {a.name: a for a in model.AILMENT_AFFIXES}
        seen = 0
        for row in affix_sheet:
            if text(row["Affix Kind"]) != "Ailment":
                continue
            affix = by_name[text(row["Affix Name"])]
            assert text(row["Ailment"]) == affix.ailment, affix.name
            assert text(row["Gem"]) == affix.gem, affix.name
            assert float(row["Top Value"]) == pytest.approx(affix.top_chance), affix.name
            assert slot_set(row["Allowed Slots"]) == affix.allowed_slots, affix.name
            seen += 1
        assert seen == len(model.AILMENT_AFFIXES)

    def test_hybrid_affixes_match(self, affix_sheet, model):
        by_name = {h.name: h for h in model.HYBRID_AFFIXES}
        seen = 0
        for row in affix_sheet:
            if text(row["Affix Kind"]) != "Hybrid":
                continue
            hybrid = by_name[text(row["Affix Name"])]
            stated = (text(row["Hybrid Part 1"]), text(row["Hybrid Part 2"]))
            assert stated == tuple(p.name for p in hybrid.parts), hybrid.name
            assert slot_set(row["Allowed Slots"]) == hybrid.allowed_slots, hybrid.name
            seen += 1
        assert seen == len(model.HYBRID_AFFIXES)


class TestTheCountsThatAreAssertedInUnreal:
    """The Unreal automation test pins these row counts by hand.

    No Python test can catch a stale number there, and it has been missed
    before, so the two counts are stated here as well and compared.
    """

    def test_the_item_base_count(self, base_sheet, model):
        assert len(base_sheet) == 55 == len(model.ITEM_BASES)

    def test_the_affix_count(self, affix_sheet, model):
        """68 after eight attribute affixes were added on 2026-08-04, then
        70 after mana leech and energy shield leech were added for #214, then
        78 after the eight increased-damage-against-a-type affixes for #213."""
        assert len(affix_sheet) == 78 == model.total_pool_size()

    def test_the_gem_count(self, model):
        """The Gems sheet count the Unreal test pins by hand.

        Not derived from the model: `affixes.py` holds no gem list. Stated here
        because a gem added to the workbook without the Unreal number being
        moved fails only in the C++ suite, which nothing on a pull request runs.
        """
        gem_rows = read_sheet("Gems")
        assert len(gem_rows) == 27

"""The Gear Rarity sheet must agree with the simulation's drop tables.

WHY THIS EXISTS. How heavily each rarity is weighted on a drop, and the upgrade
level each one requires, now live in two places: the Gear Rarity sheet of
`docs/All_Things_Cataclysm.xlsx`, which is what the game will read, and
`RARITY_DROP_WEIGHT` and `RARITY_GEAR_LEVEL_GATE` in `sim/cataclysm_sim/loot.py`,
where the drop rules are enforced and where the tuning work happens.

Two copies of the same numbers drift. This project has been bitten by exactly that
twice with `sim/cataclysm_sim/scoring.py`, which is why `sim/verify_scoring_port.py`
exists.

WHICH IS AUTHORITATIVE. The workbook, as everywhere else in this project. When
this fails, the usual fix is to change the Python to match the sheet.

WHAT WOULD GO UNNOTICED WITHOUT IT. Nothing errors when they disagree. A weight
column changed in the workbook to make Cataclysmic items rarer would leave the
simulation still reporting the old distribution, and every figure derived from it
would describe a different game from the one being played.
"""

from __future__ import annotations

import pathlib
import re
import sys

import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
DESIGN_DOCUMENT = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"
SHEET = "Gear Rarity"


def text(value) -> str:
    return "" if value is None else str(value).strip()


@pytest.fixture(scope="module")
def sheet() -> dict[str, dict]:
    import openpyxl

    book = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    if SHEET not in book.sheetnames:
        pytest.fail(f"the workbook has no sheet named {SHEET!r}")

    rows = list(book[SHEET].iter_rows(values_only=True))
    headers = [text(h) for h in rows[0]]
    out: dict[str, dict] = {}
    for raw in rows[1:]:
        if not raw or not text(raw[0]):
            continue
        record = {headers[i]: raw[i] for i in range(len(headers)) if i < len(raw)}
        out[text(record["Rarity"])] = record
    return out


@pytest.fixture(scope="module")
def model():
    if str(REPO_ROOT / "sim") not in sys.path:
        sys.path.insert(0, str(REPO_ROOT / "sim"))
    from cataclysm_sim import loot

    return loot


def test_the_sheet_was_found_and_has_every_rarity(sheet, model) -> None:
    """Read a real sheet, so the comparisons below cannot pass on an empty one."""
    import cataclysm_sim.affixes as af

    assert set(sheet) == set(af.RARITIES), (
        f"the {SHEET} sheet lists {sorted(sheet)}, and the rarity ladder is "
        f"{list(af.RARITIES)}"
    )


def test_every_drop_weight_matches(sheet, model) -> None:
    wrong = []
    for rarity, record in sheet.items():
        in_sheet = float(record["Drop Weight"])
        in_model = float(model.RARITY_DROP_WEIGHT[rarity])
        if abs(in_sheet - in_model) > 1e-9:
            wrong.append(f"{rarity}: sheet {in_sheet}, model {in_model}")

    assert not wrong, (
        "the Drop Weight column and RARITY_DROP_WEIGHT in "
        f"sim/cataclysm_sim/loot.py disagree: {'; '.join(wrong)}. The workbook "
        "is authoritative, so change the Python."
    )


def test_every_gear_level_gate_matches(sheet, model) -> None:
    wrong = []
    for rarity, record in sheet.items():
        in_sheet = int(record["Gear Level Gate"])
        in_model = int(model.RARITY_GEAR_LEVEL_GATE[rarity])
        if in_sheet != in_model:
            wrong.append(f"{rarity}: sheet {in_sheet}, model {in_model}")

    assert not wrong, (
        "the Gear Level Gate column and RARITY_GEAR_LEVEL_GATE in "
        f"sim/cataclysm_sim/loot.py disagree: {'; '.join(wrong)}. The workbook "
        "is authoritative, so change the Python."
    )


def test_every_residue_band_matches(sheet, model) -> None:
    wrong = []
    for rarity, record in sheet.items():
        in_sheet = (float(record["Residue On Drop Lowest"]),
                    float(record["Residue On Drop Highest"]))
        in_model = tuple(float(v) for v in model.RARITY_RESIDUE_BAND[rarity])
        if in_sheet != in_model:
            wrong.append(f"{rarity}: sheet {in_sheet}, model {in_model}")

    assert not wrong, (
        "the two Residue On Drop columns and RARITY_RESIDUE_BAND in "
        f"sim/cataclysm_sim/loot.py disagree: {'; '.join(wrong)}. The workbook "
        "is authoritative, so change the Python."
    )


def test_the_top_rarity_carries_the_band_the_project_owner_stated(sheet) -> None:
    """300 to 500 on a Cataclysmic drop, stated on 2026-08-18.

    Read from the sheet against the figure itself. Every other test here compares
    the sheet and the Python to each other, which passes happily if both were
    changed together by mistake.
    """
    record = sheet["Cataclysmic"]
    assert (float(record["Residue On Drop Lowest"]),
            float(record["Residue On Drop Highest"])) == (300.0, 500.0)


def test_every_band_runs_upward_and_starts_above_zero(sheet) -> None:
    bad = {}
    for rarity, record in sheet.items():
        lowest = float(record["Residue On Drop Lowest"])
        highest = float(record["Residue On Drop Highest"])
        if not 0.0 < lowest <= highest:
            bad[rarity] = (lowest, highest)
    assert not bad, (
        f"these rarities have a residue band that does not run upward from "
        f"above zero: {bad}"
    )


def test_the_gates_match_the_design_document(sheet) -> None:
    """The four stated gates, read from the sheet rather than from the model.

    `docs/Cataclysm_GDD_v2.md` section VI states these four in its rarity table:
    Legendary requires gear level 4, Mythical 6, Ascendant 8 and Cataclysmic 10.
    The other four rarities have no gate stated, so they are zero.

    THIS IS THE ONE PLACE THE NUMBERS ARE WRITTEN OUT AGAIN ON PURPOSE. Every
    other test here compares two copies to each other, which passes happily if
    both are wrong together. This one compares the sheet against the design
    document, which is the thing they are both supposed to be describing.
    """
    stated = {
        "Everyday": 0, "Quality": 0, "Superb": 0, "Masterful": 0,
        "Legendary": 4, "Mythical": 6, "Ascendant": 8, "Cataclysmic": 10,
    }
    wrong = []
    for rarity, gate in stated.items():
        in_sheet = int(sheet[rarity]["Gear Level Gate"])
        if in_sheet != gate:
            wrong.append(
                f"{rarity}: sheet {in_sheet}, design document {gate}")

    assert not wrong, (
        "the Gear Level Gate column disagrees with the rarity table in section "
        f"VI of docs/Cataclysm_GDD_v2.md: {'; '.join(wrong)}"
    )


#: Spelled-out counts the design document uses in the sentence below.
NUMBER_WORDS = {
    "one": 1, "two": 2, "three": 3, "four": 4, "five": 5,
    "six": 6, "seven": 7, "eight": 8, "nine": 9, "ten": 10,
}


def test_the_named_material_drop_rate_in_the_document_matches_the_model(model) -> None:
    """Section VI states how often a named top-tier material drops, and the
    model computes it. They have to agree.

    WHY THIS EXISTS. On 2026-08-23 issue #852 added ten upgrade stones and put
    two of them in the top rarity band, taking it from three materials to five.
    sim/cataclysm_sim/loot.py, sim/tests/test_loot.py and
    game/Source/Cataclysm/Tests/CataclysmDropRollTests.cpp were each updated to
    one drop in 1,705. The design document was missed and nothing noticed,
    because no test read that sentence. Issue #865.

    THE FIGURE IS PARSED OUT OF THE DOCUMENT rather than restated here. A test
    that writes 1,705 on both sides passes no matter what the document says,
    which is the hole this fills rather than repeats.
    """
    if not DESIGN_DOCUMENT.is_file():
        pytest.skip(f"{DESIGN_DOCUMENT.name} is not present")
    whole = DESIGN_DOCUMENT.read_text(encoding="utf-8")
    start = whole.index("# **VI. Itemization")
    section = whole[start:whole.index("# **VII.", start)]

    stated = re.search(
        r"since (\w+) materials share that tier, a named one such as "
        r"Purified Essence is one in ([\d,]+)", section)
    assert stated, (
        "section VI of docs/Cataclysm_GDD_v2.md no longer contains the sentence "
        "stating how often a named top-tier material drops, which begins 'since N "
        "materials share that tier'. If the wording changed deliberately, update "
        "this pattern rather than deleting the test.")

    document_count = NUMBER_WORDS[stated.group(1)]
    document_rate = int(stated.group(2).replace(",", ""))

    model_count = model.MATERIALS_IN_TIER["Extremely Rare"]
    model_rate = 1 / model.material_tier_distribution(0.0)["Extremely Rare"] * model_count

    assert document_count == model_count, (
        f"the design document says {document_count} materials share the top rarity "
        f"band; sim/cataclysm_sim/loot.py says {model_count}")
    assert abs(document_rate - model_rate) < 6, (
        f"the design document says a named top-tier material is one drop in "
        f"{document_rate:,}; the model computes one in {model_rate:,.0f}")

"""Tests for the gameplay tag generator.

Most of these run against a small fixture workbook built in the test, so they do
not depend on the real design data and cannot be broken by a design change. Two
tests at the end do check the real workbook, because the committed tag list being
out of date is exactly the drift this tool exists to prevent.
"""

from __future__ import annotations

import pathlib
import sys

import openpyxl
import pytest

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

import generate_gameplay_tags as gen  # noqa: E402


def make_workbook(path: pathlib.Path, tags, references=None) -> pathlib.Path:
    """Build a fixture workbook with a Tags sheet and optional reference sheets."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = gen.TAGS_SHEET
    sheet.append(["Tag Name", "Description"])
    for row in tags:
        sheet.append(list(row))

    if references:
        skills = book.create_sheet("Weapon Skills")
        skills.append(["Weapon Type", "Damage Type", "Slot", "Skill Name",
                       "Skill Description", "Tags"])
        for value in references:
            skills.append(["W", "War", "Heavy", "S", "D", value])

    book.save(path)
    return path


class TestReadingTags:
    def test_reads_tags_and_sorts_them(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [
            ("Zulu.Last", "last"), ("Alpha.First", "first"),
        ])
        assert gen.read_tags(book) == [("Alpha.First", "first"), ("Zulu.Last", "last")]

    def test_rejects_a_duplicate_tag(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [
            ("Element.War", "one"), ("Element.War", "two"),
        ])
        with pytest.raises(gen.TagError, match="defined more than once"):
            gen.read_tags(book)

    def test_rejects_a_tag_with_no_description(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [("Element.War", "")])
        with pytest.raises(gen.TagError, match="has no description"):
            gen.read_tags(book)

    @pytest.mark.parametrize("bad", ["Element War", "Element..War", ".Element", "9Element"])
    def test_rejects_an_invalid_tag_name(self, tmp_path, bad):
        book = make_workbook(tmp_path / "w.xlsx", [(bad, "d")])
        with pytest.raises(gen.TagError, match="not a valid tag name"):
            gen.read_tags(book)

    def test_rejects_a_workbook_with_no_tags_sheet(self, tmp_path):
        path = tmp_path / "empty.xlsx"
        openpyxl.Workbook().save(path)
        with pytest.raises(gen.TagError, match="no sheet named"):
            gen.read_tags(path)


def make_effect_workbook(path: pathlib.Path, buffs, debuffs, dots):
    """A fixture workbook whose three effect sheets name the given effects.

    The generator reads column A and takes everything before the first colon as
    the effect's name, which is how the real sheets are written: a row reads
    "Cripple: reduces movement speed" and the tag is built from "Cripple".
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = gen.TAGS_SHEET
    sheet.append(["Tag Name", "Description"])
    sheet.append(["Element.War", "War damage"])

    for name, rows in (("Buffs", buffs), ("Debuffs", debuffs), ("DoTs", dots)):
        made = book.create_sheet(name)
        for effect in rows:
            made.append([f"{effect}: what it does"])

    book.save(path)
    return path


class TestStatusEffectBranches:
    """The sheet an effect comes from is a segment of its tag. Issue #1145.

    WHY THIS MATTERS AND IS NOT COSMETIC. `Status.` used to be one flat branch
    holding the buffs, the debuffs and the damage over times together, so nothing
    could tell Divine Aegis from Cripple by its tag. Seven Masochist passive
    nodes are paid per debuff carried, and the only way to exclude the eighteen
    buffs was a list of names kept by hand. The project owner chose splitting the
    branch over keeping such a list, on 2026-09-04.
    """

    def test_each_sheet_gets_its_own_branch(self, tmp_path):
        found = dict(gen.read_status_effect_tags(make_effect_workbook(
            tmp_path / "effects.xlsx",
            buffs=["Divine Aegis"], debuffs=["Cripple"], dots=["Void Splinter"])))

        assert set(found) == {"Status.Buff.DivineAegis",
                              "Status.Debuff.Cripple",
                              "Status.DoT.VoidSplinter"}

    def test_a_buff_never_lands_under_the_debuff_branch(self, tmp_path):
        """The whole point, stated as its own assertion.

        `UCataclysmDebuffs::DebuffRootNames` names `Status.Debuff`, and a root
        counts itself and all its children, so a buff appearing under it would be
        counted as a debuff by every reader of that class at once.
        """
        found = dict(gen.read_status_effect_tags(make_effect_workbook(
            tmp_path / "effects.xlsx",
            buffs=["Commander", "Divine Aegis", "Demonic Rage"],
            debuffs=["Cripple"], dots=[])))

        under_debuff = [t for t in found if t.startswith("Status.Debuff.")]
        assert under_debuff == ["Status.Debuff.Cripple"], (
            f"a buff was generated under the debuff branch: {under_debuff}")

    def test_one_name_on_two_sheets_makes_one_tag_and_not_two(self, tmp_path):
        """First sheet wins, which is what it did before the split too.

        BEFORE THE SPLIT THIS FELL OUT OF THE TAG BEING THE KEY: both sheets
        produced `Status.Plague` and the second was dropped. Now they would
        produce `Status.Buff.Plague` and `Status.Debuff.Plague`, which are
        different keys, so the generator has to dedupe on the effect's NAME
        instead. Without that a design mistake becomes two tags for one effect
        and the second is applied by nothing.
        """
        found = dict(gen.read_status_effect_tags(make_effect_workbook(
            tmp_path / "effects.xlsx",
            buffs=["Plague"], debuffs=["Plague"], dots=[])))

        assert list(found) == ["Status.Buff.Plague"]

    def test_the_real_sheets_put_every_effect_under_a_kind(self):
        """No effect in the real design is left on the flat branch."""
        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")

        found = dict(gen.read_status_effect_tags(gen.WORKBOOK))
        assert len(found) > 40, "the effect sheets lost a large number of rows"

        kinds = set(gen.EFFECT_SHEETS.values())
        flat = [tag for tag in found
                if len(tag.split(".")) != 3 or tag.split(".")[1] not in kinds]
        assert not flat, (
            f"these status tags carry no kind segment, so nothing can tell "
            f"whether they harm: {sorted(flat)}")


class TestImpliedParents:
    def test_parents_are_implied(self):
        """Unreal creates parents automatically, so data may reference them."""
        implied = gen.implied_tags([("Item.Weapon.Sword", "d")])
        assert implied == {"Item", "Item.Weapon", "Item.Weapon.Sword"}

    def test_a_referenced_parent_is_not_reported_as_missing(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx",
                             [("Item.Weapon.Sword", "d")],
                             references=["Item.Weapon"])
        tags = gen.read_tags(book)
        assert gen.check_references(tags, gen.find_references(book)) == []

    def test_an_undefined_tag_is_reported(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx",
                             [("Slot.Ultimate", "d")],
                             references=["Type.Ultimate"])
        tags = gen.read_tags(book)
        problems = gen.check_references(tags, gen.find_references(book))
        assert len(problems) == 1
        assert "Type.Ultimate" in problems[0]
        assert "Weapon Skills" in problems[0]


class TestRendering:
    def test_output_is_a_valid_tag_list(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [("Element.War", "Physical, Bleed")])
        text = gen.render(gen.read_tags(book))
        assert "[/Script/GameplayTags.GameplayTagsList]" in text
        assert 'GameplayTagList=(Tag="Element.War",DevComment="Physical, Bleed")' in text

    def test_output_says_it_is_generated(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [("A.B", "d")])
        assert "GENERATED FILE" in gen.render(gen.read_tags(book))

    def test_output_is_deterministic(self, tmp_path):
        """Re-running on an unchanged sheet must produce no diff."""
        book = make_workbook(tmp_path / "w.xlsx", [("B.Two", "2"), ("A.One", "1")])
        tags = gen.read_tags(book)
        assert gen.render(tags) == gen.render(tags)

    def test_a_quote_in_a_description_is_refused(self, tmp_path):
        # The ini format cannot carry an unescaped quote. Rather than emit a
        # file Unreal will misparse, refuse to write it.
        with pytest.raises(gen.TagError, match="contains a quote"):
            gen.render([("A.B", 'a "quoted" word')])


class TestCommandLine:
    def test_writes_the_file(self, tmp_path, capsys):
        book = make_workbook(tmp_path / "w.xlsx", [("A.B", "d")])
        out = tmp_path / "sub" / "Tags.ini"
        assert gen.main(["--workbook", str(book), "--output", str(out)]) == 0
        assert out.is_file()
        assert "Wrote 1 tags" in capsys.readouterr().out

    def test_check_passes_when_current(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx", [("A.B", "d")])
        out = tmp_path / "Tags.ini"
        gen.main(["--workbook", str(book), "--output", str(out)])
        assert gen.main(["--workbook", str(book), "--output", str(out), "--check"]) == 0

    def test_check_fails_when_stale(self, tmp_path, capsys):
        book = make_workbook(tmp_path / "w.xlsx", [("A.B", "d")])
        out = tmp_path / "Tags.ini"
        gen.main(["--workbook", str(book), "--output", str(out)])

        # The sheet changes; the committed file does not.
        make_workbook(book, [("A.B", "d"), ("C.D", "e")])
        assert gen.main(["--workbook", str(book), "--output", str(out), "--check"]) == 1
        assert "out of date" in capsys.readouterr().err

    def test_strict_fails_on_an_undefined_reference(self, tmp_path, capsys):
        book = make_workbook(tmp_path / "w.xlsx",
                             [("Slot.Ultimate", "d")],
                             references=["Type.Ultimate"])
        out = tmp_path / "Tags.ini"
        # Without --strict this is only a warning, so the run still succeeds.
        assert gen.main(["--workbook", str(book), "--output", str(out)]) == 0
        assert "WARNING" in capsys.readouterr().err

        assert gen.main(["--workbook", str(book), "--output", str(out),
                         "--strict"]) == 1
        assert "FAIL" in capsys.readouterr().err

    def test_strict_passes_when_every_reference_resolves(self, tmp_path):
        book = make_workbook(tmp_path / "w.xlsx",
                             [("Slot.Ultimate", "d")],
                             references=["Slot.Ultimate"])
        out = tmp_path / "Tags.ini"
        assert gen.main(["--workbook", str(book), "--output", str(out),
                         "--strict"]) == 0

    def test_check_fails_when_the_file_is_missing(self, tmp_path, capsys):
        book = make_workbook(tmp_path / "w.xlsx", [("A.B", "d")])
        missing = tmp_path / "nope.ini"
        assert gen.main(["--workbook", str(book), "--output", str(missing), "--check"]) == 1
        assert "does not exist" in capsys.readouterr().err


class TestAgainstTheRealWorkbook:
    """The committed tag list going stale is the drift this tool prevents."""

    def test_the_committed_tag_list_is_current(self, capsys):
        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")
        assert gen.main(["--check"]) == 0, (
            "game/Config/Tags/CataclysmTags.ini is out of date. "
            "Run: python tools/generate_gameplay_tags.py"
        )

    def test_the_real_tag_vocabulary_is_valid(self):
        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")
        tags = gen.read_tags(gen.WORKBOOK)
        assert len(tags) > 100, "the Tags sheet lost a large number of rows"
        roots = {tag.split(".")[0] for tag, _ in tags}
        assert {"Element", "Type", "Slot", "Item", "Stat",
                "Keyword", "Scope", "Trigger"} <= roots

    def test_item_slot_tags_are_exactly_the_design_document_s_slots(self):
        """Issue #106. The tag list had 14 item slots; the design has 11 plus
        potions.

        It carried an Item.Slot.OffHand for a slot the design explicitly removes
        -- "There are no offhand items" -- an Item.Slot.Bracers for a piece that
        appears nowhere, and Feet and Neck where the design says Boots and
        Necklace. Nothing compared the two, so the disagreement sat there.

        GEAR_SLOTS is the simulation's reading of the Item Slots list in section
        VI, so comparing against it means the vocabulary cannot drift from the
        design without one of the two being changed deliberately. Potion is
        added here because the design lists four potion slots but they are
        consumables rather than gear, which is why the simulation's gear list
        leaves them out.
        """
        from cataclysm_sim.affixes import GEAR_SLOTS

        if not gen.WORKBOOK.is_file():
            pytest.skip("design workbook not present")

        declared = {tag for tag, _ in gen.read_tags(gen.WORKBOOK)
                    if tag.startswith("Item.Slot.")}
        expected = {f"Item.Slot.{slot}" for slot in GEAR_SLOTS} | {"Item.Slot.Potion"}

        assert declared == expected, (
            f"the Tags sheet and the design's item slots disagree.\n"
            f"  only in the Tags sheet: {sorted(declared - expected)}\n"
            f"  only in the design:     {sorted(expected - declared)}"
        )

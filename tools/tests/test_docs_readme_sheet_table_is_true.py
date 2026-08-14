"""The sheet table in docs/README.md, checked against the workbook it describes.

WHY THIS EXISTS. Issue #334. `docs/README.md` carried a table headed "Sheets in
`All_Things_Cataclysm.xlsx`" that was stale in two independent ways at once: it
listed 11 of the workbook's sheets, and **every** row count in it was wrong. The
counts were the padded ranges from the original Google Sheets export -- 1012,
1007, 961, 1000, 1005 -- rather than the number of rows holding data.

WHY THAT MATTERED RATHER THAN BEING UNTIDY. The Weapon Skills row said 1000 where
the sheet holds 398, and 398 is the figure `docs/Cataclysm_GDD_v2.md` itself uses
for the size of the skill matrix. So the two design documents gave different
sizes for the same table, and a reader arriving at the README first got the wrong
one.

THE ISSUE'S OWN FIGURES WERE WRONG FOR THREE SHEETS, and this file is why that is
worth recording rather than quietly correcting. #334 measured Buffs at 17,
Debuffs at 23 and DoTs at 7. **Those three sheets have no heading row**, so a
script that skips the first row of every sheet loses one entry from each. The
generated `game/Data/StatusEffects.csv` settles it: 18 buffs, 24 debuffs, 8
damage-over-time effects, 50 rows in total.
`test_the_headerless_sheets_agree_with_the_generated_table` below checks exactly
that, so the constant naming those three sheets cannot rot into a wrong number
without something failing.

WHAT IS ASSERTED HERE.

    docs/README.md still has a sheet table this parser can read
    it names every sheet in the workbook, and no sheet the workbook lacks
    every row count in it matches the workbook
    the count for the three heading-less sheets agrees with the generated
      status effect table, which is derived from them
    the prose count of sheets in the file listing above matches as well
"""

from __future__ import annotations

import csv
import pathlib
import re

import openpyxl
import pytest

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
README = REPO_ROOT / "docs" / "README.md"
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
STATUS_EFFECTS = REPO_ROOT / "game" / "Data" / "StatusEffects.csv"

#: The heading the table sits under.
TABLE_HEADING = "## Sheets in `All_Things_Cataclysm.xlsx`"

#: Sheets with no heading row, so their first row is data rather than a label.
#: `tools/generate_datatables.py` says the same thing in its own words: "Buffs,
#: Debuffs -- single column, no header row, each cell 'Name: Description'".
#: Getting this set wrong is what made issue #334's own measurements one short on
#: each of them, so the test below cross-checks it against generated data rather
#: than leaving it as a claim.
HEADERLESS = {"Buffs", "Debuffs", "DoTs"}


def data_rows(sheet_name: str, sheet) -> int:
    """Rows holding data, excluding a heading row where the sheet has one."""
    rows = [row for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)]
    return len(rows) if sheet_name in HEADERLESS else len(rows) - 1


@pytest.fixture(scope="module")
def workbook_counts() -> dict[str, int]:
    if not WORKBOOK.is_file():
        pytest.skip("the design workbook is not present")
    book = openpyxl.load_workbook(WORKBOOK, read_only=True)
    try:
        return {name: data_rows(name, book[name]) for name in book.sheetnames}
    finally:
        book.close()


@pytest.fixture(scope="module")
def readme() -> str:
    if not README.is_file():
        pytest.skip("docs/README.md is not present")
    return README.read_text(encoding="utf-8")


@pytest.fixture(scope="module")
def table_counts(readme: str) -> dict[str, int]:
    """Sheet name against stated row count, read out of the Markdown table.

    The names and numbers are wrapped in `**` on the Tags row, because that row
    is emphasised, so the markers are stripped rather than assumed absent.
    """
    start = readme.find(TABLE_HEADING)
    assert start != -1, (
        f"docs/README.md has no {TABLE_HEADING!r} section. If it was renamed, "
        f"update TABLE_HEADING here; if it was removed, nothing records what "
        f"the workbook contains. Issue #334.")
    end = readme.find("\n## ", start + len(TABLE_HEADING))
    section = readme[start:end if end != -1 else len(readme)]

    found: dict[str, int] = {}
    for line in section.splitlines():
        cells = [cell.strip().strip("*") for cell in line.split("|")]
        # A data row is `| Sheet | Rows | First columns |`, so splitting gives
        # an empty cell at each end.
        if len(cells) < 4 or not cells[1] or not re.fullmatch(r"\d+", cells[2]):
            continue
        found[cells[1]] = int(cells[2])

    assert found, (
        "no rows could be parsed out of the sheet table in docs/README.md. "
        "Either the table's shape changed, in which case fix this parser, or "
        "the table is gone. Every assertion below would otherwise pass having "
        "compared nothing. Issue #334.")
    return found


def test_the_table_names_every_sheet_and_no_others(table_counts,
                                                   workbook_counts) -> None:
    """The half of issue #334 that was invisible: five sheets read by the data
    pipeline were missing from the table entirely, so a reader could not tell
    they existed."""
    missing = sorted(set(workbook_counts) - set(table_counts))
    assert not missing, (
        f"the sheet table in docs/README.md does not list {missing}. Every "
        f"sheet in docs/All_Things_Cataclysm.xlsx has to appear, or a reader "
        f"cannot tell the table is incomplete. Issue #334.")

    invented = sorted(set(table_counts) - set(workbook_counts))
    assert not invented, (
        f"the sheet table in docs/README.md lists {invented}, which the "
        f"workbook does not have. Either the sheet was removed and the table "
        f"was not, or the name is misspelt. Issue #334.")


def test_every_row_count_in_the_table_is_true(table_counts,
                                              workbook_counts) -> None:
    """The other half. Every count in the table was the padded export range
    rather than the rows holding data, and the largest gap was Weapon Skills at
    1000 against a real 398."""
    wrong = {name: (stated, workbook_counts[name])
             for name, stated in table_counts.items()
             if name in workbook_counts and stated != workbook_counts[name]}
    assert not wrong, (
        "the sheet table in docs/README.md states row counts the workbook does "
        "not have. Sheet: stated, actual -- "
        + "; ".join(f"{n}: {s}, {a}" for n, (s, a) in sorted(wrong.items()))
        + ". Issue #334.")


def test_the_headerless_sheets_agree_with_the_generated_table(
        workbook_counts) -> None:
    """HEADERLESS is a claim about the workbook, and a wrong claim there makes
    every count above wrong by one without failing anything.

    `game/Data/StatusEffects.csv` is generated from these three sheets, one row
    per entry, so it is an independent count of the same thing. Issue #334's own
    measurements were one short on each of these three, which is what this
    catches.
    """
    if not STATUS_EFFECTS.is_file():
        pytest.skip("game/Data/StatusEffects.csv has not been generated")

    with STATUS_EFFECTS.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    kind_column = "Kind" if rows and "Kind" in rows[0] else "EffectKind"
    generated = {"Buffs": sum(1 for r in rows if r[kind_column] == "Buff"),
                 "Debuffs": sum(1 for r in rows if r[kind_column] == "Debuff"),
                 "DoTs": sum(1 for r in rows if r[kind_column] == "DoT")}

    for sheet, count in generated.items():
        assert workbook_counts[sheet] == count, (
            f"the {sheet} sheet is measured at {workbook_counts[sheet]} rows, "
            f"and game/Data/StatusEffects.csv holds {count} entries generated "
            f"from it. They have to agree. If the sheet gained a heading row, "
            f"remove it from HEADERLESS in this file; if it lost one, add it. "
            f"Issue #334.")


def test_the_file_listing_states_the_right_number_of_sheets(
        readme, workbook_counts) -> None:
    """The count appears twice in this file: once as a table and once as prose
    in the row describing the workbook. The prose one said 11 for as long as the
    table did, so fixing only the table would leave the same wrong number in the
    place a reader meets first."""
    expected = f"{len(workbook_counts)} sheets"
    assert expected in readme, (
        f"docs/README.md does not say the workbook has {expected}. The file "
        f"listing near the top states a sheet count in prose, and it went "
        f"stale alongside the table below it. Issue #334.")

"""The skill example tables in the design document quote numbers. Check them.

WHY THIS FILE EXISTS. `docs/Cataclysm_GDD_v2.md` has two summary tables, "War
Skill Examples" and "Demonic Skill Examples", each abbreviating skills whose full
text lives in the Weapon Skills sheet of `docs/All_Things_Cataclysm.xlsx`. The
sheet is authoritative: `game/Data/WeaponSkills.csv` is generated from it and the
game reads that. The summaries had drifted (issue #154):

- Ballista fired "every 2.5 seconds" in the document and every 2 seconds in the
  sheet.
- Fortress reflected "100% of blocked damage" in the document and 500% in the
  sheet. Anyone reading the document to implement it would have built the skill
  five times too weak.

WHAT IS CHECKED, AND WHY NOT THE WHOLE TEXT. The summaries are deliberately
shorter than the sheet: they drop facts, reorder clauses and reword freely.
Comparing the prose would fail on formatting rather than on drift, which is
exactly why `tools/tests/test_demonic_skills.py` compares skill NAMES only. So
this compares the one thing a summary must not change: **every quantity it
quotes has to appear in the full description**.

A summary may say less than the sheet. It may not say something different.

Percentages, seconds and metres are the three units the tables use. A quantity in
any other unit -- "200 HP" -- is ignored, because a summary that mentions it at
all would have to spell out the unit and the risk of drift is lower than the risk
of a false failure.
"""

from __future__ import annotations

import pathlib
import re

import pytest
from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
DESIGN_DOC = REPO_ROOT / "docs" / "Cataclysm_GDD_v2.md"
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"

#: The two summary tables, by the heading each sits under.
EXAMPLE_TABLES = ("War Skill Examples", "Demonic Skill Examples")

#: How a unit may be written, and the name this file uses for it.
UNITS = {
    "percent": r"%",
    "second": r"s(?:ec|econd)?s?\b",
    "metre": r"m(?:et(?:er|re)s?)?\b",
}

QUANTITY = re.compile(
    r"(\d+(?:\.\d+)?)\s*(" + "|".join(UNITS.values()) + ")",
    re.IGNORECASE,
)


def quantities(text: str) -> set[tuple[float, str]]:
    """Every number-and-unit pair in a piece of text."""
    found: set[tuple[float, str]] = set()
    for amount, written in QUANTITY.findall(text or ""):
        for name, pattern in UNITS.items():
            if re.fullmatch(pattern, written, re.IGNORECASE):
                found.add((float(amount), name))
                break
    return found


def workbook_descriptions() -> dict[str, str]:
    """Every designed skill's full description, by skill name."""
    sheet = load_workbook(WORKBOOK, read_only=True, data_only=True)["Weapon Skills"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(cell or "") for cell in next(rows)]
    name_column = header.index("Skill Name")
    text_column = header.index("Skill Description")

    described: dict[str, str] = {}
    for row in rows:
        if not row or not row[name_column]:
            continue
        described[str(row[name_column]).strip()] = str(row[text_column] or "")
    return described


def summary_rows(heading: str) -> list[tuple[str, str]]:
    """The (skill name, summary) pairs under one heading's table."""
    text = DESIGN_DOC.read_text(encoding="utf-8")
    start = text.find(f"## **{heading}**")
    if start == -1:
        pytest.fail(f"{DESIGN_DOC.name} has no heading '{heading}'.")

    # To the next heading of the same level, or the end of the document.
    end = text.find("\n## ", start + 1)
    table = text[start : end if end != -1 else len(text)]

    rows = []
    for line in table.splitlines():
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        # Three columns: weapon and slot, skill name, description. The header row
        # and the alignment row have no description worth reading.
        if len(cells) != 3 or not cells[1] or cells[1].startswith((":", "\\*\\*")):
            continue
        rows.append((cells[1], cells[2]))
    return rows


@pytest.mark.parametrize("heading", EXAMPLE_TABLES)
def test_the_table_is_found_and_is_not_empty(heading: str) -> None:
    """Guards every check below. An empty table would pass them all."""
    rows = summary_rows(heading)
    assert len(rows) >= 5, (
        f"Found only {len(rows)} rows under '{heading}' in {DESIGN_DOC.name}. "
        "The table was probably reformatted and this file can no longer read it."
    )


@pytest.mark.parametrize("heading", EXAMPLE_TABLES)
def test_every_summarised_skill_exists_in_the_workbook(heading: str) -> None:
    described = workbook_descriptions()
    missing = [name for name, _ in summary_rows(heading) if name not in described]
    assert not missing, (
        f"'{heading}' in {DESIGN_DOC.name} summarises skills that the Weapon "
        f"Skills sheet of {WORKBOOK.name} does not contain: {', '.join(missing)}"
    )


@pytest.mark.parametrize("heading", EXAMPLE_TABLES)
def test_every_quantity_in_the_summary_appears_in_the_full_description(
    heading: str,
) -> None:
    """A summary may say less than the sheet. It may not say something different."""
    described = workbook_descriptions()

    problems = []
    for name, summary in summary_rows(heading):
        full = described.get(name)
        if full is None:
            continue  # Covered by the test above.

        invented = quantities(summary) - quantities(full)
        for amount, unit in sorted(invented):
            problems.append(
                f"{name}: the summary says {amount:g} {unit}, which does not "
                f"appear in its description in {WORKBOOK.name}"
            )

    assert not problems, (
        f"'{heading}' in {DESIGN_DOC.name} has drifted from the workbook:\n  "
        + "\n  ".join(problems)
    )

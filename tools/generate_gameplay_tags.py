"""Generate Unreal's gameplay tag list from the Tags sheet of the design workbook.

    python tools/generate_gameplay_tags.py            # write the tag list
    python tools/generate_gameplay_tags.py --check    # verify, change nothing

The Tags sheet is the vocabulary the ability, enchantment and item systems are all
written in. It is generated rather than typed into the editor so the two cannot
drift: if an enchantment references a tag that no longer exists, it silently stops
matching items, and nothing reports it.

Output is sorted and deterministic, so re-running on an unchanged sheet produces
no diff.
"""

from __future__ import annotations

import argparse
import pathlib
import re
import sys

REPO_ROOT = pathlib.Path(__file__).resolve().parents[1]
WORKBOOK = REPO_ROOT / "docs" / "All_Things_Cataclysm.xlsx"
OUTPUT = REPO_ROOT / "game" / "Config" / "Tags" / "CataclysmTags.ini"

TAGS_SHEET = "Tags"

# Sheets and zero-based column indexes that contain tag references, checked
# against the vocabulary so the data cannot reference a tag that does not exist.
# Zero-based column indexes. The Enchantments sheet holds two tables side by
# side: positives in columns A-D with tags in D (index 3), negatives in F-I with
# tags in I (index 8). Index 7 is the negatives WEIGHT column and pointing at it
# silently skipped every negative enchantment's tags.
REFERENCE_COLUMNS = {
    "Weapon Skills": [5],
    "Enchantments": [3, 8],
}

# Unreal allows letters, digits and underscores in each segment, separated by dots.
VALID_TAG = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*(\.[A-Za-z0-9_]+)*$")

# Matches anything shaped like a dotted tag, used to find references in free text.
TAG_SHAPED = re.compile(r"\b[A-Z][A-Za-z]*(?:\.[A-Za-z][A-Za-z0-9]*)+\b")

HEADER = """; Gameplay tags for Cataclysm.
;
; GENERATED FILE -- do not edit by hand.
; Source: docs/All_Things_Cataclysm.xlsx, "Tags" sheet.
; Regenerate: python tools/generate_gameplay_tags.py
;
; Unreal loads every .ini in Config/Tags when ImportTagsFromConfig is true; see
; Config/DefaultGameplayTags.ini.

[/Script/GameplayTags.GameplayTagsList]
"""


class TagError(Exception):
    """A problem with the tag vocabulary that must stop the build."""


def read_tags(workbook: pathlib.Path) -> list[tuple[str, str]]:
    """Return (tag, description) pairs from the Tags sheet, sorted by tag."""
    import openpyxl

    book = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    if TAGS_SHEET not in book.sheetnames:
        raise TagError(f"{workbook} has no sheet named {TAGS_SHEET!r}")

    rows = [r for r in book[TAGS_SHEET].iter_rows(values_only=True) if r and r[0]]
    if not rows:
        raise TagError(f"the {TAGS_SHEET} sheet is empty")

    # Drop the header row if it looks like one.
    if str(rows[0][0]).strip().lower() in {"tag name", "tag", "name"}:
        rows = rows[1:]

    tags: dict[str, str] = {}
    problems: list[str] = []
    for index, row in enumerate(rows, start=2):
        tag = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not VALID_TAG.match(tag):
            problems.append(f"row {index}: {tag!r} is not a valid tag name")
            continue
        if tag in tags:
            problems.append(f"row {index}: {tag!r} is defined more than once")
            continue
        if not description:
            problems.append(f"row {index}: {tag!r} has no description")
            continue
        tags[tag] = description

    if problems:
        raise TagError("the Tags sheet has problems:\n  " + "\n  ".join(problems))

    return sorted(tags.items())


def implied_tags(tags: list[tuple[str, str]]) -> set[str]:
    """Every tag plus the parents Unreal creates implicitly.

    Declaring `Item.Weapon.Sword` gives you `Item.Weapon` and `Item` for free, so
    data referencing a parent is legitimate even when the sheet never lists it.
    """
    known: set[str] = set()
    for tag, _ in tags:
        parts = tag.split(".")
        for i in range(1, len(parts) + 1):
            known.add(".".join(parts[:i]))
    return known


def find_references(workbook: pathlib.Path) -> dict[str, set[str]]:
    """Return {referenced tag: set of sheets it appears in}."""
    import openpyxl

    book = openpyxl.load_workbook(workbook, data_only=True, read_only=True)
    found: dict[str, set[str]] = {}
    for sheet, columns in REFERENCE_COLUMNS.items():
        if sheet not in book.sheetnames:
            continue
        for row in book[sheet].iter_rows(values_only=True):
            for column in columns:
                if column < len(row) and row[column]:
                    for match in TAG_SHAPED.findall(str(row[column])):
                        found.setdefault(match, set()).add(sheet)
    return found


def check_references(tags: list[tuple[str, str]],
                     references: dict[str, set[str]]) -> list[str]:
    """Report references to tags the vocabulary does not define."""
    known = implied_tags(tags)
    return [f"{tag}  (used in {', '.join(sorted(sheets))})"
            for tag, sheets in sorted(references.items())
            if tag not in known]


def render(tags: list[tuple[str, str]]) -> str:
    lines = [HEADER]
    for tag, description in tags:
        # Descriptions are checked for quotes on the way in, so no escaping is
        # needed. If that ever changes this is where it breaks.
        if '"' in description:
            raise TagError(f"{tag}: description contains a quote, which the "
                           f"ini format cannot carry unescaped")
        lines.append(f'GameplayTagList=(Tag="{tag}",DevComment="{description}")\n')
    return "".join(lines)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true",
                        help="verify without writing; exit 1 if out of date")
    parser.add_argument("--strict", action="store_true",
                        help="treat undefined tag references as an error, not a "
                             "warning")
    parser.add_argument("--workbook", type=pathlib.Path, default=WORKBOOK)
    parser.add_argument("--output", type=pathlib.Path, default=OUTPUT)
    args = parser.parse_args(argv)

    try:
        tags = read_tags(args.workbook)
    except TagError as error:
        print(f"FAIL: {error}", file=sys.stderr)
        return 1

    contents = render(tags)

    # Every tag reference in the design data currently resolves. --strict keeps it
    # that way; continuous integration uses it. Without the flag this is only a
    # warning, which is the right behaviour while a sheet is mid-edit locally.
    unknown = check_references(tags, find_references(args.workbook))
    if unknown:
        label = "FAIL" if args.strict else "WARNING"
        print(f"{label}: {len(unknown)} tag reference(s) are not defined in the "
              f"Tags sheet:", file=sys.stderr)
        for line in unknown:
            print(f"  {line}", file=sys.stderr)
        if args.strict:
            print("  Either add the tag to the Tags sheet or correct the "
                  "reference.", file=sys.stderr)
            return 1

    if args.check:
        if not args.output.is_file():
            print(f"FAIL: {args.output} does not exist. Run without --check.",
                  file=sys.stderr)
            return 1
        if args.output.read_text(encoding="utf-8") != contents:
            print(f"FAIL: {args.output} is out of date with {args.workbook.name}. "
                  f"Run without --check.", file=sys.stderr)
            return 1
        print(f"{args.output.name} is up to date ({len(tags)} tags).")
        return 0

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(contents, encoding="utf-8")

    # Show a repository-relative path when the output is inside the repository,
    # and the full path otherwise. relative_to raises rather than returning the
    # absolute path when there is no common root.
    try:
        shown = args.output.resolve().relative_to(REPO_ROOT)
    except ValueError:
        shown = args.output
    print(f"Wrote {len(tags)} tags to {shown}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
